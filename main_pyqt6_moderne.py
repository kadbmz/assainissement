import sys
import polars as pl
import os
import threading
import time
from pathlib import Path
import chardet

# Gestion des ressources pour l'exécutable
def resource_path(relative_path):
    """Obtient le chemin absolu des ressources, fonctionne en développement et dans l'exécutable PyInstaller"""
    try:
        # PyInstaller crée un dossier temporaire et stocke le chemin dans _MEIPASS
        base_path = sys._MEIPASS
        print(f"🔍 Détection exécutable PyInstaller - base_path: {base_path}")
    except AttributeError:
        # En développement normal, utiliser le répertoire du script
        base_path = os.path.dirname(os.path.abspath(__file__))
        print(f"🔍 Mode développement - base_path: {base_path}")
    except Exception as e:
        print(f"❌ Erreur détection base_path: {e}")
        # Fallback : utiliser le répertoire du script
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    full_path = os.path.join(base_path, relative_path)
    print(f"📂 Chemin ressource recherché: {full_path}")
    print(f"📂 Fichier existe: {os.path.exists(full_path)}")
    
    return full_path
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QWidget, 
    QPushButton, QLabel, QTextEdit, QProgressBar, QFileDialog, 
    QMessageBox, QFrame, QScrollArea, QGridLayout, QSplitter,
    QGroupBox, QStatusBar, QMenuBar, QToolBar, QSizePolicy,
    QInputDialog, QComboBox, QDialog, QListWidget, QListWidgetItem,
    QTabWidget, QCheckBox, QLineEdit, QRadioButton, QButtonGroup
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer, QSize, QRect, QUrl
from PyQt6.QtGui import QFont, QIcon, QPalette, QColor, QPixmap, QPainter, QLinearGradient, QBrush, QPen
from PyQt6.QtMultimedia import QMediaPlayer, QAudioOutput
from PyQt6.QtMultimediaWidgets import QVideoWidget

# Importer la fenêtre de login
from login_window import LoginWindow

# Import conditionnel de cv2 avec fallback
try:
    import cv2
    CV2_AVAILABLE = True
except ImportError:
    CV2_AVAILABLE = False
    print("Attention: cv2 (opencv-python) n'est pas installé. La vidéo ne sera pas disponible.")

class WorkerThread(QThread):
    """Thread pour les opérations longues"""
    progress_updated = pyqtSignal(int)
    log_message = pyqtSignal(str, str)  # message, level
    finished_loading = pyqtSignal(object)  # df_consolidated
    finished_processing = pyqtSignal(object)  # df_traite
    finished_lettrage = pyqtSignal(object)  # df_traite_lettré
    
    def __init__(self, operation, *args):
        super().__init__()
        self.operation = operation
        self.args = args
        self.df_consolidated = None
        self.df_traite = None
    
    def run(self):
        if self.operation == "load":
            self._load_files()
        elif self.operation == "process":
            self._process_data()
        elif self.operation == "lettrage":
            self._lettrage_data()
        elif self.operation == "lettrage_filtre":
            self._lettrage_data_filtre()
    
    def _load_files(self):
        """Logique de chargement des fichiers"""
        try:
            dossier = self.args[0]
            self.log_message.emit(f"Début du chargement depuis: {dossier}", "INFO")
            
            chemin_dossier = Path(dossier)
            fichiers_txt = list(chemin_dossier.glob("*.txt"))
            fichiers_dsv = list(chemin_dossier.glob("*.dsv"))
            tous_fichiers = fichiers_txt + fichiers_dsv
            
            if not tous_fichiers:
                self.log_message.emit("Aucun fichier .txt ou .dsv trouvé!", "ERROR")
                return
            
            fichiers_cpi = [str(f) for f in tous_fichiers]
            nb_fichiers = len(fichiers_cpi)
            
            nb_txt = len(fichiers_txt)
            nb_dsv = len(fichiers_dsv)
            
            if nb_txt > 0 and nb_dsv > 0:
                self.log_message.emit(f"Trouvé {nb_fichiers} fichiers: {nb_txt} .txt et {nb_dsv} .dsv", "SUCCESS")
            elif nb_txt > 0:
                self.log_message.emit(f"Trouvé {nb_fichiers} fichiers .txt", "SUCCESS")
            else:
                self.log_message.emit(f"Trouvé {nb_fichiers} fichiers .dsv", "SUCCESS")
            
            # Calculer la taille totale
            taille_totale = sum(os.path.getsize(f) for f in fichiers_cpi)
            taille_mb = taille_totale / (1024 * 1024)
            
            self.log_message.emit(f"Taille totale: {taille_mb:.1f} MB", "INFO")
            
            start_time = time.time()
            separateur = "|"
            
            # Charger les fichiers
            lazy_dfs = []
            fichiers_echoues = []
            
            for i, fichier in enumerate(fichiers_cpi):
                try:
                    nom_fichier = Path(fichier).name
                    
                    lazy_df = pl.scan_csv(
                        fichier,
                        separator=separateur,
                        has_header=True,
                        infer_schema_length=1000,
                        ignore_errors=True,
                        encoding='utf8-lossy',
                        try_parse_dates=False,
                        null_values=['', 'NULL', 'null', 'N/A', 'n/a'],
                        truncate_ragged_lines=True,
                        raise_if_empty=False
                    )
                    
                    lazy_df = lazy_df.with_columns(
                        pl.lit(nom_fichier).alias("source_file")
                    )
                    
                    lazy_dfs.append(lazy_df)
                    
                    progression = int((i + 1) / nb_fichiers * 100)
                    self.progress_updated.emit(progression)
                    
                    if (i + 1) % 10 == 0:
                        self.log_message.emit(f"Fichiers traités: {i + 1}/{nb_fichiers}", "INFO")
                
                except Exception as err:
                    fichiers_echoues.append(fichier)
                    error_msg = f"Erreur fichier {Path(fichier).name}: {str(err)[:50]}..."
                    self.log_message.emit(error_msg, "ERROR")
                    continue
            
            if not lazy_dfs:
                self.log_message.emit("Aucun fichier n'a pu être chargé!", "ERROR")
                return
            
            # Concaténer tous les lazy frames
            self.log_message.emit("Consolidation des données...", "INFO")
            
            try:
                if len(lazy_dfs) == 1:
                    self.df_consolidated = lazy_dfs[0].collect()
                else:
                    self.df_consolidated = pl.concat(lazy_dfs, how="diagonal_relaxed").collect()
                
                end_time = time.time()
                temps_chargement = end_time - start_time
                
                nb_lignes = self.df_consolidated.height
                nb_colonnes_reelles = self.df_consolidated.width
                
                self.log_message.emit(f"Chargement terminé en {temps_chargement:.2f} secondes", "SUCCESS")
                self.log_message.emit(f"DataFrame consolidé: {nb_lignes:,} lignes × {nb_colonnes_reelles} colonnes", "SUCCESS")
                
                colonnes_detectees = self.df_consolidated.columns
                self.log_message.emit(f"Colonnes détectées: {', '.join(colonnes_detectees[:10])}{'...' if len(colonnes_detectees) > 10 else ''}", "INFO")
                
                if fichiers_echoues:
                    self.log_message.emit(f"{len(fichiers_echoues)} fichier(s) n'ont pas pu être chargés", "WARNING")
                
                self.finished_loading.emit(self.df_consolidated)
                
            except Exception as err:
                self.log_message.emit(f"Erreur lors de la consolidation: {str(err)}", "ERROR")
        
        except Exception as err:
            self.log_message.emit(f"Erreur générale: {str(err)}", "ERROR")
    
    def _process_data(self):
        """Logique de traitement des données"""
        try:
            df_consolidated = self.args[0]
            self.log_message.emit("Début du traitement des données...", "INFO")
            start_time = time.time()
            
            # Créer une copie pour le traitement
            self.df_traite = df_consolidated.clone()
            
            # Nettoyage des données
            self.log_message.emit("Nettoyage des données...", "INFO")
            self.df_traite = self.df_traite.filter(
                pl.any_horizontal(pl.col("*").is_not_null())
            )
            
            # Ajouter la colonne INST PAIEMENT
            self.log_message.emit("Ajout de la colonne INST PAIEMENT...", "INFO")
            
            conditions = {
                'Chèque': [30, 130, 31, 131, 32, 132],
                'Effet commercial': [60, 160, 61, 161],
                'Virement': [10, 110, 11, 111, 12, 112],
                'Prélèvement liaison': [20, 120],
                'Monétique': [40, 140, 141, 50, 150, 51, 151, 70, 170]
            }
            
            if "TYPEOPERATION" in self.df_traite.columns:
                self.df_traite = self.df_traite.with_columns(
                    pl.col("TYPEOPERATION").cast(pl.Int64).alias("TYPEOPERATION_int")
                ).with_columns(
                    pl.when(pl.col("TYPEOPERATION_int").is_in(conditions['Chèque']))
                    .then(pl.lit('Chèque'))
                    .when(pl.col("TYPEOPERATION_int").is_in(conditions['Effet commercial']))
                    .then(pl.lit('Effet commercial'))
                    .when(pl.col("TYPEOPERATION_int").is_in(conditions['Virement']))
                    .then(pl.lit('Virement'))
                    .when(pl.col("TYPEOPERATION_int").is_in(conditions['Prélèvement liaison']))
                    .then(pl.lit('Prélèvement liaison'))
                    .when(pl.col("TYPEOPERATION_int").is_in(conditions['Monétique']))
                    .then(pl.lit('Monétique'))
                    .otherwise(pl.lit('Non classé'))
                    .alias("INST PAIEMENT")
                ).drop("TYPEOPERATION_int")
                
                stats_type = self.df_traite.group_by("INST PAIEMENT").agg(
                    pl.len().alias("nombre_operations")
                ).sort("nombre_operations", descending=True)
                
                self.log_message.emit("Statistiques des types d'opération:", "INFO")
                for row in stats_type.iter_rows():
                    self.log_message.emit(f"  {row[0]}: {row[1]:,} opérations", "INFO")
            
            # Ajouter la colonne conditionnelle NATURE OPE CPI
            self.log_message.emit("Ajout de la colonne NATURE OPE CPI...", "INFO")
            
            if "BANQUETIRE" in self.df_traite.columns:
                self.df_traite = self.df_traite.with_columns(
                    pl.when(pl.col("BANQUETIRE").cast(pl.Utf8).is_in(["2", "02", "002"]))
                    .then(pl.lit("RETOUR"))
                    .otherwise(pl.lit("ALLER"))
                    .alias("NATURE OPE CPI")
                )
                
                stats_nature = self.df_traite.group_by("NATURE OPE CPI").agg(
                    pl.len().alias("nombre_operations")
                ).sort("nombre_operations", descending=True)
                
                self.log_message.emit("Statistiques des natures d'opération:", "INFO")
                for row in stats_nature.iter_rows():
                    self.log_message.emit(f"  {row[0]}: {row[1]:,} opérations", "INFO")
            else:
                self.log_message.emit("⚠️ Colonne 'BANQUETIRE' non trouvée - colonne NATURE OPE CPI non créée", "WARNING")
            
            # Dupliquer la colonne MONTANTOPERATION en MONTANTS et supprimer les signes négatifs
            self.log_message.emit("Duplication de MONTANTOPERATION en MONTANTS sans signes négatifs...", "INFO")
            
            if "MONTANTOPERATION" in self.df_traite.columns:
                self.df_traite = self.df_traite.with_columns(
                    pl.col("MONTANTOPERATION").alias("MONTANTS")
                ).with_columns(
                    pl.col("MONTANTS").str.replace_all(r"^\s*-\s*", "").alias("MONTANTS")
                )
                
                self.log_message.emit("Colonne MONTANTS créée avec succès (valeurs absolues)", "SUCCESS")
            else:
                self.log_message.emit("⚠️ Colonne 'MONTANTOPERATION' non trouvée - colonne MONTANTS non créée", "WARNING")
            
            # Créer les colonnes conditionnelles DEBIT et CREDIT selon SENS
            self.log_message.emit("Création des colonnes DEBIT et CREDIT...", "INFO")
            
            if "SENS" in self.df_traite.columns and "MONTANTS" in self.df_traite.columns:
                self.df_traite = self.df_traite.with_columns(
                    pl.when(pl.col("SENS").str.starts_with("D"))
                    .then(pl.col("MONTANTS"))
                    .otherwise(pl.lit(""))
                    .alias("DEBIT"),
                    pl.when(pl.col("SENS").str.starts_with("C"))
                    .then(pl.col("MONTANTS"))
                    .otherwise(pl.lit(""))
                    .alias("CREDIT")
                )
                
                total_debit = self.df_traite.filter(pl.col("DEBIT") != "").height
                total_credit = self.df_traite.filter(pl.col("CREDIT") != "").height
                
                self.log_message.emit(f"Colonnes DEBIT/CREDIT créées:", "INFO")
                self.log_message.emit(f"  Opérations au débit: {total_debit:,}", "INFO")
                self.log_message.emit(f"  Opérations au crédit: {total_credit:,}", "INFO")
            else:
                if "SENS" not in self.df_traite.columns:
                    self.log_message.emit("⚠️ Colonne 'SENS' non trouvée - colonnes DEBIT/CREDIT non créées", "WARNING")
                if "MONTANTS" not in self.df_traite.columns:
                    self.log_message.emit("⚠️ Colonne 'MONTANTS' non trouvée - colonnes DEBIT/CREDIT non créées", "WARNING")
            
            # Créer un résumé détaillé par fichier
            self.log_message.emit("Création du résumé détaillé par fichier...", "INFO")
            
            colonnes_disponibles = self.df_traite.columns
            has_sens = "SENS" in colonnes_disponibles
            
            resume_expr = [pl.len().alias("nombre_lignes")]
            
            if has_sens:
                resume_expr.extend([
                    pl.col("SENS").filter(pl.col("SENS") == "D").len().alias("nombre_lignes_debit"),
                    pl.col("SENS").filter(pl.col("SENS") == "C").len().alias("nombre_lignes_credit")
                ])
            
            # Créer le résumé et le stocker pour l'export
            df_resume = self.df_traite.group_by("source_file").agg(*resume_expr).sort("source_file")
            
            end_time = time.time()
            temps_traitement = end_time - start_time
            
            self.log_message.emit(f"Traitement terminé en {temps_traitement:.2f} secondes", "SUCCESS")
            self.log_message.emit(f"Données traitées: {self.df_traite.height:,} lignes", "SUCCESS")
            
            if has_sens:
                total_debit = self.df_traite.filter(pl.col("SENS") == "D").height
                total_credit = self.df_traite.filter(pl.col("SENS") == "C").height
                self.log_message.emit(f"Statistiques globales - Débit: {total_debit:,} lignes, Crédit: {total_credit:,} lignes", "INFO")
            
            # Émettre le signal avec les données traitées et le résumé
            self.finished_processing.emit(self.df_traite)
            
        except Exception as e:
            self.log_message.emit(f"Erreur lors du traitement: {str(e)}", "ERROR")
    
    def _lettrage_data(self):
        """Logique de lettrage des données"""
        try:
            df_traite = self.args[0]
            self.log_message.emit("Début du lettrage des opérations...", "INFO")
            start_time = time.time()
            
            # Vérifier les colonnes nécessaires pour le lettrage
            colonnes_necessaires = ["source_file", "INST PAIEMENT", "MONTANTS", "RIBTIRE", 
                                  "RIBBENEFICIAIRE", "NUM_DOCUMENT", "SENS", "DEBIT", "CREDIT"]
            
            colonnes_manquantes = [col for col in colonnes_necessaires if col not in df_traite.columns]
            
            if colonnes_manquantes:
                self.log_message.emit(f"Colonnes manquantes pour le lettrage: {', '.join(colonnes_manquantes)}", "ERROR")
                return
            
            # Nettoyer la colonne MONTANTS en supprimant les espaces
            df_traite = df_traite.with_columns(
                pl.col("MONTANTS").str.replace_all(r"^\s+|\s+$|\s+", "").alias("MONTANTS")
            )
            
            # Créer un identifiant unique pour chaque groupe de lettrage
            df_traite = df_traite.with_columns(
                pl.col("source_file").fill_null("").cast(pl.String),
                pl.col("INST PAIEMENT").fill_null("NON_CLASSÉ").cast(pl.String),
                pl.col("MONTANTS").fill_null("0").cast(pl.String),
                pl.col("RIBTIRE").fill_null("RIB_INCONNU").cast(pl.String),
                pl.col("RIBBENEFICIAIRE").fill_null("RIB_INCONNU").cast(pl.String),
                pl.col("NUM_DOCUMENT").fill_null("DOC_INCONNU").cast(pl.String),
                pl.col("SENS").fill_null("S_INCONNU").cast(pl.String),
            )
            
            # Créer le groupe de lettrage avec les valeurs nettoyées
            df_traite = df_traite.with_columns(
                (pl.col("source_file") + "_" +
                 pl.col("INST PAIEMENT") + "_" +
                 pl.col("MONTANTS") + "_" +
                 pl.col("RIBTIRE") + "_" +
                 pl.col("RIBBENEFICIAIRE") + "_" +
                 pl.col("NUM_DOCUMENT") + "_" +
                 pl.col("SENS")).alias("groupe_lettrage")
            )
            
            # Ajouter un index pour chaque groupe de lettrage
            df_traite = df_traite.sort("groupe_lettrage").with_columns(
                pl.int_range(pl.len()).over("groupe_lettrage").alias("index_groupe")
            )
            
            # Ajouter 1 pour commencer à 1 au lieu de 0
            df_traite = df_traite.with_columns(
                (pl.col("index_groupe") + 1).alias("index_groupe")
            )
            
            # Initialiser la colonne Statut avec "Paiement" par défaut
            df_traite = df_traite.with_columns(
                pl.lit("Paiement").alias("Statut")
            )
            
            # Créer groupe_lettrage_sans_sens AVEC l'index inclus dans la chaîne
            df_traite = df_traite.with_columns(
                (pl.col("groupe_lettrage")
                 .str.replace_all("_C", "")
                 .str.replace_all("_D", "")
                 .str.replace_all(".txt", "")
                 .str.replace_all(".dsv", "")
                 + "_" + 
                 pl.col("index_groupe").cast(pl.String))
                .alias("groupe_lettrage_sans_sens")
            )
            
            # Grouper par les critères de lettrage et identifier les groupes avec débits et crédits mixtes
            groupes_lettrage = df_traite.group_by("groupe_lettrage_sans_sens").agg([
                pl.col("DEBIT").filter(pl.col("DEBIT") != "").len().alias("count_debit"),
                pl.col("CREDIT").filter(pl.col("CREDIT") != "").len().alias("count_credit"),
                pl.len().alias("nombre_operations")
            ])
            
            self.log_message.emit(f"Nombre de groupes de lettrage créés: {groupes_lettrage.height}", "INFO")
            
            # Identifier les groupes qui ont à la fois des débits et des crédits
            groupes_rejet = groupes_lettrage.filter(
                (pl.col("count_debit") > 0) &
                (pl.col("count_credit") > 0)
            ).select("groupe_lettrage_sans_sens").to_series().to_list()
            
            self.log_message.emit(f"Nombre de groupes identifiés comme rejets: {len(groupes_rejet)}", "INFO")
            
            # Marquer les groupes rejetés
            if groupes_rejet:
                df_traite = df_traite.with_columns(
                    pl.when(pl.col("groupe_lettrage_sans_sens").is_in(groupes_rejet))
                    .then(pl.lit("Rejet"))
                    .otherwise(pl.col("Statut"))
                    .alias("Statut")
                )
            
            # Statistiques du lettrage
            stats_lettrage = df_traite.group_by("Statut").agg([
                pl.len().alias("nombre_operations")
            ]).sort("nombre_operations", descending=True)
            
            self.log_message.emit("Statistiques de lettrage:", "INFO")
            for row in stats_lettrage.iter_rows():
                statut, nb_ops = row[0], row[1]
                self.log_message.emit(f"  {statut}: {nb_ops:,} opérations", "INFO")
            
            # Supprimer les colonnes de lettrage temporaires
            colonnes_a_supprimer = ["groupe_lettrage_sans_sens", "index_groupe", "groupe_lettrage"]
            colonnes_existantes = [col for col in colonnes_a_supprimer if col in df_traite.columns]
            
            if colonnes_existantes:
                df_traite = df_traite.drop(colonnes_existantes)
            
            end_time = time.time()
            temps_lettrage = end_time - start_time
            
            self.log_message.emit(f"Lettrage terminé en {temps_lettrage:.2f} secondes", "SUCCESS")
            
            # Émettre le signal avec les données lettrées
            self.finished_lettrage.emit(df_traite)
            
        except Exception as e:
            self.log_message.emit(f"Erreur lors du lettrage: {str(e)}", "ERROR")
    
    def _lettrage_data_filtre(self):
        """Logique de lettrage des données avec filtre sur NATURE OPE CPI"""
        try:
            df_traite = self.args[0]
            nature_filtre = self.args[1]  # Pour information seulement, le filtrage est déjà fait
            
            self.log_message.emit("Début du lettrage des opérations avec filtre...", "INFO")
            start_time = time.time()
            
            # Vérifier les colonnes nécessaires pour le lettrage
            colonnes_necessaires = ["source_file", "INST PAIEMENT", "MONTANTS", "RIBTIRE", 
                                  "RIBBENEFICIAIRE", "NUM_DOCUMENT", "SENS", "DEBIT", "CREDIT"]
            
            colonnes_manquantes = [col for col in colonnes_necessaires if col not in df_traite.columns]
            
            if colonnes_manquantes:
                self.log_message.emit(f"Colonnes manquantes pour le lettrage: {', '.join(colonnes_manquantes)}", "ERROR")
                return
            
            # Nettoyer la colonne MONTANTS en supprimant les espaces
            df_traite = df_traite.with_columns(
                pl.col("MONTANTS").str.replace_all(r"^\s+|\s+$|\s+", "").alias("MONTANTS")
            )
            
            # Créer un identifiant unique pour chaque groupe de lettrage
            df_traite = df_traite.with_columns(
                pl.col("source_file").fill_null("").cast(pl.String),
                pl.col("INST PAIEMENT").fill_null("NON_CLASSÉ").cast(pl.String),
                pl.col("MONTANTS").fill_null("0").cast(pl.String),
                pl.col("RIBTIRE").fill_null("RIB_INCONNU").cast(pl.String),
                pl.col("RIBBENEFICIAIRE").fill_null("RIB_INCONNU").cast(pl.String),
                pl.col("NUM_DOCUMENT").fill_null("DOC_INCONNU").cast(pl.String),
                pl.col("SENS").fill_null("S_INCONNU").cast(pl.String),
            )
            
            # Créer le groupe de lettrage avec les valeurs nettoyées
            df_traite = df_traite.with_columns(
                (pl.col("source_file") + "_" +
                 pl.col("INST PAIEMENT") + "_" +
                 pl.col("MONTANTS") + "_" +
                 pl.col("RIBTIRE") + "_" +
                 pl.col("RIBBENEFICIAIRE") + "_" +
                 pl.col("NUM_DOCUMENT") + "_" +
                 pl.col("SENS")).alias("groupe_lettrage")
            )
            
            # Ajouter un index pour chaque groupe de lettrage
            df_traite = df_traite.sort("groupe_lettrage").with_columns(
                pl.int_range(pl.len()).over("groupe_lettrage").alias("index_groupe")
            )
            
            # Ajouter 1 pour commencer à 1 au lieu de 0
            df_traite = df_traite.with_columns(
                (pl.col("index_groupe") + 1).alias("index_groupe")
            )
            
            # Initialiser la colonne Statut avec "Paiement" par défaut
            df_traite = df_traite.with_columns(
                pl.lit("Paiement").alias("Statut")
            )
            
            # Créer groupe_lettrage_sans_sens AVEC l'index inclus dans la chaîne
            df_traite = df_traite.with_columns(
                (pl.col("groupe_lettrage")
                 .str.replace_all("_C", "")
                 .str.replace_all("_D", "")
                 .str.replace_all(".txt", "")
                 .str.replace_all(".dsv", "")
                 + "_" + 
                 pl.col("index_groupe").cast(pl.String))
                .alias("groupe_lettrage_sans_sens")
            )
            
            # Grouper par les critères de lettrage et identifier les groupes avec débits et crédits mixtes
            groupes_lettrage = df_traite.group_by("groupe_lettrage_sans_sens").agg([
                pl.col("DEBIT").filter(pl.col("DEBIT") != "").len().alias("count_debit"),
                pl.col("CREDIT").filter(pl.col("CREDIT") != "").len().alias("count_credit"),
                pl.len().alias("nombre_operations")
            ])
            
            self.log_message.emit(f"Nombre de groupes de lettrage créés: {groupes_lettrage.height}", "INFO")
            
            # Identifier les groupes qui ont à la fois des débits et des crédits
            groupes_rejet = groupes_lettrage.filter(
                (pl.col("count_debit") > 0) &
                (pl.col("count_credit") > 0)
            ).select("groupe_lettrage_sans_sens").to_series().to_list()
            
            self.log_message.emit(f"Nombre de groupes identifiés comme rejets: {len(groupes_rejet)}", "INFO")
            
            # Marquer les groupes rejetés
            if groupes_rejet:
                df_traite = df_traite.with_columns(
                    pl.when(pl.col("groupe_lettrage_sans_sens").is_in(groupes_rejet))
                    .then(pl.lit("Rejet"))
                    .otherwise(pl.col("Statut"))
                    .alias("Statut")
                )
            
            # Statistiques du lettrage
            stats_lettrage = df_traite.group_by("Statut").agg([
                pl.len().alias("nombre_operations")
            ])
            
            for stat in stats_lettrage.iter_rows():
                statut, nb_ops = stat[0], stat[1]
                self.log_message.emit(f"  {statut}: {nb_ops:,} opérations", "INFO")
            
            temps_lettrage = time.time() - start_time
            
            self.log_message.emit(f"Lettrage avec filtre terminé en {temps_lettrage:.2f} secondes", "SUCCESS")
            
            # Émettre le signal avec les données lettrées
            self.finished_lettrage.emit(df_traite)
            
        except Exception as e:
            self.log_message.emit(f"Erreur lors du lettrage avec filtre: {str(e)}", "ERROR")

class ModernButton(QPushButton):
    """Bouton moderne avec animations et style amélioré"""
    def __init__(self, text, color="#5ea3fa", hover_color="#60a5fa", parent=None):
        super().__init__(text, parent)
        
        # Animation parameters
        self.normal_scale = 1.0
        self.hover_scale = 1.05
        self.pressed_scale = 0.95
        
        # Animation parameters
        self._current_step = 0
        self.animation_steps = 8
        self.animation_duration = 150
        
        # Lueur animation parameters
        self.lueur_animation_timer = QTimer()
        self.lueur_animation_timer.timeout.connect(self.animer_lueur)
        self.lueur_position = 0.5  # Position de la lueur (0.0 = haut, 0.5 = centre, 1.0 = bas)
        self.lueur_target = 0.5
        self.lueur_animation_speed = 0.02
        
        # Style parameters
        self.normal_color = color
        self.hover_color = hover_color
        
        # Utiliser un stylesheet fixe pour éviter les erreurs de parsing
        self.update_lueur_style()
        
        # Setup animations
        self.setup_animations()
    
    def update_lueur_style(self):
        """Met à jour le style avec la position actuelle de la lueur"""
        lueur_stop = max(0.1, min(0.9, self.lueur_position))
        
        stylesheet = f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {self.normal_color}, stop:{lueur_stop} rgba(255, 255, 255, 0.1), 
                    stop:1 {self.normal_color});
                color: #FFFAFA;
                border: 2px solid rgba(255, 255, 255, 0.3);
                border-radius: 15px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: 600;
                min-height: 32px;
                min-width: 140px;
                text-transform: uppercase;
                letter-spacing: 1px;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {self.hover_color}, stop:{lueur_stop} rgba(255, 255, 255, 0.2), 
                    stop:1 {self.hover_color});
                border: 2px solid rgba(255, 255, 255, 0.5);
                color: #FFFAFA;
            }}
            QPushButton:pressed {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {self.normal_color}, stop:{lueur_stop} rgba(0, 0, 0, 0.1), 
                    stop:1 {self.hover_color});
                border: 2px solid rgba(255, 255, 255, 0.6);
                color: #FFFAFA;
            }}
            QPushButton:disabled {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #9ca3af, stop:0.5 rgba(255, 255, 255, 0.05), stop:1 #9ca3af);
                color: #FFFAFA;
                border: 2px solid rgba(255, 255, 255, 0.1);
            }}
        """
        self.setStyleSheet(stylesheet)
    
    def animer_lueur(self):
        """Anime la position de la lueur"""
        if abs(self.lueur_position - self.lueur_target) < 0.01:
            self.lueur_animation_timer.stop()
            return
        
        # Animation fluide vers la cible
        if self.lueur_position < self.lueur_target:
            self.lueur_position += self.lueur_animation_speed
            if self.lueur_position > self.lueur_target:
                self.lueur_position = self.lueur_target
        else:
            self.lueur_position -= self.lueur_animation_speed
            if self.lueur_position < self.lueur_target:
                self.lueur_position = self.lueur_target
        
        self.update_lueur_style()
    
    def setup_animations(self):
        """Configure les animations pour le bouton"""
        # Animation pour l'effet de scale (utilise la géométrie)
        self._original_geometry = None
        self._animation_timer = QTimer()
        self._animation_timer.setSingleShot(True)
        self._animation_timer.timeout.connect(self._animate_step)
        
        # Current animation values
        self._current_scale = self.normal_scale
        self._target_scale = self.normal_scale
    
    def enterEvent(self, event):
        """Quand le curseur entre dans le bouton"""
        if not self.isEnabled():
            return
        self._target_scale = self.hover_scale
        self.lueur_target = 0.1  # Lueur monte en haut
        self.lueur_animation_timer.start(16)  # ~60 FPS pour l'animation fluide
        self._start_animation()
        super().enterEvent(event)
    
    def leaveEvent(self, event):
        """Quand le curseur quitte le bouton"""
        if not self.isEnabled():
            return
        self._target_scale = self.normal_scale
        self.lueur_target = 0.5  # Lueur retourne au centre
        self.lueur_animation_timer.start(16)
        self._start_animation()
        super().leaveEvent(event)
    
    def mousePressEvent(self, event):
        """Quand le bouton est pressé"""
        if not self.isEnabled():
            return
        self._target_scale = self.pressed_scale
        self.lueur_target = 0.8  # Lueur va vers le bas quand pressé
        self.lueur_animation_timer.start(16)
        self._start_animation()
        super().mousePressEvent(event)
    
    def mouseReleaseEvent(self, event):
        """Quand le bouton est relâché"""
        if not self.isEnabled():
            return
        # Vérifier si le curseur est toujours sur le bouton
        if self.rect().contains(self.mapFromParent(event.pos())):
            self._target_scale = self.hover_scale
            self.lueur_target = 0.1  # Lueur reste en haut si curseur dessus
        else:
            self._target_scale = self.normal_scale
            self.lueur_target = 0.5  # Lueur retourne au centre
        self.lueur_animation_timer.start(16)
        self._start_animation()
        super().mouseReleaseEvent(event)
    
    def _start_animation(self):
        """Démarre l'animation de scale uniquement"""
        if self._original_geometry is None:
            self._original_geometry = self.geometry()
        
        self._current_step = 0
        self._animation_timer.start(self.animation_duration // self.animation_steps)
    
    def _animate_step(self):
        """Effectue une étape de l'animation"""
        if self._current_step >= self.animation_steps:
            return
        
        # Calculer le progress
        progress = (self._current_step + 1) / self.animation_steps
        
        # Animer le scale
        self._current_scale = self._current_scale + (self._target_scale - self._current_scale) * progress
        
        # Appliquer les changements
        self._apply_scale()
        
        self._current_step += 1
        
        # Continuer l'animation
        if self._current_step < self.animation_steps:
            self._animation_timer.start(self.animation_duration // self.animation_steps)
    
    def _apply_scale(self):
        """Applique le scale actuel au bouton"""
        if self._original_geometry is None:
            return
        
        # Calculer la nouvelle taille et position
        original_width = self._original_geometry.width()
        original_height = self._original_geometry.height()
        original_x = self._original_geometry.x()
        original_y = self._original_geometry.y()
        
        new_width = int(original_width * self._current_scale)
        new_height = int(original_height * self._current_scale)
        
        # Centrer le bouton pendant le scale
        new_x = original_x + (original_width - new_width) // 2
        new_y = original_y + (original_height - new_height) // 2
        
        # Appliquer la nouvelle géométrie
        self.setGeometry(new_x, new_y, new_width, new_height)
    
    def resizeEvent(self, event):
        """Gère le redimensionnement pour maintenir l'animation"""
        super().resizeEvent(event)
        if self._current_scale == self.normal_scale:
            self._original_geometry = self.geometry()

class ModernTextEdit(QTextEdit):
    """Zone de texte moderne avec style terminal dark cosmic"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("""
            QTextEdit {
                background-color: #010001;
                color: #60a5fa;
                border: none;
                border-radius: 8px;
                padding: 15px;
                font-family: 'Consolas', 'Monaco', monospace;
                font-size: 12px;
            }
            /* Style pour la scrollbar du QTextEdit */
            QScrollBar:vertical {
                background-color: #010001;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background-color: #1ecce8;
                border-radius: 6px;
                min-height: 20px;
            }
            /* Supprimer les flèches haut et bas */
            QScrollBar::add-line:vertical {
                height: 0px;
                width: 0px;
                background: none;
            }
            QScrollBar::sub-line:vertical {
                height: 0px;
                width: 0px;
                background: none;
            }
        """)
        self.setReadOnly(True)

class HistogramWidget(QWidget):
    """Histogramme simple avec axes X et Y pour afficher les pourcentages de paiements et rejets"""
    def __init__(self, width=220, height=120, parent=None):
        super().__init__(parent)
        self.setFixedSize(width, height)
        
        # Données pour l'histogramme
        self.paiement_percent = 0.0
        self.rejet_percent = 0.0
        
        # Couleurs
        self.paiement_color = QColor(16, 185, 129)  # #10b981 (vert)
        self.rejet_color = QColor(239, 68, 68)     # #ef4444 (rouge)
        self.background_color = QColor(1, 0, 1)       # #010001
        self.text_color = QColor(226, 232, 240)      # #e2e8f0
        self.grid_color = QColor(30, 41, 59)         # #1e293b
        self.axis_color = QColor(148, 163, 184)       # #94a3b8
        
        # Animation
        self.animation_steps = 20
        self.current_step = 0
        self.animation_timer = QTimer()
        self.animation_timer.timeout.connect(self.animate_to_target)
        
        # Valeurs cibles pour animation
        self.target_paiement = 0.0
        self.target_rejet = 0.0
        self.current_paiement = 0.0
        self.current_rejet = 0.0
    
    def set_values(self, paiement_percent, rejet_percent, animate=True):
        """Définit les valeurs et lance l'animation"""
        self.target_paiement = min(max(paiement_percent, 0), 100)
        self.target_rejet = min(max(rejet_percent, 0), 100)
        
        if not animate:
            self.current_paiement = self.target_paiement
            self.current_rejet = self.target_rejet
            self.update()
        else:
            self.current_step = 0
            self.animation_timer.start(30)  # 30ms entre chaque étape
    
    def animate_to_target(self):
        """Anime les valeurs vers les cibles"""
        if self.current_step >= self.animation_steps:
            self.animation_timer.stop()
            self.current_paiement = self.target_paiement
            self.current_rejet = self.target_rejet
            self.update()
            return
        
        # Calculer la progression (ease-in-out)
        progress = self.current_step / self.animation_steps
        if progress < 0.5:
            eased_progress = 2 * progress * progress
        else:
            eased_progress = 1 - pow(-2 * progress + 2, 2) / 2
        
        # Interpoler les valeurs
        self.current_paiement = self.current_paiement + (self.target_paiement - self.current_paiement) * eased_progress
        self.current_rejet = self.current_rejet + (self.target_rejet - self.current_rejet) * eased_progress
        
        self.update()
        self.current_step += 1
    
    def paintEvent(self, event):
        """Dessine l'histogramme simple avec axes X et Y"""
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        
        # Dimensions
        rect = self.rect()
        width = rect.width()
        height = rect.height()
        
        # Marges pour les axes et labels (hauteur augmentée pour l'axe Y)
        margin_left = 40   # Espace pour les labels Y
        margin_right = 20
        margin_top = 10    # augmenter la hauteur de l'axe y
        margin_bottom = 40  # Espace pour les labels X
        
        # Zone de dessin du graphique (plus de hauteur pour l'axe Y)
        chart_x = margin_left
        chart_y = margin_top
        chart_width = width - margin_left - margin_right
        chart_height = height - margin_top - margin_bottom
        
        # Fond
        painter.fillRect(rect, self.background_color)
        
        # Dessiner les axes X et Y
        painter.setPen(QPen(self.axis_color, 2))
        
        # Axe Y (vertical gauche)
        painter.drawLine(chart_x, chart_y, chart_x, chart_y + chart_height)
        
        # Axe X (horizontal bas)
        painter.drawLine(chart_x, chart_y + chart_height, chart_x + chart_width, chart_y + chart_height)
        
        # Labels de l'axe Y (pourcentages) - en gras
        painter.setPen(QPen(self.text_color))
        font = QFont("Arial", 8, QFont.Weight.Bold)  # En gras
        painter.setFont(font)
        for i in range(0, 101, 25):
            y = chart_y + chart_height - (chart_height * i / 100)
            # Aligner parfaitement avec l'axe Y
            painter.drawText(5, int(y - 5), 30, 15, 
                          Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter, f"{i}%")
        
        # Calculer la largeur des barres (avec espacement entre elles)
        bar_width = chart_width // 4  # Barres un peu plus petites pour laisser de l'espace
        bar_spacing = bar_width // 2
        
        # Position X pour les barres (avec espacement entre les deux)
        paiement_x = chart_x + bar_spacing
        rejet_x = chart_x + bar_width + bar_spacing + 10  # +10px d'espacement entre les barres
        
        # Hauteur des barres (basées sur les pourcentages)
        paiement_height = (self.current_paiement / 100) * chart_height
        rejet_height = (self.current_rejet / 100) * chart_height
        
        # Dessiner les barres avec ombres
        # Barre de paiement
        if paiement_height > 0:
            # Ombre
            shadow_rect = QRect(paiement_x + 2, chart_y + chart_height - int(paiement_height) + 2, 
                              bar_width, int(paiement_height))
            painter.fillRect(shadow_rect, QColor(0, 0, 0, 50))
            
            # Barre principale
            paiement_rect = QRect(paiement_x, chart_y + chart_height - int(paiement_height), 
                               bar_width, int(paiement_height))
            painter.fillRect(paiement_rect, self.paiement_color)
        
        # Barre de rejet
        if rejet_height > 0:
            # Ombre
            shadow_rect = QRect(rejet_x + 2, chart_y + chart_height - int(rejet_height) + 2, 
                              bar_width, int(rejet_height))
            painter.fillRect(shadow_rect, QColor(0, 0, 0, 50))
            
            # Barre principale
            rejet_rect = QRect(rejet_x, chart_y + chart_height - int(rejet_height), 
                             bar_width, int(rejet_height))
            painter.fillRect(rejet_rect, self.rejet_color)
        
        # Labels de l'axe X - alignés avec les barres
        painter.setPen(QPen(self.text_color))
        font = QFont("Arial", 9, QFont.Weight.Bold)
        painter.setFont(font)
        
        # Label "Paiements" - centré sous la barre
        paiement_center_x = paiement_x + bar_width // 2
        painter.drawText(paiement_center_x - 35, chart_y + chart_height + 8, 70, 20, 
                      Qt.AlignmentFlag.AlignCenter, "Paiements")
        
        # Label "Rejets" - centré sous la barre
        rejet_center_x = rejet_x + bar_width // 2
        painter.drawText(rejet_center_x - 25, chart_y + chart_height + 8, 50, 20, 
                      Qt.AlignmentFlag.AlignCenter, "Rejets")
        
        # Afficher les valeurs au-dessus des barres (décalées plus haut)
        font.setPointSize(10)  # Police plus grande pour meilleure visibilité
        font.setBold(True)
        painter.setFont(font)
        
        # Valeur paiement - centrée au-dessus de la barre
        if self.current_paiement > 0:
            value_y = chart_y + chart_height - int(paiement_height) - 22  # Décalage de 22px
            if value_y < chart_y - 15:  # S'assurer que ça ne sort pas trop
                value_y = chart_y - 15
            painter.setPen(QPen(self.paiement_color))
            painter.drawText(paiement_center_x - 25, value_y, 50, 20, 
                          Qt.AlignmentFlag.AlignCenter, f"{self.current_paiement:.1f}%")
        
        # Valeur rejet - centrée au-dessus de la barre
        if self.current_rejet > 0:
            value_y = chart_y + chart_height - int(rejet_height) - 22  # Décalage de 22px
            if value_y < chart_y - 15:  # S'assurer que ça ne sort pas trop
                value_y = chart_y - 15
            painter.setPen(QPen(self.rejet_color))
            painter.drawText(rejet_center_x - 25, value_y, 50, 20, 
                          Qt.AlignmentFlag.AlignCenter, f"{self.current_rejet:.1f}%")
        
        painter.end()

class CircularProgress(QWidget):
    """Jauge circulaire animée avec arc de progression"""
    def __init__(self, size=120, parent=None):
        super().__init__(parent)
        self.setFixedSize(size, size)
        
        # Propriétés de la jauge
        self.value = 0.0  # Valeur actuelle (0-100)
        self.target_value = 0.0  # Valeur cible
        self.animation_steps = 30
        self.current_step = 0
        self.animation_timer = QTimer()
        self.animation_timer.timeout.connect(self.animate_to_target)
        
        # Couleurs et styles
        self.background_color = QColor(1, 0, 1)  # #010001
        self.arc_color = QColor(30, 204, 232)  # #1ecce8
        self.text_color = QColor(226, 232, 240)  # #e2e8f0
        self.success_color = QColor(16, 185, 129)  # #10b981
        self.warning_color = QColor(245, 158, 11)  # #f59e0b
        self.danger_color = QColor(239, 68, 68)  # #ef4444
        
        # Dimensions
        self.arc_width = 12
        self.start_angle = -90  # Commence en haut
        self.span_angle = 270  # Arc de 270° (3/4 de cercle)
        
        # Animation
        self.animation_speed = 20  # ms entre chaque étape
        
    def set_value(self, value, animate=True):
        """Définit la valeur cible et lance l'animation"""
        if not animate:
            self.value = min(max(value, 0), 100)
            self.update()
        else:
            self.target_value = min(max(value, 0), 100)
            self.current_step = 0
            self.animation_timer.start(self.animation_speed)
    
    def animate_to_target(self):
        """Anime la progression vers la valeur cible"""
        if self.current_step >= self.animation_steps:
            self.animation_timer.stop()
            self.value = self.target_value
            self.update()
            return
        
        # Calculer la progression (ease-in-out)
        progress = self.current_step / self.animation_steps
        if progress < 0.5:
            eased_progress = 2 * progress * progress
        else:
            eased_progress = 1 - pow(-2 * progress + 2, 2) / 2
        
        # Interpoler entre valeur actuelle et cible
        self.value = self.value + (self.target_value - self.value) * eased_progress
        self.update()
        
        self.current_step += 1
    
    def get_color_for_value(self):
        """Retourne la couleur selon la valeur"""
        if self.value > 77:
            return self.arc_color  # Bleu (#1ecce8)
        elif self.value >= 56:
            return self.warning_color  # Orange (#f59e0b)
        else:
            return self.danger_color  # Rouge (#ef4444)
    
    def paintEvent(self, event):
        """Dessine la jauge circulaire"""
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        
        # Dimensions
        rect = self.rect()
        center = rect.center()
        radius = min(rect.width(), rect.height()) // 2 - self.arc_width
        
        # Dessiner l'arc de fond
        painter.setPen(QPen(self.background_color.lighter(150), self.arc_width))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawArc(int(center.x() - radius), int(center.y() - radius), 
                     int(radius * 2), int(radius * 2), 
                     int(self.start_angle * 16), int(self.span_angle * 16))
        
        # Dessiner l'arc de progression
        if self.value > 0:
            color = self.get_color_for_value()
            painter.setPen(QPen(color, self.arc_width, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap))
            
            # Calculer l'angle de progression
            progress_angle = (self.value / 100) * self.span_angle
            painter.drawArc(int(center.x() - radius), int(center.y() - radius), 
                         int(radius * 2), int(radius * 2), 
                         int(self.start_angle * 16), int(progress_angle * 16))
        
        # Dessiner le texte central
        painter.setPen(QPen(self.text_color))
        font = QFont("Arial", 16, QFont.Weight.Bold)
        painter.setFont(font)
        
        # Texte du pourcentage
        percentage_text = f"{int(self.value)}%"
        text_rect = painter.boundingRect(rect, Qt.AlignmentFlag.AlignCenter, percentage_text)
        painter.drawText(text_rect, Qt.AlignmentFlag.AlignCenter, percentage_text)
        
        painter.end()

class VentilationOdometerWidget(QWidget):
    """Jauge de ventilation style compteur automobile"""
    def __init__(self, width=160, height=80, parent=None):
        super().__init__(parent)
        self.setFixedSize(width, height)
        
        # Propriétés de la jauge
        self.value = 0.0  # Valeur actuelle (0-100)
        self.target_value = 0.0  # Valeur cible
        self.animation_steps = 30
        self.current_step = 0
        self.animation_timer = QTimer()
        self.animation_timer.timeout.connect(self.animate_to_target)
        
        # Couleurs et styles
        self.background_color = QColor(1, 0, 1)  # #010001
        self.text_color = QColor(226, 232, 240)  # #e2e8f0
        self.danger_color = QColor(239, 68, 68)  # #ef4444 (rouge)
        self.warning_color = QColor(245, 158, 11)  # #f59e0b (orange)
        self.success_color = QColor(16, 185, 129)  # #10b981 (vert)
        
        # Dimensions du compteur
        self.meter_width = width - 40
        self.meter_height = 25
        self.corner_radius = 8
        
        # Animation
        self.animation_speed = 20  # ms entre chaque étape
        
    def set_value(self, value, animate=True):
        """Définit la valeur cible et lance l'animation"""
        if not animate:
            self.value = min(max(value, 0), 100)
            self.update()
        else:
            self.target_value = min(max(value, 0), 100)
            self.current_step = 0
            self.animation_timer.start(self.animation_speed)
    
    def animate_to_target(self):
        """Anime la progression vers la valeur cible"""
        if self.current_step >= self.animation_steps:
            self.animation_timer.stop()
            self.value = self.target_value
            self.update()
            return
        
        # Calculer la progression (ease-in-out)
        progress = self.current_step / self.animation_steps
        if progress < 0.5:
            eased_progress = 2 * progress * progress
        else:
            eased_progress = 1 - pow(-2 * progress + 2, 2) / 2
        
        # Interpoler entre valeur actuelle et cible
        self.value = self.value + (self.target_value - self.value) * eased_progress
        self.update()
        
        self.current_step += 1
    
    def get_color_for_value(self):
        """Retourne la couleur selon la valeur"""
        if self.value <= 40:
            return self.danger_color  # Rouge (0-40%)
        elif self.value <= 67:
            return self.warning_color  # Orange (40-67%)
        else:
            return self.success_color  # Vert (>67%)
    
    def paintEvent(self, event):
        """Dessine le compteur style automobile"""
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        
        # Dimensions
        rect = self.rect()
        
        # Position du compteur
        meter_x = 20
        meter_y = rect.height() // 2 - self.meter_height // 2
        meter_rect = QRect(meter_x, meter_y, self.meter_width, self.meter_height)
        
        # Dessiner le fond du compteur
        painter.setBrush(QBrush(self.background_color.lighter(120)))
        painter.setPen(QPen(self.background_color.lighter(200), 2))
        painter.drawRoundedRect(meter_rect, self.corner_radius, self.corner_radius)
        
        # Dessiner la barre de progression
        if self.value > 0:
            color = self.get_color_for_value()
            progress_width = int((self.meter_width - 4) * self.value / 100)
            progress_rect = QRect(meter_x + 2, meter_y + 2, progress_width, self.meter_height - 4)
            
            # Dégradé subtil pour la barre
            gradient = QLinearGradient(progress_rect.left(), progress_rect.top(), 
                                     progress_rect.left(), progress_rect.bottom())
            gradient.setColorAt(0, color.lighter(110))
            gradient.setColorAt(0.5, color)
            gradient.setColorAt(1, color.darker(110))
            
            painter.setBrush(QBrush(gradient))
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawRoundedRect(progress_rect, self.corner_radius - 2, self.corner_radius - 2)
        
        # Dessiner les graduations (petits traits)
        painter.setPen(QPen(self.text_color, 1))
        for i in range(0, 101, 20):  # Graduations tous les 20%
            x = meter_x + 2 + int((self.meter_width - 4) * i / 100)
            painter.drawLine(x, meter_y + self.meter_height - 5, x, meter_y + self.meter_height - 2)
        
        # Dessiner le texte du pourcentage
        painter.setPen(QPen(self.text_color))
        font = QFont("Arial", 14, QFont.Weight.Bold)  # Police par défaut
        # Essayer d'utiliser une police digitale si disponible
        try:
            font.setFamily("Digital-7")
        except:
            pass  # Garder Arial si Digital-7 n'est pas disponible
        painter.setFont(font)
        
        percentage_text = f"{int(self.value)}%"
        text_rect = QRect(meter_x, meter_y - 20, self.meter_width, 20)
        painter.drawText(text_rect, Qt.AlignmentFlag.AlignCenter, percentage_text)
        
        painter.end()

class CPIAnalyzerModern(QMainWindow):
    def __init__(self):
        super().__init__()
        self.df_consolidated = None
        self.df_traite = None
        self.df_resume = None
        self.df_solde_par_date = None
        self.worker_thread = None
        
        # Forcer le style sombre pour la barre de titre (Windows)
        try:
            import win32gui
            import win32con
            # Tenter de définir la couleur de la barre de titre
            self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, False)
        except ImportError:
            pass  # win32gui n'est pas installé
        
        self.init_ui()
        self.apply_modern_style()
        
        # Timer pour les animations
        self.animation_timer = QTimer()
        self.animation_timer.timeout.connect(self.update_animations)
        self.animation_timer.start(50)
        
        # Afficher le message de bienvenue
        QTimer.singleShot(1000, self.afficher_message_bienvenue)
        
        # Forcer le redimensionnement correct des widgets après le lancement
        QTimer.singleShot(1500, self.forcer_redimensionnement_widgets)
    
    def afficher_dimensions(self):
        """Affiche les dimensions actuelles de l'application"""
        # Obtenir les dimensions de la fenêtre
        window_geometry = self.geometry()
        window_size = self.size()
        
        # Obtenir les dimensions de l'écran
        screen = self.screen()
        screen_geometry = screen.geometry()
        available_geometry = screen.availableGeometry()
        
        # Obtenir les dimensions physiques/réelles de l'écran
        physical_size = screen.size()
        device_pixel_ratio = screen.devicePixelRatio()
        
        # Calculer les vraies dimensions
        real_width = int(physical_size.width() * device_pixel_ratio)
        real_height = int(physical_size.height() * device_pixel_ratio)
        
        # Afficher les dimensions dans la console
        print("\n" + "="*70)
        print("📏 DIMENSIONS COMPLÈTES DE L'APPLICATION")
        print("="*70)
        print(f"🖥️  Dimensions de la fenêtre : {window_size.width()} x {window_size.height()} pixels")
        print(f"📍 Position de la fenêtre : ({window_geometry.x()}, {window_geometry.y()})")
        print(f"📺 Résolution détectée : {screen_geometry.width()} x {screen_geometry.height()} pixels")
        print(f"✅ Espace disponible : {available_geometry.width()} x {available_geometry.height()} pixels")
        print(f"🔍 Ratio pixel/device : {device_pixel_ratio:.2f}")
        print(f"📐 Dimensions physiques : {physical_size.width()} x {physical_size.height()} pixels")
        print(f" RÉSOLUTION RÉELLE : {real_width} x {real_height} pixels")
        print(f"📊 Taux d'occupation (détecté) : {(window_size.width() / available_geometry.width() * 100):.1f}% x {(window_size.height() / available_geometry.height() * 100):.1f}%")
        print(f"📊 Taux d'occupation (réel) : {(window_size.width() / real_width * 100):.1f}% x {(window_size.height() / real_height * 100):.1f}%")
        print("="*70)
    
    def forcer_redimensionnement_widgets(self):
        """Force le redimensionnement correct de tous les widgets après le lancement en mode maximisé"""
        try:
            print("🔧 Forçage du redimensionnement des widgets...")
            
            # Forcer la mise à jour de la géométrie de la fenêtre
            self.updateGeometry()
            
            # Forcer le redimensionnement du widget central
            central_widget = self.centralWidget()
            if central_widget:
                central_widget.resize(central_widget.size())
                central_widget.updateGeometry()
                
                # Forcer la mise à jour du layout principal
                if central_widget.layout():
                    central_widget.layout().activate()
                    central_widget.layout().update()
            
            # Forcer le redimensionnement des onglets s'ils existent
            if hasattr(self, 'tab_widget') and self.tab_widget:
                self.tab_widget.resize(self.tab_widget.size())
                self.tab_widget.updateGeometry()
                self.tab_widget.repaint()
                
                # Mettre à jour tous les onglets
                for i in range(self.tab_widget.count()):
                    tab = self.tab_widget.widget(i)
                    if tab:
                        tab.resize(tab.size())
                        tab.updateGeometry()
                        tab.repaint()
            
            # Forcer le redimensionnement du journal d'activités
            if hasattr(self, 'log_text') and self.log_text:
                self.log_text.resize(self.log_text.size())
                self.log_text.updateGeometry()
                self.log_text.ensureCursorVisible()
                self.log_text.repaint()
            
            # Forcer la mise à jour de tous les widgets enfants
            self._update_all_child_widgets(self)
            
            # Forcer un repaint complet de l'application
            self.repaint()
            self.update()
            
            # Forcer la mise à jour de la barre de statut
            if hasattr(self, 'status_bar') and self.status_bar:
                self.status_bar.update()
            
            print("✅ Redimensionnement des widgets forcé avec succès")
            
        except Exception as e:
            print(f"❌ Erreur lors du forçage du redimensionnement: {e}")
    
    def _update_all_child_widgets(self, widget):
        """Met à jour récursivement tous les widgets enfants"""
        try:
            if hasattr(widget, 'children'):
                for child in widget.children():
                    if hasattr(child, 'updateGeometry'):
                        child.updateGeometry()
                    if hasattr(child, 'update'):
                        child.update()
                    if hasattr(child, 'repaint'):
                        child.repaint()
                    # Récursif pour les widgets conteneurs
                    if hasattr(child, 'children') and child.children():
                        self._update_all_child_widgets(child)
        except Exception as e:
            print(f"⚠️ Erreur mise à jour widgets enfants: {e}")
    
    def resizeEvent(self, event):
        """Gère l'événement de redimensionnement de la fenêtre"""
        try:
            # Appeler la méthode parente
            super().resizeEvent(event)
            
            # Forcer la mise à jour des layouts après un court délai
            QTimer.singleShot(100, self.forcer_redimensionnement_widgets)
            
        except Exception as e:
            print(f"❌ Erreur lors du redimensionnement: {e}")
            
    def init_ui(self):
        """Initialise l'interface utilisateur"""
        self.setWindowTitle("CPI ANALYZER TOOL")
        
        # Ajouter l'icône de l'application
        icon_path = resource_path('CPI.ico')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
            print(f"✅ Icône CPI chargée: {icon_path}")
        else:
            print(f"❌ Icône CPI non trouvée: {icon_path}")
        
        self.setGeometry(100, 100, 1400, 900)
        
        # D'abord afficher la fenêtre en mode normal
        self.show()
        
        # Puis appliquer le mode maximisé après un très court délai
        QTimer.singleShot(50, lambda: self.showMaximized())
        
        # Mesurer et afficher les dimensions après un court délai
        QTimer.singleShot(1000, self.afficher_dimensions)
        
        # Widget central
        central_widget = QWidget()
        central_widget.setStyleSheet("background-color: #010001;")
        self.setCentralWidget(central_widget)
        
        # Créer le layout principal
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)
        
        # Créer un conteneur pour centrer le header
        header_container = QWidget()
        header_container.setStyleSheet("background-color: #010001;")
        header_container.setFixedHeight(200)  # Hauteur du header
        header_layout = QHBoxLayout(header_container)
        header_layout.setContentsMargins(0, 0, 0, 0)
        header_layout.setSpacing(0)
        
        # Ajouter un espace flexible à gauche et à droite pour centrer
        header_layout.addStretch()
        
        # Créer le header
        self.create_header(header_layout)
        
        # Ajouter un espace flexible à droite
        header_layout.addStretch()
        
        # Ajouter le conteneur centré au layout principal
        main_layout.addWidget(header_container)
        
        # Zone d'informations
        self.create_info_section(main_layout)
                
        # Boutons d'action
        self.create_buttons_section(main_layout)
        
        # Zone de log
        self.create_log_section(main_layout)
        
        # Ajouter un espace flexible pour que le journal puisse s'étendre
        main_layout.addStretch()
        
        # Appliquer le layout au widget central
        central_widget.setLayout(main_layout)
        
        # Barre de status
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("✅ Prêt")
    
    def create_header(self, parent_layout):
        """Crée le header avec vidéo adaptée"""
        # Header qui s'adapte automatiquement à la résolution de la vidéo
        header_height = 200  # Hauteur raisonnable
        
        # Créer le widget vidéo directement sans cadre
        self.video_widget = QVideoWidget()
        self.video_widget.setFixedHeight(header_height)
        self.video_widget.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.video_widget.setStyleSheet("""
            QVideoWidget {
                border-radius: 0px;
                background: black;
                border: none;
            }
        """)
        
        # Créer le lecteur média
        self.media_player = QMediaPlayer()
        self.media_player.setVideoOutput(self.video_widget)
        
        # Configurer le lecteur pour jouer en boucle avec effet boomerang
        self.media_player.setLoops(QMediaPlayer.Loops.Infinite)  # Boucle infinie
        
        # Configurer le lecteur pour s'arrêter à la fin
        self.media_player.mediaStatusChanged.connect(self.on_media_status_changed)
        
        # Variable pour gérer l'effet boomerang
        self.boomerang_forward = True
        self.boomerang_position = 0
        
        # Timer pour la lecture inversée manuelle
        self.boomerang_timer = QTimer()
        self.boomerang_timer.timeout.connect(self.update_boomerang_position)
        self.boomerang_timer.setInterval(40)  # ~25 FPS
        
        # Timer pour redimensionner et forcer le remplissage
        self.resize_timer = QTimer()
        self.resize_timer.setSingleShot(True)
        self.resize_timer.timeout.connect(self.force_video_fill)
        
        # Timer pour redimensionner continuellement (pour s'assurer que ça reste rempli)
        self.continuous_resize_timer = QTimer()
        self.continuous_resize_timer.timeout.connect(self.force_video_fill)
        
        # Timer pour essayer de configurer le lecteur vidéo
        self.video_config_timer = QTimer()
        self.video_config_timer.setSingleShot(True)
        self.video_config_timer.timeout.connect(self.configure_video_player)
        
        # Chemin vers la vidéo
        self.video_path = resource_path('barre title.mp4')
        
        # Logs détaillés pour le diagnostic
        self.log_message(f"🔍 Recherche vidéo: {self.video_path}", 'DEBUG')
        self.log_message(f"📂 Fichier vidéo existe: {os.path.exists(self.video_path)}", 'DEBUG')
        
        # Lister les fichiers dans le dossier pour diagnostic
        try:
            video_dir = os.path.dirname(self.video_path)
            files_in_dir = os.listdir(video_dir) if os.path.exists(video_dir) else []
            video_files = [f for f in files_in_dir if f.endswith(('.mp4', '.avi', '.mov', '.mkv'))]
            self.log_message(f"📋 Fichiers vidéo trouvés: {video_files}", 'DEBUG')
        except Exception as e:
            self.log_message(f"❌ Erreur listing dossier: {e}", 'ERROR')
        
        # Charger la vidéo si elle existe
        if os.path.exists(self.video_path):
            try:
                self.log_message(f"🎬 Chargement vidéo depuis: {self.video_path}", 'INFO')
                self.media_player.setSource(QUrl.fromLocalFile(self.video_path))
                # Démarrer à la seconde 3
                self.media_player.setPosition(3000)  # 3000ms = 3 secondes
                self.media_player.play()
                self.log_message("Vidéo d'en-tête chargée et démarrée à la seconde 3", 'SUCCESS')
                
                # Ajouter directement le widget vidéo au layout parent
                parent_layout.addWidget(self.video_widget)
                
                # Démarrer le redimensionnement continu
                self.continuous_resize_timer.start(50)  # Toutes les 50ms (plus fréquent)
                
                # Premier redimensionnement après un court délai
                self.resize_timer.start(200)
                
                # Configurer le lecteur vidéo après un délai
                self.video_config_timer.start(1000)
                
                # Adapter les dimensions du header IMMÉDIATEMENT
                self.adapt_header_to_video_resolution()
            except Exception as e:
                self.log_message(f"Erreur lors du chargement de la vidéo: {e}", 'ERROR')
                self.create_fallback_header(parent_layout)
        else:
            self.log_message("❌ Vidéo non trouvée - Utilisation de l'en-tête de secours", 'ERROR')
            self.log_message(f"📂 Chemin recherché: {self.video_path}", 'ERROR')
            self.create_fallback_header(parent_layout)
    
    def adapt_header_to_video_resolution(self):
        """Adapte les bordures et marges du header pour occuper tout l'écran"""
        if hasattr(self, 'media_player') and self.media_player:
            try:
                # Obtenir les dimensions de l'écran disponible
                screen = self.screen()
                available_geometry = screen.availableGeometry()
                screen_width = available_geometry.width()
                screen_height = available_geometry.height()
                
                # Obtenir les dimensions réelles de la vidéo
                video_size = self.video_widget.size()
                
                # Attendre que la vidéo soit réellement chargée (dimensions > 0)
                if video_size.width() > 0 and video_size.height() > 0:
                    # Utiliser toute la largeur de l'écran pour la vidéo
                    self.video_widget.setFixedWidth(screen_width)
                    
                    # Ajuster la hauteur pour ne pas cacher les autres éléments
                    header_height = min(video_size.height(), 200)  # Limiter à 200px maximum
                    self.video_widget.setFixedHeight(header_height)
                    
                    self.log_message(f" Vidéo élargie pour occuper tout l'écran", 'SUCCESS')
                    self.log_message(f"📏 Largeur écran : {screen_width} pixels", 'INFO')
                    self.log_message(f"📐 Hauteur écran : {screen_height} pixels", 'INFO')
                    self.log_message(f"🎬 Vidéo redimensionnée : {screen_width} x {header_height} pixels", 'INFO')
                    self.log_message(f"📊 Taux d'occupation : 100% x {(header_height/screen_height*100):.1f}%", 'INFO')
                    
                    # Afficher dans la console
                    print("\n" + "="*70)
                    print("🎬 VIDÉO ÉLARGIE - OCCUPATION COMPLÈTE DE L'ÉCRAN")
                    print("="*70)
                    print(f"📏 Largeur de l'écran : {screen_width} pixels")
                    print(f"📐 Hauteur disponible : {screen_height} pixels")
                    print(f"🎬 Vidéo redimensionnée : {screen_width} x {header_height} pixels")
                    print(f"📊 Occupation : 100% x {(header_height/screen_height*100):.1f}%")
                    print("="*70)
                else:
                    # La vidéo n'est pas encore chargée, réessayer dans 100ms
                    QTimer.singleShot(100, self.adapt_header_to_video_resolution)
                        
            except Exception as e:
                self.log_message(f"Erreur adaptation header : {e}", 'DEBUG')
    
    def force_video_fill(self):
        """Force la vidéo à remplir tout l'espace sans marges"""
        if hasattr(self, 'video_widget') and self.video_widget:
            try:
                # Obtenir les dimensions de l'écran
                screen = self.screen()
                available_geometry = screen.availableGeometry()
                screen_width = available_geometry.width()
                
                # Forcer la vidéo à remplir toute la largeur
                self.video_widget.setFixedWidth(screen_width)
                
            except Exception as e:
                self.log_message(f"Erreur force video fill: {e}", 'DEBUG')
    
    def configure_video_player(self):
        """Configure le lecteur vidéo pour essayer de supprimer les bandes noires"""
        if hasattr(self, 'media_player') and self.media_player:
            try:
                # Essayer de configurer différentes propriétés pour le remplissage
                # Note: Ces propriétés peuvent ne pas fonctionner selon la version PyQt6
                if hasattr(self.media_player, 'setVideoSink'):
                    pass  # Pour les versions plus récentes
                if hasattr(self.media_player, 'setVideoOutput'):
                    pass  # Déjà configuré
                
                # Forcer un nouveau redimensionnement
                self.force_video_fill()
                
                pass  # Supprimer le message de configuration vidéo
            except Exception as e:
                self.log_message(f"Erreur configuration lecteur: {e}", 'DEBUG')
    
    def create_fallback_header(self, header_layout):
        """Crée un en-tête de secours si la vidéo n'est pas disponible"""
        title_label = QLabel("CPI ANALYZER TOOL")
        title_label.setStyleSheet("""
            QLabel {
                color: white;
                font-size: 36px;
                font-weight: bold;
                background: transparent;
            }
        """)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        subtitle_label = QLabel("Interface")
        subtitle_label.setStyleSheet("""
            QLabel {
                color: #e0e7ff;
                font-size: 18px;
                background: transparent;
                margin-top: 10px;
            }
        """)
        subtitle_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        header_layout.addWidget(title_label)
        header_layout.addWidget(subtitle_label)
    
    def on_media_status_changed(self, status):
        """Gère les changements de statut du média avec boucle de 3 à 7 secondes"""
        from PyQt6.QtMultimedia import QMediaPlayer
        
        if status == QMediaPlayer.MediaStatus.EndOfMedia:
            # La vidéo est terminée - revenir à la seconde 7
            self.media_player.setPosition(7000)  # 7000ms = 7 secondes
            self.media_player.play()
            self.log_message("Vidéo terminée - Retour à la seconde 7", 'INFO')
                
        elif status == QMediaPlayer.MediaStatus.LoadedMedia:
            pass  # Supprimer le message de chargement vidéo
            
        elif status == QMediaPlayer.MediaStatus.InvalidMedia:
            self.log_message("Média invalide - Utilisation de l'en-tête de secours", 'ERROR')
            # Remplacer par l'en-tête de secours
            parent = self.video_widget.parent()
            if parent:
                layout = parent.layout()
                if layout:
                    layout.removeWidget(self.video_widget)
                    self.video_widget.deleteLater()
                    self.create_fallback_header(layout)
    
    def update_boomerang_position(self):
        """Met à jour la position pour la lecture inversée manuelle"""
        if hasattr(self, 'media_player') and self.media_player:
            if not self.boomerang_forward:
                # Lecture inversée : reculer de 40ms (~1 frame à 25fps)
                self.boomerang_position -= 40
                if self.boomerang_position <= 0:
                    # Retour au début : repasser en lecture normale
                    self.boomerang_forward = True
                    self.boomerang_timer.stop()
                    self.media_player.setPosition(0)
                    self.media_player.setPlaybackRate(1.0)
                    self.media_player.play()
                    self.log_message("Effet boomerang : Fin lecture inversée", 'INFO')
                else:
                    # Continuer lecture inversée
                    self.media_player.setPosition(self.boomerang_position)
    
    def create_info_section(self, parent_layout):
        """Crée la section d'informations"""
        info_group = QGroupBox("📊 Tableau de bord")
        info_group.setStyleSheet("""
            QGroupBox {
                font-size: 16px;
                font-weight: bold;
                color: #e2e8f0;
                border: 2px solid #1ecce8;
                border-radius: 8px;
                margin-top: 5px;
                padding-top: 5px;
                background-color: #010001;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: #60a5fa;
            }
        """)
        
        info_layout = QGridLayout(info_group)
        info_layout.setContentsMargins(8, 8, 8, 8)  # Marges réduites
        info_layout.setSpacing(4)  # Espacement réduit entre les éléments
        
        # Labels d'information - première ligne (fichiers et taille)
        self.lbl_nb_fichiers = QLabel("Nbr de fichiers CPI : 0")
        self.lbl_taille_totale = QLabel("Taille Totale : 0 MB")
        
        # Labels d'information - deuxième ligne (statistiques)
        self.lbl_total_lignes = QLabel("Nbr Total des lignes : 0")
        self.lbl_nb_paiements = QLabel("Nbr Total des Paiements : 0")
        self.lbl_nb_rejets = QLabel("Nbr Total des Rejets : 0")
        
        # Labels d'information - instruments de paiement
        self.lbl_cheques = QLabel("Nbr Opr Chèques : 0")
        self.lbl_effets = QLabel("Nbr Opr Effets : 0")
        self.lbl_virements = QLabel("Nbr Opr Virements : 0")
        self.lbl_monetiques = QLabel("Nbr Opr Monétiques : 0")
        self.lbl_prelevements = QLabel("Nbr Opr Prélèvements : 0")
        
        # Histogramme des paiements/rejets
        self.histogram_widget = HistogramWidget(width=220, height=100)
        
        # Jauge de progression pour le taux de rapprochement
        self.progress_jauge = CircularProgress(size=100)
        self.progress_jauge.set_value(0, animate=False)  # Initialiser à 0%
        
        # Jauge de ventilation style compteur automobile
        self.ventilation_jauge = VentilationOdometerWidget(width=160, height=80)
        self.ventilation_jauge.set_value(0, animate=False)  # Initialiser à 0%
        
        # État du filtrage
        self.filtre_actif = None  # None, "Chèque", "Virement", "Effet commercial", "Monétique", "Prélèvement liaison"
        
        # Variables pour l'animation des chiffres
        self.chiffres_animation_timer = QTimer()
        self.chiffres_animation_timer.timeout.connect(self.animer_chiffres)
        self.animation_en_cours = False
        self.valeurs_cibles = {}
        self.valeurs_actuelles = {}
        self.animation_steps = 20
        self.animation_step_actuel = 0
        
        # Style pour tous les labels
        for label in [self.lbl_nb_fichiers, self.lbl_taille_totale, 
                     self.lbl_total_lignes, self.lbl_nb_paiements, self.lbl_nb_rejets]:
            label.setStyleSheet("""
                QLabel {
                    font-size: 14px;
                    color: #e2e8f0;
                    padding: 4px;
                    background-color: #010001;
                    border-radius: 6px;
                    border: none;
                }
            """)
        
        # Style pour les labels d'instruments de paiement (cliquables)
        for label in [self.lbl_cheques, self.lbl_virements, self.lbl_effets, 
                     self.lbl_monetiques, self.lbl_prelevements]:
            label.setStyleSheet("""
                QLabel {
                    font-size: 14px;
                    color: #e2e8f0;
                    padding: 4px;
                    background-color: #010001;
                    border-radius: 6px;
                    border: none;
                }
                QLabel:hover {
                    background-color: #010001;
                    color: #4a9eff;
                }
                QLabel.active {
                    background-color: #1e40af;
                    color: #ffffff;
                }
            """)
        
        # PREMIER CÔTÉ - Colonne de gauche
        
        # Ligne 1 - Nbr de fichiers CPI
        info_layout.addWidget(self.lbl_nb_fichiers, 0, 0)
        
        # Ligne 2 - Taille Totale
        info_layout.addWidget(self.lbl_taille_totale, 1, 0)
        
        # Ligne 3 - Nbr Total des lignes
        info_layout.addWidget(self.lbl_total_lignes, 2, 0)
        
        # Ligne 4 - Nbr Total des Paiements
        info_layout.addWidget(self.lbl_nb_paiements, 3, 0)
        
        # Ligne 5 - Nbr Total des Rejets
        info_layout.addWidget(self.lbl_nb_rejets, 4, 0)
        
        # DEUXIÈME CÔTÉ - Colonne de droite
        
        # Ligne 1 - Nbr Opr Chèques
        info_layout.addWidget(self.lbl_cheques, 0, 1)
        
        # Ligne 2 - Nbr Opr Effets
        info_layout.addWidget(self.lbl_effets, 1, 1)
        
        # Ligne 3 - Nbr Opr Virements
        info_layout.addWidget(self.lbl_virements, 2, 1)
        
        # Ligne 4 - Nbr Opr Monétiques
        info_layout.addWidget(self.lbl_monetiques, 3, 1)
        
        # Ligne 5 - Nbr Opr Prélèvements
        info_layout.addWidget(self.lbl_prelevements, 4, 1)
        
        # QUATRIÈME COLONNE - Histogramme des paiements/rejets
        # Créer un conteneur pour centrer l'histogramme verticalement
        histogram_container = QWidget()
        histogram_layout = QVBoxLayout(histogram_container)
        histogram_layout.setContentsMargins(0, 10, 0, 0)  # Marge en haut pour aligner
        histogram_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Ajouter un label au-dessus de l'histogramme
        histogram_label = QLabel("Taux de Paiements & Rejets CPI")
        histogram_label.setStyleSheet("""
            QLabel {
                font-size: 14px;
                color: #60a5fa;
                font-weight: bold;
                padding: 2px;
            }
        """)
        histogram_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        histogram_layout.addWidget(histogram_label)
        
        # Ajouter l'histogramme
        histogram_layout.addWidget(self.histogram_widget)
        
        # Ajouter le conteneur de l'histogramme à la quatrième colonne sur toutes les lignes
        info_layout.addWidget(histogram_container, 0, 2, 5, 1)  # (row, col, rowSpan, colSpan)
        
        # CINQUIÈME COLONNE - Jauge d'exhaustivité
        # Créer un conteneur pour centrer la jauge verticalement
        jauge_container = QWidget()
        jauge_layout = QVBoxLayout(jauge_container)
        jauge_layout.setContentsMargins(0, 10, 0, 0)  # Marge en haut pour aligner
        jauge_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Ajouter un label au-dessus de la jauge
        jauge_label = QLabel("Taux d'exhaustivité CPI")
        jauge_label.setStyleSheet("""
            QLabel {
                font-size: 14px;
                color: #60a5fa;
                font-weight: bold;
                padding: 2px;
            }
        """)
        jauge_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        jauge_layout.addWidget(jauge_label)
        
        # Ajouter la jauge
        jauge_layout.addWidget(self.progress_jauge)
        
        # Ajouter le conteneur de la jauge à la cinquième colonne sur toutes les lignes
        info_layout.addWidget(jauge_container, 0, 3, 5, 1)  # (row, col, rowSpan, colSpan)
        
        # SIXIÈME COLONNE - Jauge de ventilation
        # Créer un conteneur pour centrer la jauge verticalement
        ventilation_container = QWidget()
        ventilation_layout = QVBoxLayout(ventilation_container)
        ventilation_layout.setContentsMargins(0, 10, 0, 0)  # Marge en haut pour aligner
        ventilation_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Ajouter un label au-dessus de la jauge
        ventilation_label = QLabel("Taux de ventilation")
        ventilation_label.setStyleSheet("""
            QLabel {
                font-size: 14px;
                color: #60a5fa;
                font-weight: bold;
                padding: 2px;
            }
        """)
        ventilation_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ventilation_layout.addWidget(ventilation_label)
        
        # Ajouter la jauge de ventilation
        ventilation_layout.addWidget(self.ventilation_jauge)
        
        # Ajouter le conteneur de la ventilation à la sixième colonne sur toutes les lignes
        info_layout.addWidget(ventilation_container, 0, 4, 5, 1)  # (row, col, rowSpan, colSpan)
        
        parent_layout.addWidget(info_group)
        
        # Connecter les événements de clic aux labels d'instruments de paiement
        self.lbl_cheques.instrument = "Chèque"
        self.lbl_virements.instrument = "Virement"
        self.lbl_effets.instrument = "Effet commercial"
        self.lbl_monetiques.instrument = "Monétique"
        self.lbl_prelevements.instrument = "Prélèvement liaison"
        
        # Utiliser eventFilter pour une meilleure gestion des clics
        for label in [self.lbl_cheques, self.lbl_virements, self.lbl_effets, self.lbl_monetiques, self.lbl_prelevements]:
            label.installEventFilter(self)
    
    def appliquer_filtre(self, instrument):
        """Applique ou retire un filtre par instrument de paiement"""
        if self.filtre_actif == instrument:
            # Retirer le filtre
            self.filtre_actif = None
        else:
            # Appliquer le nouveau filtre
            self.filtre_actif = instrument
        
        # Mettre à jour l'apparence des labels
        self.mettre_a_jour_apparence_filtres()
        
        # Mettre à jour les statistiques si les données sont disponibles
        if hasattr(self, 'df_traite') and self.df_traite is not None:
            self.mettre_a_jour_statistiques_filtrees()
        else:
            self.log_message("Aucune donnée traitée disponible pour le filtrage", 'WARNING')
            # Mettre à jour quand même l'apparence pour montrer que le clic fonctionne
            self.mettre_a_jour_apparence_filtres()
    
    def mettre_a_jour_apparence_filtres(self):
        """Met à jour l'apparence des labels de filtre"""
        labels_map = {
            "Chèque": self.lbl_cheques,
            "Virement": self.lbl_virements,
            "Effet commercial": self.lbl_effets,
            "Monétique": self.lbl_monetiques,
            "Prélèvement liaison": self.lbl_prelevements
        }
        
        for instrument, label in labels_map.items():
            if self.filtre_actif == instrument:
                label.setProperty("class", "active")
                # Appliquer le style actif avec marges réduites
                label.setStyleSheet("""
                    QLabel {
                        font-size: 14px;
                        color: #4a9eff;
                        padding: 4px;
                        background-color: #010001;
                        border-radius: 6px;
                        border: none;
                    }
                    QLabel:hover {
                        background-color: #010001;
                        color: #4a9eff;
                    }
                """)
            else:
                label.setProperty("class", "")
                # Réinitialiser le style normal avec marges réduites
                label.setStyleSheet("""
                    QLabel {
                        font-size: 14px;
                        color: #e2e8f0;
                        padding: 4px;
                        background-color: #010001;
                        border-radius: 6px;
                        border: none;
                    }
                    QLabel:hover {
                        background-color: #010001;
                        color: #4a9eff;
                    }
                """)
    
    def mettre_a_jour_statistiques_filtrees(self):
        """Met à jour les statistiques du tableau de bord selon le filtre actif avec animation"""
        try:
            if not hasattr(self, 'df_traite') or self.df_traite is None:
                return
            
            # Filtrer les données selon l'instrument sélectionné
            if self.filtre_actif:
                df_filtre = self.df_traite.filter(pl.col("INST PAIEMENT") == self.filtre_actif)
            else:
                df_filtre = self.df_traite
            
            # Calculer les statistiques sur les données filtrées
            total_lignes = df_filtre.height
            
            # Statistiques de paiements/rejets
            colonnes_disponibles = df_filtre.columns
            has_statut = "Statut" in colonnes_disponibles
            
            if has_statut:
                stats_statut = df_filtre.group_by("Statut").agg(
                    pl.len().alias("count")
                )
                dict_statut = {row[0]: row[1] for row in stats_statut.iter_rows()}
                nb_paiements = dict_statut.get('Paiement', 0)
                nb_rejets = dict_statut.get('Rejet', 0)
            else:
                nb_paiements = 0
                nb_rejets = 0
            
            # Définir les valeurs cibles pour l'animation
            self.valeurs_cibles = {
                'total_lignes': total_lignes,
                'nb_paiements': nb_paiements,
                'nb_rejets': nb_rejets
            }
            
            # Initialiser les valeurs actuelles avec les valeurs actuelles des labels
            # Extraire les valeurs actuelles depuis les labels
            try:
                current_total_text = self.lbl_total_lignes.text().replace("Nbr Total des lignes : ", "").replace(",", "")
                current_paiements_text = self.lbl_nb_paiements.text().replace("Nbr Total des Paiements : ", "").replace(",", "")
                current_rejets_text = self.lbl_nb_rejets.text().replace("Nbr Total des Rejets : ", "").replace(",", "")
                
                current_total = int(current_total_text) if current_total_text.isdigit() else 0
                current_paiements = int(current_paiements_text) if current_paiements_text.isdigit() else 0
                current_rejets = int(current_rejets_text) if current_rejets_text.isdigit() else 0
                
                self.valeurs_actuelles = {
                    'total_lignes': current_total,
                    'nb_paiements': current_paiements,
                    'nb_rejets': current_rejets
                }
            except:
                # Si extraction échoue, utiliser les valeurs cibles comme point de départ
                self.valeurs_actuelles = self.valeurs_cibles.copy()
            
                        
            # Démarrer l'animation
            self.demarrer_animation()
            
            # Mettre à jour l'histogramme avec les nouvelles statistiques
            self.update_histogram(nb_paiements, nb_rejets, total_lignes)
            
            # Mettre à jour les instruments de paiement (toujours sur toutes les données)
            if "INST PAIEMENT" in self.df_traite.columns:
                stats_inst = self.df_traite.group_by("INST PAIEMENT").agg(
                    pl.len().alias("count")
                )
                dict_inst = {row[0]: row[1] for row in stats_inst.iter_rows()}
                nb_cheques = dict_inst.get("Chèque", 0)
                nb_virements = dict_inst.get("Virement", 0)
                nb_effets = dict_inst.get("Effet commercial", 0)
                nb_monetiques = dict_inst.get("Monétique", 0)
                nb_prelevements = dict_inst.get("Prélèvement liaison", 0)
            else:
                nb_cheques = nb_virements = nb_effets = nb_monetiques = nb_prelevements = 0
            
            self.lbl_cheques.setText(f"Nbr Opr Chèques : {nb_cheques:,}")
            self.lbl_virements.setText(f"Nbr Opr Virements : {nb_virements:,}")
            self.lbl_effets.setText(f"Nbr Opr Effets : {nb_effets:,}")
            self.lbl_monetiques.setText(f"Nbr Opr Monétiques : {nb_monetiques:,}")
            self.lbl_prelevements.setText(f"Nbr Opr Prélèvements : {nb_prelevements:,}")
            
        except Exception as e:
            self.log_message(f"Erreur mise à jour statistiques filtrées: {str(e)}", 'ERROR')
    
    def demarrer_animation(self):
        """Démarre l'animation des chiffres"""
        if self.animation_en_cours:
            self.chiffres_animation_timer.stop()
        
        self.animation_step_actuel = 0
        self.animation_en_cours = True
        
        # Utiliser un timer répétitif pour l'animation fluide
        self.chiffres_animation_timer.start(30)  # 30ms entre chaque étape
    
    def animer_chiffres(self):
        """Anime les chiffres vers les valeurs cibles"""
        if self.animation_step_actuel >= self.animation_steps:
            self.chiffres_animation_timer.stop()
            self.animation_en_cours = False
            self.valeurs_actuelles = self.valeurs_cibles.copy()
            return
        
        # Calculer le progrès de l'animation (0.0 à 1.0)
        progress = self.animation_step_actuel / self.animation_steps
        
        # Fonction d'animation (ease-in-out pour un effet naturel)
        if progress < 0.5:
            eased_progress = 2 * progress * progress
        else:
            eased_progress = 1 - pow(-2 * progress + 2, 2) / 2
        
        # Interpoler les valeurs actuelles vers les valeurs cibles
        for key, valeur_cible in self.valeurs_cibles.items():
            valeur_actuelle = self.valeurs_actuelles.get(key, valeur_cible)
            valeur_animee = valeur_actuelle + (valeur_cible - valeur_actuelle) * eased_progress
            self.valeurs_actuelles[key] = valeur_animee
        
        # Mettre à jour les labels avec les valeurs animées
        total_lignes_anime = int(self.valeurs_actuelles['total_lignes'])
        paiements_anime = int(self.valeurs_actuelles['nb_paiements'])
        rejets_anime = int(self.valeurs_actuelles['nb_rejets'])
        
        self.lbl_total_lignes.setText(f"Nbr Total des lignes : {total_lignes_anime:,}")
        self.lbl_nb_paiements.setText(f"Nbr Total des Paiements : {paiements_anime:,}")
        self.lbl_nb_rejets.setText(f"Nbr Total des Rejets : {rejets_anime:,}")
        
        self.animation_step_actuel += 1
    
    def update_histogram(self, nb_paiements=None, nb_rejets=None, total_operations=None):
        """Met à jour l'histogramme des paiements/rejets"""
        try:
            if nb_paiements is None or nb_rejets is None or total_operations is None:
                # Utiliser les valeurs actuelles des labels
                try:
                    paiement_text = self.lbl_nb_paiements.text().replace("Nbr Total des Paiements : ", "").replace(",", "")
                    rejet_text = self.lbl_nb_rejets.text().replace("Nbr Total des Rejets : ", "").replace(",", "")
                    total_text = self.lbl_total_lignes.text().replace("Nbr Total des lignes : ", "").replace(",", "")
                    
                    nb_paiements = int(paiement_text) if paiement_text.isdigit() else 0
                    nb_rejets = int(rejet_text) if rejet_text.isdigit() else 0
                    total_operations = int(total_text) if total_text.isdigit() else 0
                except:
                    nb_paiements = nb_rejets = total_operations = 0
            
            # Calculer les pourcentages
            if total_operations > 0:
                paiement_percent = (nb_paiements / total_operations) * 100
                rejet_percent = (nb_rejets / total_operations) * 100
            else:
                paiement_percent = rejet_percent = 0.0
            
            # Mettre à jour l'histogramme avec animation
            if hasattr(self, 'histogram_widget') and self.histogram_widget:
                self.histogram_widget.set_values(paiement_percent, rejet_percent, animate=True)
            
        except Exception as e:
            self.log_message(f"Erreur mise à jour histogramme: {str(e)}", 'ERROR')
    
    def update_progress_jauge(self, taux_rapprochement=None):
        """Met à jour la jauge de progression avec le taux de rapprochement"""
        try:
            # Si aucun taux n'est fourni, le calculer depuis les données de rapprochement
            if taux_rapprochement is None:
                if hasattr(self, 'rapprochement_data') and self.rapprochement_data is not None:
                    # Calculer le taux exactement comme dans l'onglet Exhaustivité
                    total_rapprochement = self.rapprochement_data.height
                    rapprochees = self.rapprochement_data.filter(pl.col('NCP').is_not_null()).height
                    taux_rapprochement = (rapprochees / total_rapprochement * 100) if total_rapprochement > 0 else 0
                else:
                    # Si pas de données, mettre à 0
                    taux_rapprochement = 0
            
            # Arrêter l'animation en cours si elle existe
            if hasattr(self, 'progress_jauge') and self.progress_jauge:
                self.progress_jauge.animation_timer.stop()
                
                # Animer vers la nouvelle valeur
                self.progress_jauge.set_value(taux_rapprochement, animate=True)
                
                # Loguer la mise à jour avec le détail du calcul
                self.log_message(f"Taux d'exhaustivité : {taux_rapprochement:.1f}%", 'INFO')
                
        except Exception as e:
            self.log_message(f"❌ Erreur mise à jour jauge: {str(e)}", 'ERROR')
    
    def update_ventilation_jauge(self):
        """Met à jour la jauge de ventilation avec le taux de ventilation correcte"""
        try:
            # Utiliser le taux de ventilation correcte calculé lors du rapprochement BKHIS
            taux_ventilation = 0.0
            
            if hasattr(self, 'taux_ventilation_actuel'):
                taux_ventilation = self.taux_ventilation_actuel
                self.log_message(f"🌪️ Taux de ventilation correcte: {taux_ventilation:.1f}%", 'INFO')
            else:
                self.log_message("⚠️ Taux de ventilation non disponible, utilisation de 0%", 'WARNING')
            
            # Mettre à jour la jauge de ventilation
            if hasattr(self, 'ventilation_jauge') and self.ventilation_jauge:
                self.ventilation_jauge.animation_timer.stop()
                self.ventilation_jauge.set_value(taux_ventilation, animate=True)
                
        except Exception as e:
            self.log_message(f"❌ Erreur mise à jour jauge ventilation: {str(e)}", 'ERROR')
    
    def eventFilter(self, obj, event):
        """Gère les événements de clic sur les labels d'instruments"""
        from PyQt6.QtCore import QEvent
        
        if event.type() == QEvent.Type.MouseButtonPress:
            if hasattr(obj, 'instrument'):
                self.appliquer_filtre(obj.instrument)
                return True
        return super().eventFilter(obj, event)
    
    def create_buttons_section(self, parent_layout):
        """Crée la section des boutons"""
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)
        
        # Boutons modernes principaux
        self.btn_charger = ModernButton("CHAR.FICH.CPI", "#10b981", "#059669")
        self.btn_traitement = ModernButton("CALC.SOLDES.CPI", "#3b82f6", "#2563eb")
        self.btn_ident_rejets = ModernButton("IDENT.REJETS ET PAIEMENTS", "#1ecce8", "#1bb8d4")
        self.btn_exporter = ModernButton("Exporter", "#FFFAFA", "#F8F8FF")
        
        # Connecter les signaux des boutons principaux
        self.btn_charger.clicked.connect(self.charger_fichiers)
        self.btn_traitement.clicked.connect(self.lancer_traitement)
        self.btn_ident_rejets.clicked.connect(self.ident_rejets_paiements)
        self.btn_exporter.clicked.connect(self.exporter_donnees)
        
        # Désactiver les boutons principaux initialement
        self.btn_traitement.setEnabled(False)
        self.btn_ident_rejets.setEnabled(False)
        self.btn_exporter.setEnabled(False)
        
        buttons_layout.addWidget(self.btn_charger)
        buttons_layout.addWidget(self.btn_traitement)
        buttons_layout.addWidget(self.btn_ident_rejets)
        buttons_layout.addWidget(self.btn_exporter)
        
        parent_layout.addLayout(buttons_layout)
        
        # Créer une deuxième rangée pour les boutons de vérification
        verification_layout = QHBoxLayout()
        verification_layout.setSpacing(15)
        
        # Boutons de vérification (non opérationnels)
        self.btn_ver_exhaustivite = ModernButton("Ver.Exhaustivité", "#f59e0b", "#d97706")
        self.btn_ver_ventilation = ModernButton("Ver.Ventilation", "#8b5cf6", "#7c3aed")
        
        # AJOUT DU BOUTON RAPPROCH DELTA.CPI
        self.btn_rapproch_delta = ModernButton("RAPPROCH DELTA.CPI", "#06b6d4", "#0891b2")
        
        # AJOUT DU BOUTON DE RÉINITIALISATION
        self.btn_reset = ModernButton("Réinitialiser", "#ef4444", "#dc2626")
        
        # Connecter les signaux des boutons de vérification
        self.btn_ver_exhaustivite.clicked.connect(self.verifier_exhaustivite)
        self.btn_ver_ventilation.clicked.connect(self.verifier_ventilation)
        self.btn_rapproch_delta.clicked.connect(self.rapprochement_delta_cpi)
        self.btn_reset.clicked.connect(self.reinitialiser_programme)
        
        # Désactiver les boutons de vérification (uniquement VER.VENTilation reste désactivé)
        self.btn_ver_exhaustivite.setEnabled(True)  # Activé pour charger BKHIS
        self.btn_ver_ventilation.setEnabled(True)  # Activé pour la fonctionnalité de ventilation
        self.btn_rapproch_delta.setEnabled(True)  # Activé dès le départ
        
        verification_layout.addWidget(self.btn_ver_exhaustivite)
        verification_layout.addWidget(self.btn_ver_ventilation)
        verification_layout.addWidget(self.btn_rapproch_delta)  # Ajouter le bouton RAPPROCH DELTA.CPI
        verification_layout.addWidget(self.btn_reset)  # Ajouter le bouton de réinitialisation
        
        parent_layout.addLayout(verification_layout)
    
    def create_log_section(self, parent_layout):
        """Crée la section de log"""
        log_group = QGroupBox("📝 Terminal - Journal d'activités")
        log_group.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        log_group.setStyleSheet("""
            QGroupBox {
                font-size: 16px;
                font-weight: bold;
                color: #e2e8f0;
                border: 2px solid #1ecce8;
                border-radius: 8px;
                margin-top: 5px;
                padding-top: 5px;
                padding-bottom: 1px;
                background-color: #010001;
            }
        """)
        
        log_layout = QVBoxLayout(log_group)
        log_layout.setContentsMargins(8, 8, 8, 0)  # Marge inférieure supprimée
        
        # Ajouter directement le ModernTextEdit (avec sa propre scrollbar)
        self.log_text = ModernTextEdit()
        self.log_text.setViewportMargins(0, 0, 0, 0)
        self.log_text.setStyleSheet(self.log_text.styleSheet() + "\nQTextEdit { margin-bottom: 0px; }")
        self.log_text.setMinimumHeight(120)  # Hauteur minimum
        self.log_text.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        
        log_layout.addWidget(self.log_text)
        
        parent_layout.addWidget(log_group)
    
    def apply_modern_style(self):
        """Applique le style moderne à l'application"""
        self.setStyleSheet("""
            QMainWindow {
                background-color: #010001;
            }
            QWidget {
                font-family: 'Segoe UI', Arial, sans-serif;
                background: transparent;
            }
            /* Tenter de contrôler la barre de titre sur Windows */
            QTitleBar {
                background-color: #010001;
                color: #e2e8f0;
            }
            QGroupBox {
                background-color: #010001;
                color: #e2e8f0;
                border: 2px solid #334155;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: #60a5fa;
            }
            QLabel {
                color: #e2e8f0;
                background: transparent;
            }
            QProgressBar {
                border: 2px solid #334155;
                border-radius: 6px;
                text-align: center;
                font-weight: bold;
                height: 25px;
                background-color: #010001;
                color: #e2e8f0;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3b82f6, stop:0.5 #60a5fa, stop:1 #3b82f6);
                border-radius: 6px;
            }
        """)
    
    def log_message(self, message, level='INFO'):
        """Ajoute un message au journal avec couleurs"""
        timestamp = time.strftime("%H:%M:%S", time.localtime())
        
        # Couleurs selon le niveau
        colors = {
            'INFO': '#10b981',
            'SUCCESS': '#10b981',
            'WARNING': '#f59e0b',
            'ERROR': '#ef4444',
            'DEBUG': '#6b7280'
        }
        
        color = colors.get(level.upper(), '#10b981')
        
        formatted_message = f'<span style="color: #6b7280;">[{timestamp}]</span> <span style="color: {color}; font-weight: bold;">[{level}]</span> <span style="color: #e5e7eb;">{message}</span><br>'
        
        # Vérifier si log_text existe avant de l'utiliser
        if hasattr(self, 'log_text') and self.log_text:
            self.log_text.insertHtml(formatted_message)
            self.log_text.ensureCursorVisible()
        
        # Mettre à jour la barre de status
        if hasattr(self, 'status_bar') and self.status_bar:
            if level.upper() == 'ERROR':
                self.status_bar.showMessage(f"Erreur: {message}")
            elif level.upper() == 'SUCCESS':
                self.status_bar.showMessage(f"{message}")
            else:
                self.status_bar.showMessage(f"{message}")
        
        # Afficher dans la console pour les erreurs
        if level.upper() == 'ERROR':
            print(f"ERREUR: {message}")
    
    def reinitialiser_programme(self):
        """Réinitialise le programme en conservant les 3 premiers messages du journal"""
        try:
            # 🔥 CONSERVER LES 3 PREMIERS MESSAGES DU JOURNAL
            if hasattr(self, 'log_text') and self.log_text:
                # 🔥 APPROCHE: Utiliser le texte brut au lieu du HTML
                texte_brut = self.log_text.toPlainText()
                
                # Extraire les 3 premiers messages du texte brut
                lignes_texte = texte_brut.split('\n')
                premieres_lignes_html = []
                messages_conserves = 0
                
                for ligne in lignes_texte:
                    ligne_propre = ligne.strip()
                    if ligne_propre and ('[INFO]' in ligne_propre or '[SUCCESS]' in ligne_propre or '[ERROR]' in ligne_propre or '[DEBUG]' in ligne_propre or '[WARNING]' in ligne_propre):
                        # Vérifier si c'est un message avec timestamp
                        if '[' in ligne_propre and ']' in ligne_propre:
                            # Convertir la ligne texte en format HTML
                            timestamp_part = ligne_propre.split(']')[0] + ']'
                            level_part = '[' + ligne_propre.split('[')[2].split(']')[0] + ']'
                            message_part = ']'.join(ligne_propre.split(']')[2:])
                            
                            # Recréer le format HTML
                            colors = {'INFO': '#10b981', 'SUCCESS': '#10b981', 'WARNING': '#f59e0b', 'ERROR': '#ef4444', 'DEBUG': '#6b7280'}
                            level = level_part.strip('[]')
                            color = colors.get(level, '#10b981')
                            
                            ligne_html = f'<span style="color: #6b7280;">{timestamp_part}</span> <span style="color: {color}; font-weight: bold;">{level_part}</span> <span style="color: #e5e7eb;">{message_part}</span>'
                            premieres_lignes_html.append(ligne_html)
                            messages_conserves += 1
                            if messages_conserves >= 4:
                                break
                
                # Vider complètement le journal
                self.log_text.clear()
                
                # Restaurer uniquement les 3 premiers messages
                if premieres_lignes_html:
                    for ligne in premieres_lignes_html:
                        self.log_text.insertHtml(ligne + '<br>')
                
                # PAS DE MESSAGE DE RÉINITIALISATION (nettoyage silencieux)
            
            # RÉINITIALISER LES DONNÉES
            self.df_consolidated = None
            self.df_traite = None
            self.df_resume = None
            self.df_solde_par_date = None
            self.df_bkhis = None
            self.rapprochement_data = None
            self.file_path = None
            
            # RÉINITIALISER L'INTERFACE
            if hasattr(self, 'lbl_nb_fichiers'):
                self.lbl_nb_fichiers.setText("Nbr de fichiers CPI : 0")
            if hasattr(self, 'lbl_taille_totale'):
                self.lbl_taille_totale.setText("Taille Totale : 0 MB")
            if hasattr(self, 'lbl_total_lignes'):
                self.lbl_total_lignes.setText("Nbr Total des lignes : 0")
            if hasattr(self, 'lbl_nb_paiements'):
                self.lbl_nb_paiements.setText("Nbr Total des Paiements : 0")
            if hasattr(self, 'lbl_nb_rejets'):
                self.lbl_nb_rejets.setText("Nbr Total des Rejets : 0")
            if hasattr(self, 'lbl_cheques'):
                self.lbl_cheques.setText("Nbr Opr Chèques : 0")
            if hasattr(self, 'lbl_effets'):
                self.lbl_effets.setText("Nbr Opr Effets : 0")
            if hasattr(self, 'lbl_virements'):
                self.lbl_virements.setText("Nbr Opr Virements : 0")
            if hasattr(self, 'lbl_monetiques'):
                self.lbl_monetiques.setText("Nbr Opr Monétiques : 0")
            if hasattr(self, 'lbl_prelevements'):
                self.lbl_prelevements.setText("Nbr Opr Prélèvements : 0")
            
            # 🔥 RÉINITIALISER LES ÉTATS
            self.filtre_actif = None
            self.chargement_en_cours = False
            
            # 🔥 RÉACTIVER LES BOUTONS
            if hasattr(self, 'btn_traitement'):
                self.btn_traitement.setEnabled(False)
            if hasattr(self, 'btn_ident_rejets'):
                self.btn_ident_rejets.setEnabled(False)
            if hasattr(self, 'btn_exporter'):
                self.btn_exporter.setEnabled(False)
            if hasattr(self, 'btn_ver_exhaustivite'):
                self.btn_ver_exhaustivite.setEnabled(True)
            
            # 🔥 PAS DE MESSAGE FINAL (nettoyage silencieux)
            
            # 🔄 RÉINITIALISER LA JAUGE DE PROGRESSION
            if hasattr(self, 'progress_jauge') and self.progress_jauge:
                self.progress_jauge.set_value(0, animate=True)  # Animation vers 0%
            
            # 🔄 RÉINITIALISER LA JAUGE DE VENTILATION
            if hasattr(self, 'ventilation_jauge') and self.ventilation_jauge:
                self.ventilation_jauge.set_value(0, animate=True)  # Animation vers 0%
            
            # 🔄 RÉINITIALISER L'HISTOGRAMME
            if hasattr(self, 'histogram_widget') and self.histogram_widget:
                self.histogram_widget.set_values(0, 0, animate=True)  # Animation vers 0%
            
        except Exception as e:
            self.log_message(f"❌ Erreur lors de la réinitialisation: {str(e)}", 'ERROR')
    
    def rapprochement_delta_cpi(self):
        """Gère le rapprochement DELTA.CPI avec la boîte de dialogue Phase 3"""
        try:
            self.log_message("Ouverture de l'interface RAPPROCH DELTA.CPI...", 'INFO')
            
            # Importer le module Phase 3 (avec le nom exact du fichier)
            try:
                import importlib.util
                import sys
                
                # Ajouter le répertoire courant au chemin Python si nécessaire
                current_dir = os.path.dirname(os.path.abspath(__file__))
                if current_dir not in sys.path:
                    sys.path.insert(0, current_dir)
                
                # Importer le module avec son nom exact
                spec = importlib.util.spec_from_file_location("Phase3", os.path.join(current_dir, "Phase 3.py"))
                phase3_module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(phase3_module)
                
                # Récupérer la classe RapprochementDialog
                RapprochementDialog = getattr(phase3_module, 'RapprochementDialog')
                
                self.log_message("✅ Module Phase 3 importé avec succès", 'SUCCESS')
            except Exception as e:
                QMessageBox.critical(self, "Erreur", 
                    f"Impossible d'importer le module Phase 3:\n{str(e)}\n\n"
                    f"Vérifiez que le fichier 'Phase 3.py' existe dans le même répertoire.")
                return
            
            # Récupérer les données
            df_traite = getattr(self, 'df_traite', None)
            df_bkhis = getattr(self, 'df_bkhis', None)
            
            # Créer et afficher la boîte de dialogue
            dialog = RapprochementDialog(
                parent=self,
                df_traite=df_traite,
                df_bkhis=df_bkhis,
                log_callback=self.log_message
            )
            
            # Afficher la boîte de dialogue
            result = dialog.exec()
            
            if result == QDialog.DialogCode.Accepted:
                self.log_message("✅ Rapprochement DELTA.CPI terminé avec succès", 'SUCCESS')
            else:
                self.log_message("🚫 Rapprochement DELTA.CPI annulé", 'INFO')
            
        except Exception as e:
            self.log_message(f"❌ Erreur lors de l'ouverture RAPPROCH DELTA.CPI: {str(e)}", 'ERROR')
            QMessageBox.critical(self, "Erreur", f"Échec de l'ouverture:\n{str(e)}")
    
    def afficher_message_bienvenue(self):
        """Affiche le message de bienvenue"""
        messages = [
            ("Bienvenue dans le programme CPI ANALYZER. Solution dédiée à l'assainissement de la télécompensation bancaire. Efficiency, Accuracy & Control📈", 'INFO'),
            ("Project Lead & Mentor : Mr HADOUCHE Madjid", 'INFO'),
            ("Data Engineer : Mr HOUHOU Tarek", 'INFO'),
            ("Developer : Mr BOUMAZZA Abdelkader", 'INFO'),
        ]
        
        for message, level in messages:
            QTimer.singleShot(messages.index((message, level)) * 500, 
                            lambda m=message, l=level: self.log_message(m, l))
    
    def update_animations(self):
        """Met à jour les animations (placeholder pour futures animations)"""
        pass
    
    def charger_fichiers(self):
        """Charge les fichiers avec la configuration flexible"""
        if hasattr(self, 'chargement_en_cours') and self.chargement_en_cours:
            QMessageBox.warning(self, "Attention", "Chargement déjà en cours!")
            return
        
        dossier = QFileDialog.getExistingDirectory(
            self, 
            "Sélectionner le dossier des fichiers CPI",
            os.path.expanduser("~"),
            QFileDialog.Option.ShowDirsOnly
        )
        
        if not dossier:
            return
        
        # Récupérer la liste des fichiers pour affichage immédiat
        from pathlib import Path
        fichiers_txt = list(Path(dossier).glob("*.txt"))
        fichiers_dsv = list(Path(dossier).glob("*.dsv"))
        tous_fichiers = fichiers_txt + fichiers_dsv
        self.fichiers_cpi = [str(f) for f in tous_fichiers]
        
        # Mettre à jour les informations immédiatement
        nb_fichiers = len(self.fichiers_cpi)
        taille_mb = sum(os.path.getsize(f) for f in self.fichiers_cpi) / (1024 * 1024) if self.fichiers_cpi else 0
        
        self.lbl_nb_fichiers.setText(f"Fichiers: {nb_fichiers}")
        self.lbl_taille_totale.setText(f"Taille: {taille_mb:.1f}MB")
        
        # Initialiser les statistiques à 0 (seront mises à jour après traitement)
        self.lbl_total_lignes.setText("Nbr Total lignes : 0")
        self.lbl_cheques.setText("Chèques : 0")
        self.lbl_virements.setText("Virements : 0")
        self.lbl_effets.setText("Effets : 0")
        self.lbl_monetiques.setText("Monétiques : 0")
        self.lbl_prelevements.setText("Prélèvements : 0")
        self.lbl_nb_paiements.setText("Nbr de Paiements : 0")
        self.lbl_nb_rejets.setText("Nbr de Rejets : 0")
        
        # Lancer le thread de chargement
        self.worker_thread = WorkerThread("load", dossier)
        self.worker_thread.log_message.connect(self.log_message)
        self.worker_thread.finished_loading.connect(self.on_loading_finished)
        self.worker_thread.start()
    
    def on_loading_finished(self, df_consolidated):
        """Appelé quand le chargement est terminé"""
        self.df_consolidated = df_consolidated
        
        # Réactiver les boutons
        self.btn_charger.setEnabled(True)
        self.btn_traitement.setEnabled(True)
        self.btn_ident_rejets.setEnabled(True)
        self.btn_exporter.setEnabled(True)
        
        # Mettre à jour les informations
        nb_fichiers = len(self.fichiers_cpi) if hasattr(self, 'fichiers_cpi') else 0
        taille_mb = sum(os.path.getsize(f) for f in self.fichiers_cpi) / (1024 * 1024) if hasattr(self, 'fichiers_cpi') else 0
        
        self.lbl_nb_fichiers.setText(f"Fichiers: {nb_fichiers}")
        self.lbl_taille_totale.setText(f"Taille: {taille_mb:.1f}MB")
        
        self.log_message("Chargement terminé avec succès!", 'SUCCESS')
    
    def on_processing_finished(self, df_traite):
        """Appelé quand le traitement est terminé"""
        self.df_traite = df_traite
        
        # Mettre à jour les statistiques du tableau de bord
        self.mettre_a_jour_statistiques(df_traite)
        
        # Créer le résumé par fichier
        self.log_message("Création du résumé détaillé par fichier...", 'INFO')
        
        colonnes_disponibles = self.df_traite.columns
        has_sens = "SENS" in colonnes_disponibles
        has_statut = "Statut" in colonnes_disponibles
        
        resume_expr = [pl.len().alias("nombre_lignes")]
        
        if has_sens:
            resume_expr.extend([
                pl.col("SENS").filter(pl.col("SENS") == "D").len().alias("nombre_lignes_debit"),
                pl.col("SENS").filter(pl.col("SENS") == "C").len().alias("nombre_lignes_credit")
            ])
        
        # Créer le résumé
        self.df_resume = self.df_traite.group_by("source_file").agg(
            *resume_expr
        ).sort("source_file")
        
        # Mettre à jour les statistiques du tableau de bord (maintenant la colonne Statut existe)
        self.mettre_a_jour_statistiques(self.df_traite)
        
        # Calcul du solde par date de règlement
        self.log_message("Calcul du solde par date de règlement...", 'INFO')
        self.calculer_solde_par_date()
        
        # Afficher les statistiques globales
        if has_sens:
            total_debit = self.df_traite.filter(pl.col("SENS") == "D").height
            total_credit = self.df_traite.filter(pl.col("SENS") == "C").height
            self.log_message(f"Statistiques globales - Débit: {total_debit:,} lignes, Crédit: {total_credit:,} lignes", 'INFO')
        
        # Construire le message de succès
        message_succes = f"Traitement terminé!\n\n✅ Données nettoyées:\n- {self.df_traite.height:,} lignes traitées\n- {self.df_resume.height:,} fichiers résumés"
        
        if has_sens:
            message_succes += f"\n\n📊 Statistiques Débit/Crédit:\n- Débit: {total_debit:,} lignes\n- Crédit: {total_credit:,} lignes"
        
        # Afficher le message de succès
        QMessageBox.information(self, "Succès", message_succes)
        
        # Réactiver le bouton traitement
        self.btn_traitement.setEnabled(True)
        
        self.log_message("Traitement terminé avec succès!", 'SUCCESS')
    
    def mettre_a_jour_statistiques_base(self, df_traite):
        """Met à jour les statistiques de base dans le tableau de bord (avant lettrage)"""
        try:
            # Nombre total de lignes
            total_lignes = df_traite.height
            
            # Statistiques de débits/crédits
            colonnes_disponibles = df_traite.columns
            has_sens = "SENS" in colonnes_disponibles
            
            if has_sens:
                stats_sens = df_traite.group_by("SENS").agg(
                    pl.len().alias("count")
                )
                
                # Extraire les comptes
                dict_stats = {row[0]: row[1] for row in stats_sens.iter_rows()}
                total_debits = dict_stats.get('D', 0)
                total_credits = dict_stats.get('C', 0)
            else:
                total_debits = 0
                total_credits = 0
            
            # Mettre à jour les labels (paiements/rejets restent à 0 pour l'instant)
            self.lbl_total_lignes.setText(f"Nbr Total lignes : {total_lignes:,}")
            self.lbl_cheques.setText("Chèques : 0")
            self.lbl_virements.setText("Virements : 0")
            self.lbl_effets.setText("Effets : 0")
            self.lbl_monetiques.setText("Monétiques : 0")
            self.lbl_prelevements.setText("Prélèvements : 0")
            self.lbl_nb_paiements.setText(f"Nbr de Paiements : 0")
            self.lbl_nb_rejets.setText(f"Nbr de Rejets : 0")
            
            self.log_message("Statistiques de base du tableau de bord mises à jour", 'SUCCESS')
            
        except Exception as e:
            self.log_message(f"Erreur mise à jour statistiques de base: {str(e)}", 'ERROR')
    
    def mettre_a_jour_statistiques(self, df_traite):
        """Met à jour les statistiques complètes dans le tableau de bord (après lettrage)"""
        try:
            # Nombre total de lignes
            total_lignes = df_traite.height
            
            # Statistiques de débits/crédits
            colonnes_disponibles = df_traite.columns
            has_sens = "SENS" in colonnes_disponibles
            has_statut = "Statut" in colonnes_disponibles
            
            if has_sens:
                stats_sens = df_traite.group_by("SENS").agg(
                    pl.len().alias("count")
                )
                
                # Extraire les comptes
                dict_stats = {row[0]: row[1] for row in stats_sens.iter_rows()}
                total_debits = dict_stats.get('D', 0)
                total_credits = dict_stats.get('C', 0)
            else:
                total_debits = 0
                total_credits = 0
            
            # Statistiques de paiements/rejets (disponibles après lettrage)
            if has_statut:
                stats_statut = df_traite.group_by("Statut").agg(
                    pl.len().alias("count")
                )
                
                dict_statut = {row[0]: row[1] for row in stats_statut.iter_rows()}
                nb_paiements = dict_statut.get('Paiement', 0)
                nb_rejets = dict_statut.get('Rejet', 0)
            else:
                nb_paiements = 0
                nb_rejets = 0
            
            # Mettre à jour les labels
            self.lbl_total_lignes.setText(f"Nbr Total lignes : {total_lignes:,}")
            self.lbl_nb_paiements.setText(f"Nbr de Paiements : {nb_paiements:,}")
            self.lbl_nb_rejets.setText(f"Nbr de Rejets : {nb_rejets:,}")
            
            # Calculer les statistiques par instrument de paiement
            colonnes_disponibles = df_traite.columns
            has_inst_paiement = "INST PAIEMENT" in colonnes_disponibles
            
            if has_inst_paiement:
                stats_inst = df_traite.group_by("INST PAIEMENT").agg(
                    pl.len().alias("count")
                )
                
                dict_inst = {row[0]: row[1] for row in stats_inst.iter_rows()}
                nb_cheques = dict_inst.get("Chèque", 0)
                nb_virements = dict_inst.get("Virement", 0)
                nb_effets = dict_inst.get("Effet commercial", 0)
                nb_monetiques = dict_inst.get("Monétique", 0)
                nb_prelevements = dict_inst.get("Prélèvement liaison", 0)
            else:
                nb_cheques = 0
                nb_virements = 0
                nb_effets = 0
                nb_monetiques = 0
                nb_prelevements = 0
            
            self.lbl_cheques.setText(f"Chèques : {nb_cheques:,}")
            self.lbl_virements.setText(f"Virements : {nb_virements:,}")
            self.lbl_effets.setText(f"Effets : {nb_effets:,}")
            self.lbl_monetiques.setText(f"Monétiques : {nb_monetiques:,}")
            self.lbl_prelevements.setText(f"Prélèvements : {nb_prelevements:,}")
            
            # Mettre à jour l'histogramme avec les statistiques complètes
            self.update_histogram(nb_paiements, nb_rejets, total_lignes)
            
            self.log_message("Statistiques complètes du tableau de bord mises à jour", 'SUCCESS')
            
        except Exception as e:
            self.log_message(f"Erreur mise à jour statistiques: {str(e)}", 'ERROR')
    
    def lancer_traitement(self):
        """Lance le traitement des données (version flexible)"""
        if self.df_consolidated is None:
            QMessageBox.warning(self, "Attention", "Veuillez d'abord charger les fichiers!")
            return
        
        # Désactiver les boutons pendant le traitement
        self.btn_traitement.setEnabled(False)
        
        # Lancer le thread de traitement
        self.worker_thread = WorkerThread("process", self.df_consolidated)
        self.worker_thread.log_message.connect(self.log_message)
        self.worker_thread.finished_processing.connect(self.on_processing_finished)
        self.worker_thread.start()
    
    def calculer_solde_par_date(self):
        """Calcule le solde par date de règlement en arrière-plan"""
        try:
            # Vérifier si les colonnes nécessaires existent
            colonnes_disponibles = self.df_traite.columns
            has_datereglement = "DATEREGLEMENT" in colonnes_disponibles
            has_montantoperation = "MONTANTOPERATION" in colonnes_disponibles
            
            if not has_datereglement:
                self.log_message("⚠️ Colonne 'DATEREGLEMENT' non trouvée - calcul du solde par date non disponible", 'WARNING')
                self.df_solde_par_date = None
                return
            
            if not has_montantoperation:
                self.log_message("⚠️ Colonne 'MONTANTOPERATION' non trouvée - calcul du solde par date non disponible", 'WARNING')
                self.df_solde_par_date = None
                return
            
            # Convertir MONTANTOPERATION en numérique (nettoyer les espaces, remplacer virgule par point) et grouper par DATEREGLEMENT et source_file
            self.df_solde_par_date = self.df_traite.with_columns(
                pl.col("MONTANTOPERATION")
                .str.replace_all(r"^\s+|\s+$", "")  # Supprimer les espaces au début et à la fin
                .str.replace_all(",", ".")  # Remplacer la virgule par le point
                .cast(pl.Float64)  # Convertir en numérique
            ).group_by(["DATEREGLEMENT", "source_file"]).agg(
                pl.col("MONTANTOPERATION").sum().alias("solde_total")
            ).sort(["DATEREGLEMENT", "source_file"]).with_columns(
                # Ajouter la colonne Sens selon la logique: si solde_total négatif -> D, sinon -> C
                pl.when(pl.col("solde_total") < 0)
                .then(pl.lit("D"))
                .otherwise(pl.lit("C"))
                .alias("Sens")
            )
            
            # Afficher les statistiques du calcul
            nb_dates = self.df_solde_par_date.height
            solde_global = self.df_solde_par_date.select(pl.col("solde_total").sum()).item()
            
            self.log_message(f"Calcul du solde terminé: {nb_dates:,} dates uniques", 'SUCCESS')
            self.log_message(f"Solde global: {solde_global:,.2f}", 'SUCCESS')
            
        except Exception as e:
            self.log_message(f"Erreur lors du calcul du solde par date: {str(e)}", 'ERROR')
            self.df_solde_par_date = None
    
    def ident_rejets_paiements(self):
        """Lettrage des opérations pour identifier les rejets et paiements avec sélection"""
        if self.df_traite is None:
            self.log_message("Veuillez d'abord lancer le traitement des données!", 'ERROR')
            return
        
        # Vérifier si les colonnes nécessaires existent pour le lettrage
        colonnes_necessaires = ["source_file", "INST PAIEMENT", "MONTANTS", "RIBTIRE", 
                                  "RIBBENEFICIAIRE", "NUM_DOCUMENT", "SENS", "DEBIT", "CREDIT"]
        
        colonnes_manquantes = [col for col in colonnes_necessaires if col not in self.df_traite.columns]
        
        if colonnes_manquantes:
            self.log_message(f"Colonnes manquantes pour le lettrage: {', '.join(colonnes_manquantes)}", 'ERROR')
            self.log_message("Veuillez d'abord lancer le traitement des données!", 'ERROR')
            return
        
        # Afficher la boîte de dialogue de sélection
        self.afficher_dialogue_selection_lettrage()
    
    def afficher_dialogue_selection_lettrage(self):
        """Affiche une boîte de dialogue pour sélectionner les instruments et statuts"""
        # Récupérer les instruments de paiement disponibles
        instruments_disponibles = []
        if "INST PAIEMENT" in self.df_traite.columns:
            instruments_disponibles = self.df_traite["INST PAIEMENT"].unique().to_list()
            instruments_disponibles = [str(inst).strip() for inst in instruments_disponibles if inst is not None and str(inst).strip() != ""]
            instruments_disponibles = sorted(list(set(instruments_disponibles)))
        
        if not instruments_disponibles:
            QMessageBox.warning(self, "Attention", "Aucun instrument de paiement trouvé dans les données!")
            return
        
        # Créer la boîte de dialogue
        dialog = QDialog(self)
        dialog.setWindowTitle("Sélection des critères de lettrage")
        dialog.setFixedSize(500, 400)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #010001;
                color: #e2e8f0;
            }
            QLabel {
                color: #e2e8f0;
                font-size: 14px;
                padding: 5px;
            }
            QGroupBox {
                color: #e2e8f0;
                border: 2px solid #1ecce8;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                background-color: #010001;
            }
            QGroupBox::title {
                color: #60a5fa;
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
            QListWidget {
                background-color: #010001;
                color: #e2e8f0;
                border: 1px solid #1ecce8;
                border-radius: 6px;
                padding: 5px;
            }
            QListWidget::item {
                padding: 3px;
                border-radius: 3px;
            }
            QListWidget::item:selected {
                background-color: #1ecce8;
                color: #010001;
            }
            QPushButton {
                background-color: #1ecce8;
                color: #010001;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #1bb8d4;
            }
            QPushButton:pressed {
                background-color: #1498a8;
            }
            QPushButton:disabled {
                background-color: #6b7280;
                color: #9ca3af;
            }
        """)
        
        layout = QVBoxLayout(dialog)
        
        # Titre
        title_label = QLabel("Sélectionner les critères pour le lettrage")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #60a5fa; padding: 15px;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # Section des instruments de paiement
        instruments_group = QGroupBox("📋 Instruments de paiement")
        instruments_layout = QVBoxLayout(instruments_group)
        
        instruments_list = QListWidget()
        instruments_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        
        # Ajouter les instruments disponibles
        for instrument in instruments_disponibles:
            item = QListWidgetItem(instrument)
            item.setSelected(True)  # Sélectionner tous par défaut
            instruments_list.addItem(item)
        
        instruments_layout.addWidget(instruments_list)
        layout.addWidget(instruments_group)
        
        # Section des natures d'opération CPI
        nature_group = QGroupBox("🏷️ Nature OPE CPI")
        nature_layout = QVBoxLayout(nature_group)
        
        # Radio buttons pour la nature d'opération
        self.radio_tous = QRadioButton("Toutes les natures (RETOUR + ALLER)")
        self.radio_retour = QRadioButton("Uniquement les RETOUR")
        self.radio_aller = QRadioButton("Uniquement les ALLER")
        
        self.radio_tous.setChecked(True)  # Par défaut
        
        nature_layout.addWidget(self.radio_tous)
        nature_layout.addWidget(self.radio_retour)
        nature_layout.addWidget(self.radio_aller)
        
        layout.addWidget(nature_group)
        
        # Boutons d'action
        buttons_layout = QHBoxLayout()
        
        lancer_button = QPushButton("Lancer le lettrage")
        annuler_button = QPushButton("Annuler")
        
        lancer_button.clicked.connect(lambda: self.lancer_lettrage_filtre(dialog, instruments_list))
        annuler_button.clicked.connect(dialog.reject)
        
        buttons_layout.addStretch()
        buttons_layout.addWidget(lancer_button)
        buttons_layout.addWidget(annuler_button)
        
        layout.addLayout(buttons_layout)
        
        # Afficher la boîte de dialogue
        dialog.exec()
    
    def lancer_lettrage_filtre(self, dialog, instruments_list):
        """Lance le lettrage avec les filtres sélectionnés"""
        # Récupérer les instruments sélectionnés
        selected_instruments = [instruments_list.item(i).text() for i in range(instruments_list.count()) 
                              if instruments_list.item(i).isSelected()]
        
        if not selected_instruments:
            QMessageBox.warning(dialog, "Attention", "Veuillez sélectionner au moins un instrument de paiement!")
            return
        
        # Récupérer la nature d'opération sélectionnée
        nature_filtre = "tous"
        if self.radio_retour.isChecked():
            nature_filtre = "RETOUR"
        elif self.radio_aller.isChecked():
            nature_filtre = "ALLER"
        
        # Filtrer les données selon les critères
        df_filtre = self.df_traite.filter(pl.col("INST PAIEMENT").is_in(selected_instruments))
        
        # Appliquer le filtre sur la nature d'opération si nécessaire
        if nature_filtre != "tous":
            if "NATURE OPE CPI" in df_filtre.columns:
                df_filtre = df_filtre.filter(pl.col("NATURE OPE CPI") == nature_filtre)
                self.log_message(f"Filtre NATURE OPE CPI appliqué: {nature_filtre}", 'INFO')
            else:
                self.log_message("⚠️ Colonne 'NATURE OPE CPI' non trouvée - filtre ignoré", 'WARNING')
        
        self.log_message(f"Lancement du lettrage avec filtres:", 'INFO')
        self.log_message(f"  Instruments: {', '.join(selected_instruments)}", 'INFO')
        self.log_message(f"  Nature OPE CPI: {nature_filtre}", 'INFO')
        self.log_message(f"  Lignes à traiter: {df_filtre.height:,}", 'INFO')
        
        # Fermer la boîte de dialogue
        dialog.accept()
        
        # Désactiver les boutons pendant le lettrage
        self.btn_ident_rejets.setEnabled(False)
        
        # Lancer le thread de lettrage avec les données filtrées et la nature
        self.worker_thread = WorkerThread("lettrage_filtre", df_filtre, nature_filtre)
        self.worker_thread.log_message.connect(self.log_message)
        self.worker_thread.finished_lettrage.connect(self.on_lettrage_finished_filtre)
        self.worker_thread.start()
    
    def on_lettrage_finished_filtre(self, df_traite_lettre):
        """Appelé quand le lettrage avec filtres est terminé"""
        # Si c'est le premier lettrage, remplacer complètement les données
        if "groupe_lettrage_sans_sens" not in self.df_traite.columns:
            self.df_traite = df_traite_lettre
            self.log_message("Premier lettrage effectué - données remplies", 'INFO')
        else:
            # Fusionner les résultats avec les données existantes
            # Créer un mapping des statuts pour les lignes traitées
            mapping_statuts = df_traite_lettre.select([
                "groupe_lettrage_sans_sens", 
                "Statut"
            ]).unique()
            
            # Mettre à jour les statuts dans les données originales en utilisant le mapping
            for ligne in mapping_statuts.iter_rows():
                groupe, statut = ligne[0], ligne[1]
                self.df_traite = self.df_traite.with_columns(
                    pl.when(pl.col("groupe_lettrage_sans_sens") == groupe)
                    .then(pl.lit(statut))
                    .otherwise(pl.col("Statut"))
                    .alias("Statut")
                )
            self.log_message("Lettrage complémentaire effectué - données fusionnées", 'INFO')
        
        # Mettre à jour les statistiques du tableau de bord
        self.mettre_a_jour_statistiques(self.df_traite)
        
        # Afficher les statistiques finales
        stats_lettrage = self.df_traite.group_by("Statut").agg([
            pl.len().alias("nombre_operations")
        ]).sort("nombre_operations", descending=True)
        
        self.log_message("Statistiques finales de lettrage:", 'INFO')
        for row in stats_lettrage.iter_rows():
            statut, nb_ops = row[0], row[1]
            self.log_message(f"  {statut}: {nb_ops:,} opérations", 'INFO')
        
        # Afficher le message de succès
        QMessageBox.information(self, "Lettrage terminé", 
                          f"Lettrage des opérations terminé avec succès!\n\n"
                          f"Opérations traitées: {df_traite_lettre.height:,}\n"
                          f"Total dans la base: {self.df_traite.height:,}\n\n"
                          f"Consultez le journal pour les statistiques détaillées.")
        
        # Réactiver le bouton
        self.btn_ident_rejets.setEnabled(True)
    
    def on_lettrage_finished(self, df_traite_lettre):
        """Appelé quand le lettrage est terminé"""
        self.df_traite = df_traite_lettre
        
        # Mettre à jour les statistiques du tableau de bord (maintenant la colonne Statut existe)
        self.mettre_a_jour_statistiques(self.df_traite)
        
        # Afficher les statistiques finales
        stats_lettrage = self.df_traite.group_by("Statut").agg([
            pl.len().alias("nombre_operations")
        ]).sort("nombre_operations", descending=True)
        
        self.log_message("Statistiques finales de lettrage:", 'INFO')
        for row in stats_lettrage.iter_rows():
            statut, nb_ops = row[0], row[1]
            self.log_message(f"  {statut}: {nb_ops:,} opérations", 'INFO')
        
        # Afficher le message de succès
        QMessageBox.information(self, "Lettrage terminé", 
                          f"Lettrage des opérations terminé avec succès!\n\n"
                          f"Opérations traitées: {self.df_traite.height:,}\n\n"
                          f"Consultez le journal pour les statistiques détaillées.")
        
        # Réactiver le bouton
        self.btn_ident_rejets.setEnabled(True)
        
        self.log_message("Lettrage terminé avec succès!", 'SUCCESS')
    
    def exporter_donnees(self):
        """Exporter les données avec options avancées dans des onglets"""
        if self.df_consolidated is None and self.df_traite is None:
            QMessageBox.warning(self, "Attention", "Aucune donnée à exporter!")
            return
        
        # Déterminer quelles données exporter
        data_to_export = self.df_traite if self.df_traite is not None else self.df_consolidated
        data_name = "traitées" if self.df_traite is not None else "brutes"
        
        # Créer une boîte de dialogue personnalisée avec onglets
        dialog = QDialog(self)
        dialog.setWindowTitle("Exporter les données - Options avancées")
        dialog.setFixedSize(700, 750)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #010001;
                color: #e2e8f0;
            }
            QLabel {
                color: #e2e8f0;
                font-size: 14px;
                padding: 5px;
            }
            QGroupBox {
                color: #e2e8f0;
                border: 2px solid #1ecce8;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                background-color: #010001;
            }
            QGroupBox::title {
                color: #60a5fa;
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
            QListWidget {
                background-color: #010001;
                color: #e2e8f0;
                border: 1px solid #1ecce8;
                border-radius: 6px;
                padding: 5px;
            }
            /* Style scrollbar pour QListWidget - même style que le terminal */
            QListWidget QScrollBar:vertical {
                background-color: #010001;
                width: 12px;
                border-radius: 6px;
            }
            QListWidget QScrollBar::handle:vertical {
                background-color: #1ecce8;
                border-radius: 6px;
                min-height: 20px;
            }
            /* Supprimer les flèches haut et bas */
            QListWidget QScrollBar::add-line:vertical {
                height: 0px;
                width: 0px;
                background: none;
            }
            QListWidget QScrollBar::sub-line:vertical {
                height: 0px;
                width: 0px;
                background: none;
            }
            QListWidget::item {
                padding: 3px;
                border-radius: 3px;
            }
            QListWidget::item:selected {
                background-color: #1ecce8;
                color: #010001;
            }
            QComboBox {
                background-color: #010001;
                color: #e2e8f0;
                border: 2px solid #1ecce8;
                border-radius: 6px;
                padding: 5px;
                font-size: 14px;
            }
            QPushButton {
                background-color: #1ecce8;
                color: #010001;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #1bb8d4;
            }
            QPushButton:pressed {
                background-color: #1498a8;
            }
            QTabWidget::pane {
                border: 2px solid #1ecce8;
                border-radius: 8px;
                background-color: #010001;
            }
            QTabBar::tab {
                background-color: #010001;
                color: #e2e8f0;
                border: 2px solid #1ecce8;
                border-bottom: none;
                border-radius: 6px 6px 0 0;
                padding: 8px 16px;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                background-color: #1ecce8;
                color: #010001;
            }
            QTabBar::tab:hover {
                background-color: #1bb8d4;
                color: #010001;
            }
        """)
        
        layout = QVBoxLayout(dialog)
        
        # Informations sur les données
        info_label = QLabel(f"Données à exporter: {data_name} ({data_to_export.height:,} lignes)")
        info_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #60a5fa; padding: 10px;")
        layout.addWidget(info_label)
        
        # Créer les onglets
        tab_widget = QTabWidget()
        
        # Onglet 1: Base de données
        tab1 = self._create_database_export_tab(data_to_export)
        tab_widget.addTab(tab1, "Base de données")
        
        # Onglet 2: Récapitulatif statistique
        tab2 = self._create_stats_export_tab(data_to_export)
        tab_widget.addTab(tab2, "Récapitulatif statistique")
        
        # Onglet 3: BKHIS DT
        if hasattr(self, 'df_bkhis') and self.df_bkhis is not None:
            tab3 = self._create_bkhis_export_tab()
            tab_widget.addTab(tab3, "BKHIS DT")
        
        # Onglet 4: Exhaustivité
        if hasattr(self, 'rapprochement_data') and self.rapprochement_data is not None:
            tab4 = self._create_exhaustivite_export_tab()
            tab_widget.addTab(tab4, "Exhaustivité")
        
        # Onglet 5: Ventilation (LEFT JOIN)
        tab5 = self._create_ventilation_export_tab()
        tab_widget.addTab(tab5, "Ventilation")
        
        layout.addWidget(tab_widget)
        
        # Boutons d'action
        buttons_layout = QHBoxLayout()
        
        export_button = QPushButton("Exporter")
        cancel_button = QPushButton("Annuler")
        
        buttons_layout.addWidget(export_button)
        buttons_layout.addWidget(cancel_button)
        layout.addLayout(buttons_layout)
        
        # Connecter les signaux
        export_button.clicked.connect(lambda: self._perform_tabbed_export(
            dialog, tab_widget, data_to_export, data_name
        ))
        
        cancel_button.clicked.connect(dialog.reject)
        
        # Afficher la boîte de dialogue
        dialog.exec()
    
    def _create_database_export_tab(self, data_to_export):
        """Créer l'onglet d'export de base de données"""
        tab_widget = QWidget()
        layout = QVBoxLayout(tab_widget)
        
        # Section format d'export
        format_group = QGroupBox("Format d'export")
        format_layout = QVBoxLayout(format_group)
        
        format_combo = QComboBox()
        format_combo.addItems(["Parquet", "CSV", "TXT", "DSV", "Excel"])
        format_combo.setCurrentText("CSV")
        format_layout.addWidget(format_combo)
        layout.addWidget(format_group)
        
        # Section colonnes à exporter
        columns_group = QGroupBox("Colonnes à exporter")
        columns_layout = QVBoxLayout(columns_group)
        
        columns_list = QListWidget()
        columns_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        
        # Ajouter toutes les colonnes disponibles
        for col in data_to_export.columns:
            item = QListWidgetItem(col)
            item.setSelected(True)  # Sélectionner toutes par défaut
            columns_list.addItem(item)
        
        columns_layout.addWidget(columns_list)
        layout.addWidget(columns_group)
        
        # Section instruments de paiement
        if "INST PAIEMENT" in data_to_export.columns:
            instruments_group = QGroupBox("Instruments de paiement")
            instruments_layout = QVBoxLayout(instruments_group)
            
            instruments_list = QListWidget()
            instruments_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
            
            # Récupérer les instruments uniques depuis les données
            instruments = data_to_export["INST PAIEMENT"].unique().to_list()
            instruments = [str(inst).strip() for inst in instruments if inst is not None and str(inst).strip() != ""]
            
            # Trier les instruments pour un affichage cohérent
            instruments = sorted(list(set(instruments)))  # Éliminer les doublons et trier
            
            # Ajouter uniquement les instruments qui existent réellement
            for instrument in instruments:
                item = QListWidgetItem(instrument)
                item.setSelected(True)
                instruments_list.addItem(item)
            
            instruments_layout.addWidget(instruments_list)
            layout.addWidget(instruments_group)
        
        # Section statut
        if "Statut" in data_to_export.columns:
            status_group = QGroupBox("Statut")
            status_layout = QVBoxLayout(status_group)
            
            status_list = QListWidget()
            status_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
            
            # Ajouter les statuts possibles
            possible_statuses = ["Paiement", "Rejet"]
            for status in possible_statuses:
                item = QListWidgetItem(status)
                item.setSelected(True)
                status_list.addItem(item)
            
            status_layout.addWidget(status_list)
            layout.addWidget(status_group)
        
        # Stocker les widgets pour accès ultérieur
        tab_widget.format_combo = format_combo
        tab_widget.columns_list = columns_list
        tab_widget.instruments_list = instruments_list if "INST PAIEMENT" in data_to_export.columns else None
        tab_widget.status_list = status_list if "Statut" in data_to_export.columns else None
        
        return tab_widget
    
    def _create_stats_export_tab(self, data_to_export):
        """Créer l'onglet d'export de récapitulatif statistique"""
        tab_widget = QWidget()
        layout = QVBoxLayout(tab_widget)
        
        # Titre de l'onglet
        title_label = QLabel("📊 Export des Récapitulatifs Statistiques")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #60a5fa; padding: 10px;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # Description
        desc_label = QLabel("Exportez les récapitulatifs statistiques avec les feuilles P.Ver.Exhaustivité, P.Ver.Ventilation, Stat générale et Récap")
        desc_label.setStyleSheet("font-size: 14px; color: #e2e8f0; padding: 10px;")
        desc_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(desc_label)
        
        # Section options d'export
        options_group = QGroupBox("Options d'export")
        options_layout = QVBoxLayout(options_group)
        
        # Options de sélection individuelle des feuilles
        sheets_group = QGroupBox("Sélectionner les feuilles à exporter")
        sheets_layout = QVBoxLayout(sheets_group)
        
        # Checkbox pour P.Ver.Exhaustivité
        self.include_exhaustivite = QCheckBox("P.Ver.Exhaustivité")
        self.include_exhaustivite.setChecked(True)  # Par défaut inclus
        self.include_exhaustivite.setStyleSheet("font-size: 13px; color: #e2e8f0; padding: 5px;")
        sheets_layout.addWidget(self.include_exhaustivite)
        
        # Checkbox pour P.Ver.Ventilation  
        self.include_ventilation = QCheckBox("P.Ver.Ventilation")
        self.include_ventilation.setChecked(True)  # Par défaut inclus
        self.include_ventilation.setStyleSheet("font-size: 13px; color: #e2e8f0; padding: 5px;")
        sheets_layout.addWidget(self.include_ventilation)
        
        # Checkbox pour Stat générale
        self.include_stat_gen = QCheckBox("Stat générale")
        self.include_stat_gen.setChecked(True)  # Par défaut inclus
        self.include_stat_gen.setStyleSheet("font-size: 13px; color: #e2e8f0; padding: 5px;")
        sheets_layout.addWidget(self.include_stat_gen)
        
        # Checkbox pour Récap
        self.include_recap = QCheckBox("Récap")
        self.include_recap.setChecked(True)  # Par défaut inclus
        self.include_recap.setStyleSheet("font-size: 13px; color: #e2e8f0; padding: 5px;")
        sheets_layout.addWidget(self.include_recap)
        
        options_layout.addWidget(sheets_group)
        
        # Format d'export (uniquement Excel pour les statistiques)
        format_label = QLabel("Format: Excel (.xlsx) - Uniquement disponible pour les récapitulatifs statistiques")
        format_label.setStyleSheet("font-size: 14px; color: #1ecce8; padding: 10px;")
        options_layout.addWidget(format_label)
        
        layout.addWidget(options_group)
        
        # Section informations sur les données disponibles
        info_group = QGroupBox("Informations sur les données")
        info_layout = QVBoxLayout(info_group)
        
        # Vérifier les données disponibles
        has_solde_data = hasattr(self, 'df_solde_par_date') and self.df_solde_par_date is not None
        has_traite_data = hasattr(self, 'df_traite') and self.df_traite is not None
        
        solde_info = QLabel(f"• P.Ver.Exhaustivité: {'✅ Disponible' if has_solde_data else '❌ Non disponible'}")
        solde_info.setStyleSheet("font-size: 13px; color: #28a745 if has_solde_data else #dc3545; padding: 5px;")
        info_layout.addWidget(solde_info)
        
        ventilation_info = QLabel(f"• P.Ver.Ventilation: {'✅ Disponible' if has_traite_data else '❌ Non disponible'}")
        ventilation_info.setStyleSheet("font-size: 13px; color: #28a745 if has_traite_data else #dc3545; padding: 5px;")
        info_layout.addWidget(ventilation_info)
        
        stat_gen_info = QLabel(f"• Stat générale: {'✅ Disponible' if has_traite_data else '❌ Non disponible'}")
        stat_gen_info.setStyleSheet("font-size: 13px; color: #28a745 if has_traite_data else #dc3545; padding: 5px;")
        info_layout.addWidget(stat_gen_info)
        
        recap_info = QLabel(f"• Récap: {'✅ Disponible' if has_traite_data else '❌ Non disponible'}")
        recap_info.setStyleSheet("font-size: 13px; color: #28a745 if has_traite_data else #dc3545; padding: 5px;")
        info_layout.addWidget(recap_info)
        
        if has_solde_data:
            solde_count_info = QLabel(f"  → {self.df_solde_par_date.height:,} dates de règlement")
            solde_count_info.setStyleSheet("font-size: 12px; color: #e2e8f0; padding-left: 20px;")
            info_layout.addWidget(solde_count_info)
        
        if has_traite_data:
            traite_count_info = QLabel(f"  → {self.df_traite.height:,} opérations à analyser")
            traite_count_info.setStyleSheet("font-size: 12px; color: #e2e8f0; padding-left: 20px;")
            info_layout.addWidget(traite_count_info)
        
        layout.addWidget(info_group)
        
        # Espace flexible
        layout.addStretch()
        
        return tab_widget
    
    def _create_bkhis_export_tab(self):
        """Créer l'onglet d'export BKHIS DT avec options de lettrage"""
        tab_widget = QWidget()
        layout = QVBoxLayout(tab_widget)
        
        # Informations sur les données BKHIS
        info_group = QGroupBox("Informations BKHIS")
        info_layout = QVBoxLayout(info_group)
        
        info_label = QLabel(f"Données BKHIS disponibles: {self.df_bkhis.height:,} lignes × {self.df_bkhis.width} colonnes")
        info_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #60a5fa; padding: 5px;")
        info_layout.addWidget(info_label)
        
        # Vérifier si la colonne LETTRAGE existe
        has_lettrage = "LETTRAGE" in self.df_bkhis.columns
        if has_lettrage:
            # Compter les opérations lettrées
            nb_lettre = self.df_bkhis.select(pl.col('LETTRAGE').count()).item()
            taux_lettrage = (nb_lettre / self.df_bkhis.height * 100) if self.df_bkhis.height > 0 else 0
            lettrage_info = QLabel(f"Opérations lettrées: {nb_lettre:,} ({taux_lettrage:.1f}%)")
            lettrage_info.setStyleSheet("font-size: 12px; color: #10b981; padding-left: 20px;")
            info_layout.addWidget(lettrage_info)
        
        layout.addWidget(info_group)
        
        # Section format d'export
        format_group = QGroupBox("Format d'export")
        format_layout = QVBoxLayout(format_group)
        
        format_combo = QComboBox()
        format_combo.addItems(["Excel", "CSV", "TXT", "DSV", "Parquet"])
        format_combo.setCurrentText("Excel")
        format_layout.addWidget(format_combo)
        layout.addWidget(format_group)
        
        # Section filtre de lettrage
        if has_lettrage:
            lettrage_group = QGroupBox("Filtre de lettrage")
            lettrage_layout = QVBoxLayout(lettrage_group)
            
            lettrage_combo = QComboBox()
            lettrage_combo.addItems(["Toute la base", "Opérations lettrées uniquement", "Opérations non lettrées uniquement"])
            lettrage_combo.setCurrentText("Toute la base")
            lettrage_layout.addWidget(lettrage_combo)
            layout.addWidget(lettrage_group)
        else:
            lettrage_combo = None
        
        # Section colonnes à exporter
        columns_group = QGroupBox("Colonnes à exporter")
        columns_layout = QVBoxLayout(columns_group)
        
        columns_list = QListWidget()
        columns_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        
        # Ajouter toutes les colonnes disponibles
        for col in self.df_bkhis.columns:
            item = QListWidgetItem(col)
            item.setSelected(True)  # Sélectionner toutes par défaut
            columns_list.addItem(item)
        
        columns_layout.addWidget(columns_list)
        layout.addWidget(columns_group)
        
        # Section NCP (si disponible)
        if "NCP" in self.df_bkhis.columns:
            ncp_group = QGroupBox("Filtre par NCP")
            ncp_layout = QVBoxLayout(ncp_group)
            
            ncp_list = QListWidget()
            ncp_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
            
            # Récupérer les NCP uniques
            ncp_uniques = self.df_bkhis["NCP"].unique().to_list()
            ncp_uniques = [str(ncp).strip() for ncp in ncp_uniques if ncp is not None and str(ncp).strip() != ""]
            ncp_uniques = sorted(list(set(ncp_uniques)))  # Éliminer les doublons et trier
            
            # Limiter à 100 NCP pour éviter la surcharge
            if len(ncp_uniques) > 100:
                ncp_uniques = ncp_uniques[:100]
                info_ncp = QLabel(f"Affichage des 100 premiers NCP sur {len(self.df_bkhis['NCP'].unique().to_list())} au total")
                info_ncp.setStyleSheet("font-size: 11px; color: #f59e0b; padding: 5px;")
                ncp_layout.addWidget(info_ncp)
            
            # Ajouter les NCP
            for ncp in ncp_uniques:
                item = QListWidgetItem(ncp)
                item.setSelected(True)  # Sélectionner tous par défaut
                ncp_list.addItem(item)
            
            ncp_layout.addWidget(ncp_list)
            layout.addWidget(ncp_group)
        else:
            ncp_list = None
        
        # Espace flexible
        layout.addStretch()
        
        # Stocker les widgets pour accès ultérieur
        tab_widget.format_combo = format_combo
        tab_widget.lettrage_combo = lettrage_combo
        tab_widget.columns_list = columns_list
        tab_widget.ncp_list = ncp_list
        
        return tab_widget
    
    def _perform_tabbed_export(self, dialog, tab_widget, data_to_export, data_name):
        """Gérer l'export selon l'onglet sélectionné"""
        current_tab_index = tab_widget.currentIndex()
        
        if current_tab_index == 0:
            # Onglet Base de données - utiliser l'export existant
            current_tab = tab_widget.currentWidget()
            self._perform_filtered_export(
                dialog, 
                current_tab.format_combo, 
                current_tab.columns_list,
                current_tab.instruments_list,
                current_tab.status_list,
                data_to_export, 
                data_name
            )
        elif current_tab_index == 1:
            # Onglet Récapitulatif statistique - implémenter l'export
            self._perform_stats_export(dialog, data_to_export, data_name, 
                                       self.include_exhaustivite.isChecked(), 
                                       self.include_ventilation.isChecked(),
                                       self.include_stat_gen.isChecked(),
                                       self.include_recap.isChecked())
        elif current_tab_index == 2:
            # Onglet BKHIS DT - exporter les données BKHIS
            current_tab = tab_widget.currentWidget()
            self._perform_bkhis_export(dialog, current_tab)
        elif current_tab_index == 3:
            # Onglet Exhaustivité - exporter les résultats de rapprochement
            current_tab = tab_widget.currentWidget()
            self._perform_exhaustivite_export(dialog, current_tab)
        elif current_tab_index == 4:
            # Onglet Ventilation - effectuer le LEFT JOIN
            self._perform_ventilation_export(dialog)
    
    def _perform_ventilation_export(self, dialog):
        try:
            # Vérifier que les fichiers sont sélectionnés
            recap_file = self.recap_file_path.text().strip()
            exhaust_file = self.exhaust_file_path.text().strip()
            
            if not recap_file:
                QMessageBox.warning(self, "Attention", "Veuillez sélectionner le fichier Récapitulatif statistique!")
                return
            
            if not exhaust_file:
                QMessageBox.warning(self, "Attention", "Veuillez sélectionner le fichier Exhaustivité!")
                return
            
            self.log_message("Début de la ventilation - LEFT JOIN...", 'INFO')
            self.log_message(f"Fichier Récapitulatif: {recap_file}", 'INFO')
            self.log_message(f"Fichier Exhaustivité: {exhaust_file}", 'INFO')
            
            # Charger les données avec pandas (nécessaire pour openpyxl)
            import pandas as pd
            
            # Charger la feuille Sld.CPI.calc.instru depuis le fichier récapitulatif
            try:
                df_recap = pd.read_excel(recap_file, sheet_name="Sld.CPI.calc.instru")
                self.log_message(f"Feuille 'Sld.CPI.calc.instru' chargée: {len(df_recap)} lignes", 'SUCCESS')
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Impossible de charger la feuille 'Sld.CPI.calc.instru':\n{str(e)}")
                return
            
            # Charger la feuille Result Exhaustivité depuis le fichier exhaustivité
            try:
                df_exhaust = pd.read_excel(exhaust_file, sheet_name="Result Exhaustivité")
                self.log_message(f"Feuille 'Result Exhaustivité' chargée: {len(df_exhaust)} lignes", 'SUCCESS')
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Impossible de charger la feuille 'Result Exhaustivité':\n{str(e)}")
                return
            
            # Vérifier les colonnes nécessaires
            if "DATE REGLEMENT" not in df_recap.columns:
                QMessageBox.critical(self, "Erreur", "La colonne 'DATE REGLEMENT' n'existe pas dans la feuille 'Sld.CPI.calc.instru'")
                return
            
            if "DATE REGLEMENT CPI" not in df_exhaust.columns:
                QMessageBox.critical(self, "Erreur", "La colonne 'DATE REGLEMENT CPI' n'existe pas dans la feuille 'Result Exhaustivité'")
                return
            
            # Convertir les dates en string pour éviter les problèmes de format
            df_recap["DATE REGLEMENT"] = df_recap["DATE REGLEMENT"].astype(str)
            df_exhaust["DATE REGLEMENT CPI"] = df_exhaust["DATE REGLEMENT CPI"].astype(str)
            
            # Effectuer le LEFT JOIN
            self.log_message("Effectuation du LEFT JOIN...", 'INFO')
            df_ventilation = df_recap.merge(
                df_exhaust,
                left_on="DATE REGLEMENT",
                right_on="DATE REGLEMENT CPI",
                how="left",
                suffixes=("", "_exhaust")
            )
            
            self.log_message(f"LEFT JOIN terminé: {len(df_ventilation)} lignes résultantes", 'SUCCESS')
            
            # Ajouter la feuille Ventilation au fichier Excel d'exhaustivité
            try:
                from openpyxl import load_workbook
                
                # Charger le fichier Excel existant
                wb = load_workbook(exhaust_file)
                
                # Supprimer la feuille Ventilation si elle existe déjà
                if "Ventilation" in wb.sheetnames:
                    del wb["Ventilation"]
                    self.log_message("Ancienne feuille 'Ventilation' supprimée", 'INFO')
                
                # Créer la nouvelle feuille Ventilation
                ws_ventilation = wb.create_sheet("Ventilation")
                
                # Écrire les données
                for r_idx, row in enumerate(df_ventilation.itertuples(index=False), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws_ventilation.cell(row=r_idx, column=c_idx, value=value)
                
                # Mettre la couleur de l'onglet en orange
                ws_ventilation.sheet_properties.tabColor = "FF6600"
                
                # Sauvegarder le fichier
                wb.save(exhaust_file)
                
                self.log_message(f"Feuille 'Ventilation' créée avec succès dans: {exhaust_file}", 'SUCCESS')
                self.log_message(f"Nombre de lignes dans la ventilation: {len(df_ventilation)}", 'INFO')
                
                # Afficher les statistiques
                nb_match = len(df_ventilation[df_ventilation["DATE REGLEMENT CPI"].notna()])
                nb_no_match = len(df_ventilation[df_ventilation["DATE REGLEMENT CPI"].isna()])
                
                self.log_message(f"Statistiques du LEFT JOIN:", 'INFO')
                self.log_message(f"  - Correspondances trouvées: {nb_match}", 'INFO')
                self.log_message(f"  - Sans correspondance: {nb_no_match}", 'INFO')
                
                QMessageBox.information(self, "Succès", 
                    f"Ventilation effectuée avec succès!\n\n"
                    f"Fichier: {exhaust_file}\n"
                    f"Feuille créée: Ventilation\n"
                    f"Lignes totales: {len(df_ventilation)}\n"
                    f"Correspondances: {nb_match}\n"
                    f"Sans correspondance: {nb_no_match}")
                
                dialog.accept()
                
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Impossible de créer la feuille 'Ventilation':\n{str(e)}")
                return
            
        except Exception as e:
            self.log_message(f"Erreur lors de la ventilation: {str(e)}", 'ERROR')
            QMessageBox.critical(self, "Erreur", f"Échec de la ventilation:\n{str(e)}")
    
    def _perform_stats_export(self, dialog, data_to_export, data_name, include_exhaustivite, include_ventilation, include_stat_gen, include_recap):
        """Effectue l'export des récapitulatifs statistiques avec sélection individuelle"""
        try:
            # Vérifier les données nécessaires
            has_solde_data = hasattr(self, 'df_solde_par_date') and self.df_solde_par_date is not None
            has_traite_data = hasattr(self, 'df_traite') and self.df_traite is not None
            
            if not has_solde_data and not has_traite_data:
                QMessageBox.warning(self, "Attention", 
                    "Aucune donnée statistique disponible!\n\n"
                    "Veuillez d'abord lancer le traitement des données.")
                return
            
            # Demander le fichier de destination (uniquement Excel)
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Exporter les récapitulatifs statistiques",
                f"CPI_STATS_RECAPITULATIF_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Fichier Excel (*.xlsx)"
            )
            
            if not file_path:
                return
            
            self.log_message("Début de l'export des récapitulatifs statistiques...", 'INFO')
            
            # Afficher les feuilles sélectionnées
            self.log_message("Feuilles à exporter :", 'INFO')
            if include_exhaustivite:
                self.log_message("  ✓ P.Ver.Exhaustivité", 'INFO')
            if include_ventilation:
                self.log_message("  ✓ P.Ver.Ventilation", 'INFO')
            if include_stat_gen:
                self.log_message("  ✓ Stat générale", 'INFO')
            if include_recap:
                self.log_message("  ✓ Récap", 'INFO')
            
            # Importer les bibliothèques nécessaires
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Créer le workbook
            wb = openpyxl.Workbook()
            
            # Supprimer la feuille par défaut
            if wb.sheetnames:
                wb.remove(wb[wb.sheetnames[0]])
            
            # Ajouter la feuille Sld.CPI.calc.date si les données de solde existent
            if has_solde_data and include_exhaustivite:
                self.log_message("Création de la feuille Sld.CPI.calc.date...", 'INFO')
                ws_exhaustivite = wb.create_sheet("Sld.CPI.calc.date")
                
                # Ajouter les données du solde par date
                for r in dataframe_to_rows(self.df_solde_par_date.to_pandas(), index=False, header=True):
                    ws_exhaustivite.append(r)
                
                self.log_message(f"Feuille 'Sld.CPI.calc.date' ajoutée avec {self.df_solde_par_date.height:,} dates", 'SUCCESS')
            
            # Ajouter la feuille Sld.CPI.calc.instru si les données traitées existent
            if has_traite_data and include_ventilation:
                self.log_message("Création de la feuille Sld.CPI.calc.instru...", 'INFO')
                
                # Vérifier si les colonnes nécessaires existent pour la ventilation
                colonnes_disponibles = self.df_traite.columns
                has_datereglement = "DATEREGLEMENT" in colonnes_disponibles
                has_montantoperation = "MONTANTOPERATION" in colonnes_disponibles
                has_inst_paiement = "INST PAIEMENT" in colonnes_disponibles
                
                if has_datereglement and has_montantoperation and has_inst_paiement:
                    # Préparer les données pour la ventilation (adapté de main.py)
                    df_ventilation = self.df_traite.with_columns(
                        pl.col("MONTANTOPERATION")
                        .str.replace_all(r"^\s+|\s+$", "")  # Supprimer les espaces
                        .str.replace_all(",", ".")  # Remplacer virgule par point
                        .cast(pl.Float64)  # Convertir en numérique
                    ).group_by(["source_file", "DATEREGLEMENT", "INST PAIEMENT"]).agg(
                        pl.col("MONTANTOPERATION").sum().alias("SOLDES"),
                        pl.col("MONTANTOPERATION").len().alias("NOMBRE")  # Compter le nombre de montants
                    ).sort(["source_file", "DATEREGLEMENT", "INST PAIEMENT"]).rename({
                        "DATEREGLEMENT": "DATE REGLEMENT"
                    }).with_columns(
                        # Ajouter la colonne CHAPITRE selon INST PAIEMENT
                        pl.when(pl.col("INST PAIEMENT") == "Chèque")
                        .then(pl.lit("325210"))
                        .when(pl.col("INST PAIEMENT") == "Effet commercial")
                        .then(pl.lit("325240"))
                        .when(pl.col("INST PAIEMENT") == "Monétique")
                        .then(pl.lit("325550"))
                        .when(pl.col("INST PAIEMENT") == "Virement")
                        .then(pl.lit("325230"))
                        .when(pl.col("INST PAIEMENT") == "Prélèvement liaison")
                        .then(pl.lit("325250"))
                        .otherwise(pl.lit(""))
                        .alias("CHAPITRE"),
                        # Ajouter la colonne NCP selon INST PAIEMENT
                        pl.when(pl.col("INST PAIEMENT") == "Chèque")
                        .then(pl.lit("0004325210-67"))
                        .when(pl.col("INST PAIEMENT") == "Effet commercial")
                        .then(pl.lit("0004325240-74"))
                        .when(pl.col("INST PAIEMENT") == "Monétique")
                        .then(pl.lit("0004160550-26"))
                        .when(pl.col("INST PAIEMENT") == "Virement")
                        .then(pl.lit("0004325230-07"))
                        .when(pl.col("INST PAIEMENT") == "Prélèvement liaison")
                        .then(pl.lit("0004325250-44"))
                        .otherwise(pl.lit(""))
                        .alias("NCP")
                    ).with_columns(
                        # Ajouter la colonne SENS selon le signe de SOLDES
                        pl.when(pl.col("SOLDES") < 0)
                        .then(pl.lit("D"))
                        .otherwise(pl.lit("C"))
                        .alias("SENS")
                    )
                    
                    # Réorganiser les colonnes pour mettre NOMBRE entre INST PAIEMENT et SOLDES
                    df_ventilation = df_ventilation.select([
                        "source_file", 
                        "DATE REGLEMENT", 
                        "CHAPITRE", 
                        "NCP", 
                        "INST PAIEMENT", 
                        "NOMBRE",  # Colonne en majuscule
                        "SOLDES", 
                        "SENS"
                    ])
                    
                    # Créer la feuille Sld.CPI.calc.instru
                    ws_ventilation = wb.create_sheet("Sld.CPI.calc.instru")
                    
                    # Ajouter les données de ventilation
                    for r in dataframe_to_rows(df_ventilation.to_pandas(), index=False, header=True):
                        ws_ventilation.append(r)
                    
                    self.log_message(f"Feuille 'Sld.CPI.calc.instru' ajoutée avec {df_ventilation.height:,} ventilations", 'SUCCESS')
                    
                    # Afficher un résumé de la ventilation
                    total_general = df_ventilation.select(pl.col("SOLDES").sum()).item()
                    self.log_message(f"Total général ventilé: {total_general:,.2f}", 'INFO')
                    
                else:
                    self.log_message("⚠️ Colonnes manquantes pour Sld.CPI.calc.instru - feuille non créée", 'WARNING')
                    if not has_datereglement:
                        self.log_message("  - Colonne 'DATEREGLEMENT' manquante", 'WARNING')
                    if not has_montantoperation:
                        self.log_message("  - Colonne 'MONTANTOPERATION' manquante", 'WARNING')
                    if not has_inst_paiement:
                        self.log_message("  - Colonne 'INST PAIEMENT' manquante", 'WARNING')
            
            # Ajouter la feuille Stat générale si les données traitées existent et si sélectionnée
            if has_traite_data and include_stat_gen:
                self.log_message("Création de la feuille Stat générale...", 'INFO')
                
                # Vérifier les colonnes nécessaires pour Stat générale
                colonnes_disponibles = self.df_traite.columns
                has_datereglement = "DATEREGLEMENT" in colonnes_disponibles
                has_inst_paiement = "INST PAIEMENT" in colonnes_disponibles
                has_nature_ope = "NATURE OPE CPI" in colonnes_disponibles
                has_sens = "SENS" in colonnes_disponibles
                has_montantoperation = "MONTANTOPERATION" in colonnes_disponibles
                
                if has_datereglement and has_inst_paiement and has_sens and has_montantoperation:
                    # Préparer les données pour Stat générale
                    df_stat_gen = self.df_traite.with_columns(
                        pl.col("MONTANTOPERATION")
                        .str.replace_all(r"^\s+|\s+$", "")  # Supprimer les espaces
                        .str.replace_all(",", ".")  # Remplacer virgule par point
                        .cast(pl.Float64)  # Convertir en numérique
                    ).with_columns(
                        # Ajouter la colonne NCP selon INST PAIEMENT
                        pl.when(pl.col("INST PAIEMENT") == "Chèque")
                        .then(pl.lit("0004325210-67"))
                        .when(pl.col("INST PAIEMENT") == "Effet commercial")
                        .then(pl.lit("0004325240-74"))
                        .when(pl.col("INST PAIEMENT") == "Monétique")
                        .then(pl.lit("0004160550-26"))
                        .when(pl.col("INST PAIEMENT") == "Virement")
                        .then(pl.lit("0004325230-07"))
                        .when(pl.col("INST PAIEMENT") == "Prélèvement liaison")
                        .then(pl.lit("0004325250-44"))
                        .otherwise(pl.lit(""))
                        .alias("NCP")
                    )
                    
                    # Grouper par DATEREGLEMENT, INST PAIEMENT, NCP, NATURE OPE CPI
                    df_stat_gen = df_stat_gen.group_by([
                        "DATEREGLEMENT", 
                        "INST PAIEMENT", 
                        "NCP", 
                        "NATURE OPE CPI"
                    ]).agg(
                        # Compter les débits et crédits
                        pl.col("SENS").filter(pl.col("SENS") == "D").len().alias("Nbr Debit"),
                        pl.col("SENS").filter(pl.col("SENS") == "C").len().alias("Nbr Credit"),
                        # Sommer les débits et crédits
                        pl.col("MONTANTOPERATION").filter(pl.col("SENS") == "D").sum().alias("Val Debit"),
                        pl.col("MONTANTOPERATION").filter(pl.col("SENS") == "C").sum().alias("Val Credit")
                    ).with_columns(
                        # Calculer Nbr Total
                        (pl.col("Nbr Debit") + pl.col("Nbr Credit")).alias("Nbr Total"),
                        # Calculer Solde (Val Credit + Val Debit)
                        (pl.col("Val Credit") + pl.col("Val Debit")).alias("Solde")
                    ).sort(["DATEREGLEMENT", "INST PAIEMENT", "NATURE OPE CPI"])
                    
                    # Si NATURE OPE CPI n'existe pas, utiliser une valeur par défaut
                    if not has_nature_ope:
                        df_stat_gen = df_stat_gen.with_columns(
                            pl.lit("Non spécifié").alias("NATURE OPE CPI")
                        )
                    
                    # Renommer DATEREGLEMENT en "Date Reglement"
                    df_stat_gen = df_stat_gen.rename({
                        "DATEREGLEMENT": "Date Reglement"
                    })
                    
                    # Trier avec les nouveaux noms de colonnes
                    df_stat_gen = df_stat_gen.sort(["Date Reglement", "INST PAIEMENT", "NATURE OPE CPI"])
                    
                    # Ajouter la colonne SENS selon la logique : si Solde est négatif alors D sinon C
                    df_stat_gen = df_stat_gen.with_columns(
                        pl.when(pl.col("Solde") < 0)
                        .then(pl.lit("D"))
                        .otherwise(pl.lit("C"))
                        .alias("Sens")
                    )
                    
                    # Ajouter la colonne Observation avec la logique conditionnelle
                    df_stat_gen = df_stat_gen.with_columns(
                        pl.when(
                            (pl.col("NATURE OPE CPI") == "RETOUR") & (pl.col("Sens") == "C")
                        ).then(pl.lit("A Examiner"))
                        .when(
                            (pl.col("NATURE OPE CPI") == "ALLER") & (pl.col("Sens") == "D")
                        ).then(pl.lit("A Examiner"))
                        .otherwise(pl.lit("Rien a signaler"))
                        .alias("Observation")
                    )
                    
                    # Réorganiser les colonnes dans l'ordre demandé
                    df_stat_gen = df_stat_gen.select([
                        "Date Reglement",
                        "INST PAIEMENT", 
                        "NCP",
                        "NATURE OPE CPI",
                        "Nbr Debit",
                        "Nbr Credit",
                        "Nbr Total",
                        "Val Debit",
                        "Val Credit",
                        "Solde",
                        "Sens",
                        "Observation"
                    ])
                    
                    # Créer la feuille Stat générale
                    ws_stat_gen = wb.create_sheet("Stat générale")
                    
                    # Ajouter les données de Stat générale
                    for r in dataframe_to_rows(df_stat_gen.to_pandas(), index=False, header=True):
                        ws_stat_gen.append(r)
                    
                    # Appliquer le formatage en rouge pour les cellules "A Examiner"
                    from openpyxl.styles import Font, PatternFill
                    
                    red_font = Font(color="FF0000")  # Rouge
                    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Fond rouge clair
                    
                    # Trouver l'index de la colonne "Observation" (12ème colonne, donc index 11)
                    observation_col_index = 12  # Colonnes Excel commencent à 1
                    
                    # Parcourir toutes les lignes et formater les cellules "A Examiner"
                    for row in ws_stat_gen.iter_rows(min_row=2, max_row=ws_stat_gen.max_row, min_col=observation_col_index, max_col=observation_col_index):
                        for cell in row:
                            if cell.value == "A Examiner":
                                cell.font = red_font
                                cell.fill = red_fill
                    
                    self.log_message(f"Feuille 'Stat générale' ajoutée avec {df_stat_gen.height:,} lignes", 'SUCCESS')
                    
                    # Afficher un résumé de la Stat générale
                    total_debit_gen = df_stat_gen.select(pl.col("Val Debit").sum()).item()
                    total_credit_gen = df_stat_gen.select(pl.col("Val Credit").sum()).item()
                    solde_general = df_stat_gen.select(pl.col("Solde").sum()).item()
                    
                    self.log_message(f"Résumé Stat générale:", 'INFO')
                    self.log_message(f"  Total Débit: {total_debit_gen:,.2f}", 'INFO')
                    self.log_message(f"  Total Crédit: {total_credit_gen:,.2f}", 'INFO')
                    self.log_message(f"  Solde Général: {solde_general:,.2f}", 'INFO')
                    
                else:
                    self.log_message("⚠️ Colonnes manquantes pour Stat générale - feuille non créée", 'WARNING')
                    if not has_datereglement:
                        self.log_message("  - Colonne 'DATEREGLEMENT' manquante", 'WARNING')
                    if not has_inst_paiement:
                        self.log_message("  - Colonne 'INST PAIEMENT' manquante", 'WARNING')
                    if not has_sens:
                        self.log_message("  - Colonne 'SENS' manquante", 'WARNING')
                    if not has_montantoperation:
                        self.log_message("  - Colonne 'MONTANTOPERATION' manquante", 'WARNING')
            
            # Ajouter la feuille Récap si les données traitées existent et si sélectionnée
            if has_traite_data and include_recap:
                self.log_message("Création de la feuille Récap...", 'INFO')
                
                # Vérifier les colonnes nécessaires pour Récap
                colonnes_disponibles = self.df_traite.columns
                has_inst_paiement = "INST PAIEMENT" in colonnes_disponibles
                has_nature_ope = "NATURE OPE CPI" in colonnes_disponibles
                has_sens = "SENS" in colonnes_disponibles
                has_montantoperation = "MONTANTOPERATION" in colonnes_disponibles
                
                if has_inst_paiement and has_sens and has_montantoperation:
                    # Préparer les données pour Récap (similaire à Stat générale mais sans Date Reglement)
                    df_recap = self.df_traite.with_columns(
                        pl.col("MONTANTOPERATION")
                        .str.replace_all(r"^\s+|\s+$", "")  # Supprimer les espaces
                        .str.replace_all(",", ".")  # Remplacer virgule par point
                        .cast(pl.Float64)  # Convertir en numérique
                    ).with_columns(
                        # Ajouter la colonne NCP selon INST PAIEMENT
                        pl.when(pl.col("INST PAIEMENT") == "Chèque")
                        .then(pl.lit("0004325210-67"))
                        .when(pl.col("INST PAIEMENT") == "Effet commercial")
                        .then(pl.lit("0004325240-74"))
                        .when(pl.col("INST PAIEMENT") == "Monétique")
                        .then(pl.lit("0004160550-26"))
                        .when(pl.col("INST PAIEMENT") == "Virement")
                        .then(pl.lit("0004325230-07"))
                        .when(pl.col("INST PAIEMENT") == "Prélèvement liaison")
                        .then(pl.lit("0004325250-44"))
                        .otherwise(pl.lit(""))
                        .alias("NCP")
                    )
                    
                    # Grouper par INST PAIEMENT, NCP, NATURE OPE CPI (sans Date Reglement)
                    df_recap = df_recap.group_by([
                        "INST PAIEMENT", 
                        "NCP", 
                        "NATURE OPE CPI"
                    ]).agg(
                        # Compter les débits et crédits
                        pl.col("SENS").filter(pl.col("SENS") == "D").len().alias("Nbr Debit"),
                        pl.col("SENS").filter(pl.col("SENS") == "C").len().alias("Nbr Credit"),
                        # Sommer les débits et crédits
                        pl.col("MONTANTOPERATION").filter(pl.col("SENS") == "D").sum().alias("Val Debit"),
                        pl.col("MONTANTOPERATION").filter(pl.col("SENS") == "C").sum().alias("Val Credit")
                    ).with_columns(
                        # Calculer Nbr Total
                        (pl.col("Nbr Debit") + pl.col("Nbr Credit")).alias("Nbr Total"),
                        # Calculer Solde (Val Credit + Val Debit)
                        (pl.col("Val Credit") + pl.col("Val Debit")).alias("Solde")
                    ).with_columns(
                        # Ajouter la colonne SENS selon la logique : si Solde est négatif alors D sinon C
                        pl.when(pl.col("Solde") < 0)
                        .then(pl.lit("D"))
                        .otherwise(pl.lit("C"))
                        .alias("Sens")
                    ).with_columns(
                        # Ajouter la colonne Observation avec la logique conditionnelle
                        pl.when(
                            (pl.col("NATURE OPE CPI") == "RETOUR") & (pl.col("Sens") == "C")
                        ).then(pl.lit("A Examiner"))
                        .when(
                            (pl.col("NATURE OPE CPI") == "ALLER") & (pl.col("Sens") == "D")
                        ).then(pl.lit("A Examiner"))
                        .otherwise(pl.lit("Rien a signaler"))
                        .alias("Observation")
                    )
                    
                    # Si NATURE OPE CPI n'existe pas, utiliser une valeur par défaut
                    if not has_nature_ope:
                        df_recap = df_recap.with_columns(
                            pl.lit("Non spécifié").alias("NATURE OPE CPI")
                        )
                    
                    # Calculer le nombre d'opérations anormales à partir de la logique de Stat générale
                    # On crée d'abord les données comme dans Stat générale pour identifier les "A Examiner"
                    df_stat_gen_for_count = self.df_traite.with_columns(
                        pl.col("MONTANTOPERATION")
                        .str.replace_all(r"^\s+|\s+$", "")  # Supprimer les espaces
                        .str.replace_all(",", ".")  # Remplacer virgule par point
                        .cast(pl.Float64)  # Convertir en numérique
                    ).with_columns(
                        # Ajouter la colonne NCP selon INST PAIEMENT
                        pl.when(pl.col("INST PAIEMENT") == "Chèque")
                        .then(pl.lit("0004325210-67"))
                        .when(pl.col("INST PAIEMENT") == "Effet commercial")
                        .then(pl.lit("0004325240-74"))
                        .when(pl.col("INST PAIEMENT") == "Monétique")
                        .then(pl.lit("0004160550-26"))
                        .when(pl.col("INST PAIEMENT") == "Virement")
                        .then(pl.lit("0004325230-07"))
                        .when(pl.col("INST PAIEMENT") == "Prélèvement liaison")
                        .then(pl.lit("0004325250-44"))
                        .otherwise(pl.lit(""))
                        .alias("NCP")
                    )
                    
                    # Si NATURE OPE CPI n'existe pas, utiliser une valeur par défaut
                    if not has_nature_ope:
                        df_stat_gen_for_count = df_stat_gen_for_count.with_columns(
                            pl.lit("Non spécifié").alias("NATURE OPE CPI")
                        )
                    
                    # Grouper par Date Reglement, INST PAIEMENT, NCP, NATURE OPE CPI comme dans Stat générale
                    df_stat_gen_for_count = df_stat_gen_for_count.group_by([
                        "DATEREGLEMENT",
                        "INST PAIEMENT", 
                        "NCP", 
                        "NATURE OPE CPI"
                    ]).agg(
                        # Compter les débits et crédits
                        pl.col("SENS").filter(pl.col("SENS") == "D").len().alias("Nbr Debit"),
                        pl.col("SENS").filter(pl.col("SENS") == "C").len().alias("Nbr Credit"),
                        # Sommer les débits et crédits
                        pl.col("MONTANTOPERATION").filter(pl.col("SENS") == "D").sum().alias("Val Debit"),
                        pl.col("MONTANTOPERATION").filter(pl.col("SENS") == "C").sum().alias("Val Credit")
                    ).with_columns(
                        # Calculer Nbr Total
                        (pl.col("Nbr Debit") + pl.col("Nbr Credit")).alias("Nbr Total"),
                        # Calculer Solde (Val Credit + Val Debit)
                        (pl.col("Val Credit") + pl.col("Val Debit")).alias("Solde")
                    ).with_columns(
                        # Ajouter la colonne SENS selon la logique : si Solde est négatif alors D sinon C
                        pl.when(pl.col("Solde") < 0)
                        .then(pl.lit("D"))
                        .otherwise(pl.lit("C"))
                        .alias("Sens")
                    ).with_columns(
                        # Ajouter la colonne Observation avec la logique conditionnelle
                        pl.when(
                            (pl.col("NATURE OPE CPI") == "RETOUR") & (pl.col("Sens") == "C")
                        ).then(pl.lit("A Examiner"))
                        .when(
                            (pl.col("NATURE OPE CPI") == "ALLER") & (pl.col("Sens") == "D")
                        ).then(pl.lit("A Examiner"))
                        .otherwise(pl.lit("Rien a signaler"))
                        .alias("Observation")
                    )
                    
                    # Compter les opérations "A Examiner" par instrument de paiement ET par Nature ope CPI
                    df_count_anormales = df_stat_gen_for_count.filter(
                        pl.col("Observation") == "A Examiner"
                    ).group_by(["INST PAIEMENT", "NATURE OPE CPI"]).agg(
                        pl.len().alias("Nbr ope anormales")
                    )
                    
                    # Joindre les données récap avec le compte des opérations anormales par instrument et nature
                    df_recap = df_recap.join(
                        df_count_anormales,
                        on=["INST PAIEMENT", "NATURE OPE CPI"],
                        how="left"
                    ).with_columns(
                        # Remplacer les valeurs nulles par 0 pour les opérations anormales
                        pl.col("Nbr ope anormales").fill_null(0)
                    )
                    
                    # Réorganiser les colonnes dans l'ordre demandé (sans Date Reglement et Observation)
                    df_recap = df_recap.select([
                        "INST PAIEMENT", 
                        "NCP",
                        "NATURE OPE CPI",
                        "Nbr Debit",
                        "Nbr Credit",
                        "Nbr Total",
                        "Val Debit",
                        "Val Credit",
                        "Solde",
                        "Sens",
                        "Nbr ope anormales"
                    ]).sort(["INST PAIEMENT", "NATURE OPE CPI"])
                    
                    # Créer la feuille Récap
                    ws_recap = wb.create_sheet("Récap")
                    
                    # Ajouter les données de Récap (premier tableau)
                    for r in dataframe_to_rows(df_recap.to_pandas(), index=False, header=True):
                        ws_recap.append(r)
                    
                    # Ajouter une ligne vide pour séparer les tableaux
                    ws_recap.append([])
                    
                    # Ajouter un titre pour le deuxième tableau
                    ws_recap.append(["SITUATION DES PAIEMENTS & REJETS"])
                    ws_recap.append([])
                    
                    # Créer le deuxième tableau avec STATUT (Rejets et Paiements)
                    # Grouper par INST PAIEMENT, NCP, NATURE OPE CPI, STATUT
                    df_statut = self.df_traite.with_columns(
                        pl.col("MONTANTOPERATION")
                        .str.replace_all(r"^\s+|\s+$", "")  # Supprimer les espaces
                        .str.replace_all(",", ".")  # Remplacer virgule par point
                        .cast(pl.Float64)  # Convertir en numérique
                    ).with_columns(
                        # Ajouter la colonne NCP selon INST PAIEMENT
                        pl.when(pl.col("INST PAIEMENT") == "Chèque")
                        .then(pl.lit("0004325210-67"))
                        .when(pl.col("INST PAIEMENT") == "Effet commercial")
                        .then(pl.lit("0004325240-74"))
                        .when(pl.col("INST PAIEMENT") == "Monétique")
                        .then(pl.lit("0004160550-26"))
                        .when(pl.col("INST PAIEMENT") == "Virement")
                        .then(pl.lit("0004325230-07"))
                        .when(pl.col("INST PAIEMENT") == "Prélèvement liaison")
                        .then(pl.lit("0004325250-44"))
                        .otherwise(pl.lit(""))
                        .alias("NCP")
                    )
                    
                    # Si NATURE OPE CPI n'existe pas, utiliser une valeur par défaut
                    if not has_nature_ope:
                        df_statut = df_statut.with_columns(
                            pl.lit("Non spécifié").alias("NATURE OPE CPI")
                        )
                    
                    # Grouper par INST PAIEMENT, NCP, NATURE OPE CPI, STATUT
                    df_statut = df_statut.group_by([
                        "INST PAIEMENT", 
                        "NCP", 
                        "NATURE OPE CPI",
                        "Statut"
                    ]).agg(
                        # Compter les débits et crédits
                        pl.col("SENS").filter(pl.col("SENS") == "D").len().alias("Nbr Debit"),
                        pl.col("SENS").filter(pl.col("SENS") == "C").len().alias("Nbr Credit"),
                        # Sommer les débits et crédits
                        pl.col("MONTANTOPERATION").filter(pl.col("SENS") == "D").sum().alias("Val Debit"),
                        pl.col("MONTANTOPERATION").filter(pl.col("SENS") == "C").sum().alias("Val Credit")
                    ).with_columns(
                        # Calculer Nbr Total
                        (pl.col("Nbr Debit") + pl.col("Nbr Credit")).alias("Nbr Total"),
                        # Calculer Solde (Val Credit + Val Debit)
                        (pl.col("Val Credit") + pl.col("Val Debit")).alias("Solde")
                    ).with_columns(
                        # Ajouter la colonne SENS selon la logique : si Solde est négatif alors D sinon C, mais si Solde = 0 alors vide
                        pl.when(pl.col("Solde") == 0)
                        .then(pl.lit(""))
                        .when(pl.col("Solde") < 0)
                        .then(pl.lit("D"))
                        .otherwise(pl.lit("C"))
                        .alias("Sens")
                    ).sort(["INST PAIEMENT", "NCP", "NATURE OPE CPI", "Statut"])
                    
                    # Réorganiser les colonnes dans l'ordre demandé
                    df_statut = df_statut.select([
                        "INST PAIEMENT",
                        "NCP", 
                        "NATURE OPE CPI",
                        "Statut",
                        "Nbr Debit",
                        "Nbr Credit",
                        "Nbr Total",
                        "Val Debit",
                        "Val Credit",
                        "Solde",
                        "Sens"
                    ])
                    
                    # Ajouter les données du deuxième tableau
                    for r in dataframe_to_rows(df_statut.to_pandas(), index=False, header=True):
                        ws_recap.append(r)
                    
                    self.log_message(f"Feuille 'Récap' ajoutée avec {df_recap.height:,} lignes (premier tableau) et {df_statut.height:,} lignes (tableau statuts)", 'SUCCESS')
                    
                    # Afficher un résumé de la Récap
                    total_anormales = df_recap.select(pl.col("Nbr ope anormales").sum()).item()
                    self.log_message(f"Résumé Récap:", 'INFO')
                    self.log_message(f"  Total opérations anormales: {total_anormales:,}", 'INFO')
                    
                else:
                    self.log_message("⚠️ Colonnes manquantes pour Récap - feuille non créée", 'WARNING')
                    if not has_inst_paiement:
                        self.log_message("  - Colonne 'INST PAIEMENT' manquante", 'WARNING')
                    if not has_sens:
                        self.log_message("  - Colonne 'SENS' manquante", 'WARNING')
                    if not has_montantoperation:
                        self.log_message("  - Colonne 'MONTANTOPERATION' manquante", 'WARNING')
            
            # Sauvegarder le fichier
            wb.save(file_path)
            
            # Compter le nombre de feuilles créées
            nb_feuilles = len(wb.sheetnames)
            
            self.log_message(f"Export statistique réussi: {file_path}", 'SUCCESS')
            self.log_message(f"Feuilles créées: {', '.join(wb.sheetnames)} ({nb_feuilles} feuilles)", 'INFO')
            
            QMessageBox.information(self, "Succès", 
                f"Export des récapitulatifs statistiques réussi!\n\n"
                f"Fichier: {file_path}\n"
                f"Feuilles créées: {nb_feuilles}\n"
                f"{', '.join(wb.sheetnames)}")
            
            dialog.accept()
            
        except Exception as e:
            self.log_message(f"Erreur lors de l'export statistique: {str(e)}", 'ERROR')
            QMessageBox.critical(self, "Erreur", f"Échec de l'export statistique:\n{str(e)}")
    
    def _perform_bkhis_export(self, dialog, tab_widget):
        """Effectue l'export des données BKHIS avec filtres"""
        try:
            # Récupérer les paramètres
            format_export = tab_widget.format_combo.currentText()
            lettrage_filter = tab_widget.lettrage_combo.currentText() if tab_widget.lettrage_combo else "Toute la base"
            
            # Récupérer les colonnes sélectionnées
            selected_columns = [tab_widget.columns_list.item(i).text() for i in range(tab_widget.columns_list.count()) 
                               if tab_widget.columns_list.item(i).isSelected()]
            
            if not selected_columns:
                QMessageBox.warning(self, "Attention", "Veuillez sélectionner au moins une colonne à exporter!")
                return
            
            # Récupérer les NCP sélectionnés
            selected_ncp = []
            if tab_widget.ncp_list:
                selected_ncp = [tab_widget.ncp_list.item(i).text() for i in range(tab_widget.ncp_list.count()) 
                               if tab_widget.ncp_list.item(i).isSelected()]
            
            # Appliquer les filtres
            df_filtered = self.df_bkhis
            
            # Filtrer par lettrage
            if lettrage_filter == "Opérations lettrées uniquement":
                df_filtered = df_filtered.filter(pl.col('LETTRAGE') == 'L')
                self.log_message(f"Filtre: Opérations lettrées uniquement ({df_filtered.height:,} lignes)", 'INFO')
            elif lettrage_filter == "Opérations non lettrées uniquement":
                df_filtered = df_filtered.filter(pl.col('LETTRAGE').is_null() | (pl.col('LETTRAGE') != 'L'))
                self.log_message(f"Filtre: Opérations non lettrées uniquement ({df_filtered.height:,} lignes)", 'INFO')
            else:
                self.log_message(f"Filtre: Toute la base ({df_filtered.height:,} lignes)", 'INFO')
            
            # Filtrer par NCP
            if selected_ncp:
                # Convertir les NCP sélectionnés en chaînes puis en nombres si nécessaire
                try:
                    # Essayer de convertir en nombres
                    selected_ncp_numeric = []
                    for ncp in selected_ncp:
                        try:
                            selected_ncp_numeric.append(int(ncp))
                        except ValueError:
                            selected_ncp_numeric.append(ncp)  # Garder en string si conversion échoue
                    
                    # Vérifier le type de la colonne NCP
                    ncp_dtype = self.df_bkhis['NCP'].dtype
                    
                    if ncp_dtype in [pl.Int64, pl.Int32, pl.Float64, pl.Float32]:
                        # Filtrer avec des nombres
                        df_filtered = df_filtered.filter(pl.col('NCP').is_in(selected_ncp_numeric))
                    else:
                        # Filtrer avec des chaînes
                        df_filtered = df_filtered.filter(pl.col('NCP').is_in(selected_ncp))
                    
                    self.log_message(f"Filtre NCP: {len(selected_ncp)} comptes sélectionnés ({df_filtered.height:,} lignes)", 'INFO')
                except Exception as e:
                    self.log_message(f"Erreur lors du filtrage NCP: {str(e)}", 'WARNING')
                    # Continuer sans filtrage NCP en cas d'erreur
            
            # Sélectionner les colonnes
            df_export = df_filtered.select(selected_columns)
            
            # Demander le fichier de destination
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                f"Exporter les données BKHIS ({format_export})",
                f"BKHIS_Export_{lettrage_filter.replace(' ', '_')}",
                self._get_file_extension(format_export)
            )
            
            if not file_path:
                return
            
            self.log_message(f"Export BKHIS: {file_path}", 'INFO')
            self.log_message(f"Format: {format_export} | Colonnes: {len(selected_columns)} | Lignes: {df_export.height:,}", 'INFO')
            
            # Exporter selon le format
            if format_export == "Excel":
                import pandas as pd
                from openpyxl import Workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                wb = Workbook()
                ws = wb.active
                ws.title = "BKHIS_DT"
                
                # Ajouter les données
                for r in dataframe_to_rows(df_export.to_pandas(), index=False, header=True):
                    ws.append(r)
                
                # Appliquer un formatage simple
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # En-tête en gras
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1ecce8", end_color="1ecce8", fill_type="solid")
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Ajuster la largeur des colonnes
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width
                
                wb.save(file_path)
                
            elif format_export == "CSV":
                df_export.write_csv(file_path, separator=';')
                
            elif format_export == "TXT":
                df_export.write_csv(file_path, separator='\t')
                
            elif format_export == "DSV":
                df_export.write_csv(file_path, separator='|')
                
            elif format_export == "Parquet":
                df_export.write_parquet(file_path)
            
            self.log_message(f"Export BKHIS réussi: {file_path}", 'SUCCESS')
            self.log_message(f"Lignes exportées: {df_export.height:,}", 'SUCCESS')
            
            QMessageBox.information(self, "Succès", 
                f"Export des données BKHIS réussi!\n\n"
                f"Fichier: {file_path}\n"
                f"Format: {format_export}\n"
                f"Lignes exportées: {df_export.height:,}\n"
                f"Colonnes: {len(selected_columns)}\n"
                f"Filtre: {lettrage_filter}")
            
            dialog.accept()
            
        except Exception as e:
            self.log_message(f"Erreur lors de l'export BKHIS: {str(e)}", 'ERROR')
            QMessageBox.critical(self, "Erreur", f"Échec de l'export BKHIS:\n{str(e)}")
    
    def _create_exhaustivite_export_tab(self):
        """Crée l'onglet d'export des résultats de rapprochement (Exhaustivité)"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Informations sur les données de rapprochement
        info_group = QGroupBox("📊 Informations sur le rapprochement")
        info_layout = QVBoxLayout(info_group)
        
        if hasattr(self, 'rapprochement_data') and self.rapprochement_data is not None:
            total_rapprochement = self.rapprochement_data.height
            rapprochees = self.rapprochement_data.filter(pl.col('NCP').is_not_null()).height
            non_rapprochees = total_rapprochement - rapprochees
            taux_rapprochement = (rapprochees / total_rapprochement * 100) if total_rapprochement > 0 else 0
            
            info_text = f"""
            <b>Total des écritures CPI:</b> {total_rapprochement:,}<br>
            <b>Écritures rapprochées:</b> {rapprochees:,} ({taux_rapprochement:.1f}%)<br>
            <b>Écritures non rapprochées:</b> {non_rapprochees:,} ({100-taux_rapprochement:.1f}%)<br>
            <b>Colonnes disponibles:</b> {len(self.rapprochement_data.columns)}
            """
            
            info_label = QLabel(info_text)
            info_label.setWordWrap(True)
            info_layout.addWidget(info_label)
        
        layout.addWidget(info_group)
        
        # Format d'export
        format_group = QGroupBox("📄 Format d'export")
        format_layout = QHBoxLayout(format_group)
        
        tab.format_combo = QComboBox()
        tab.format_combo.addItems(["Excel", "CSV", "TXT", "DSV", "Parquet"])
        format_layout.addWidget(QLabel("Format:"))
        format_layout.addWidget(tab.format_combo)
        
        layout.addWidget(format_group)
        
        # Filtre de rapprochement
        filter_group = QGroupBox("🔍 Filtre de rapprochement")
        filter_layout = QVBoxLayout(filter_group)
        
        tab.filter_combo = QComboBox()
        tab.filter_combo.addItems([
            "Toute la base (rapprochées + non rapprochées)",
            "Uniquement les écritures rapprochées",
            "Uniquement les écritures non rapprochées"
        ])
        filter_layout.addWidget(QLabel("Sélection:"))
        filter_layout.addWidget(tab.filter_combo)
        
        layout.addWidget(filter_group)
        
        # Sélection des colonnes
        columns_group = QGroupBox("📋 Colonnes à exporter")
        columns_layout = QVBoxLayout(columns_group)
        
        tab.columns_list = QListWidget()
        tab.columns_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        
        # Ajouter les colonnes disponibles
        if hasattr(self, 'rapprochement_data') and self.rapprochement_data is not None:
            # Créer une copie temporaire pour générer les colonnes d'export
            temp_data = self.rapprochement_data.clone()
            
            # 🔥 CRÉER LA COLONNE SOLDE BKHIS POUR L'AFFICHAGE
            if 'SENS' in temp_data.columns and 'MONT' in temp_data.columns:
                temp_data = temp_data.with_columns([
                    pl.when(pl.col('SENS') == 'C')
                    .then(pl.col('MONT'))
                    .otherwise(-pl.col('MONT'))
                    .alias('SOLDE BKHIS')
                ])
            
            # Définir l'ordre des colonnes pour l'affichage
            ordre_affichage = [
                'N°', 'DATE REGLEMENT CPI', 'SENS CPI', 'SOLDE CPI', 
                'AGE', 'NCP', 'DCO', 'OPE', 'PIE', 'SENS', 'SOLDE BKHIS', 
                'LETTRAGE', 'LIBELLE_OPR', 'STATUT', 'ECART'
            ]
            
            # Ajouter les colonnes dans l'ordre
            colonnes_affichees = []
            for col in ordre_affichage:
                if col in temp_data.columns:
                    colonnes_affichees.append(col)
            
            # Ajouter les colonnes restantes (sauf MONT)
            for col in temp_data.columns:
                if col not in colonnes_affichees and col != 'MONT':
                    colonnes_affichees.append(col)
            
            # Afficher les colonnes dans la liste
            for col in colonnes_affichees:
                item = QListWidgetItem(col)
                item.setSelected(True)  # Sélectionner toutes par défaut
                tab.columns_list.addItem(item)
        
        columns_layout.addWidget(tab.columns_list)
        
        # Boutons de sélection rapide
        columns_buttons_layout = QHBoxLayout()
        select_all_btn = QPushButton("Tout sélectionner")
        deselect_all_btn = QPushButton("Tout désélectionner")
        
        select_all_btn.clicked.connect(lambda: self._select_all_columns(tab.columns_list))
        deselect_all_btn.clicked.connect(lambda: self._deselect_all_columns(tab.columns_list))
        
        columns_buttons_layout.addWidget(select_all_btn)
        columns_buttons_layout.addWidget(deselect_all_btn)
        columns_layout.addLayout(columns_buttons_layout)
        
        layout.addWidget(columns_group)
        
        # Espace flexible
        layout.addStretch()
        
        return tab
    
    def _create_ventilation_export_tab(self):
        """Créer l'onglet d'export Ventilation (LEFT JOIN)"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Titre de l'onglet
        title_label = QLabel("Export Ventilation")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #60a5fa; padding: 10px;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # Description
        desc_label = QLabel("Effectue un LEFT JOIN entre les feuilles 'Sld.CPI.calc.instru' et 'Result Exhaustivité'")
        desc_label.setStyleSheet("font-size: 14px; color: #e2e8f0; padding: 10px;")
        desc_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(desc_label)
        
        # Section des fichiers à charger
        files_group = QGroupBox("📁 Fichiers Excel à charger")
        files_layout = QVBoxLayout(files_group)
        
        # Fichier Récapitulatif statistique
        recap_layout = QHBoxLayout()
        recap_label = QLabel("Fichier Récapitulatif statistique:")
        recap_label.setStyleSheet("font-size: 13px; color: #e2e8f0; padding: 5px;")
        recap_layout.addWidget(recap_label)
        
        self.recap_file_path = QLineEdit()
        self.recap_file_path.setPlaceholderText("Sélectionner le fichier Excel contenant 'Sld.CPI.calc.instru'...")
        self.recap_file_path.setStyleSheet("""
            QLineEdit {
                background-color: #010001;
                color: #e2e8f0;
                border: 1px solid #1ecce8;
                border-radius: 4px;
                padding: 5px;
                font-size: 12px;
            }
        """)
        recap_layout.addWidget(self.recap_file_path)
        
        recap_browse_btn = QPushButton("Parcourir...")
        recap_browse_btn.setStyleSheet("""
            QPushButton {
                background-color: #1ecce8;
                color: #010001;
                border: none;
                border-radius: 4px;
                padding: 5px 10px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #1bb8d4;
            }
        """)
        recap_browse_btn.clicked.connect(self._browse_recap_file)
        recap_layout.addWidget(recap_browse_btn)
        
        files_layout.addLayout(recap_layout)
        
        # Fichier Exhaustivité
        exhaust_layout = QHBoxLayout()
        exhaust_label = QLabel("Fichier Exhaustivité:")
        exhaust_label.setStyleSheet("font-size: 13px; color: #e2e8f0; padding: 5px;")
        exhaust_layout.addWidget(exhaust_label)
        
        self.exhaust_file_path = QLineEdit()
        self.exhaust_file_path.setPlaceholderText("Sélectionner le fichier Excel contenant 'Result Exhaustivité'...")
        self.exhaust_file_path.setStyleSheet("""
            QLineEdit {
                background-color: #010001;
                color: #e2e8f0;
                border: 1px solid #1ecce8;
                border-radius: 4px;
                padding: 5px;
                font-size: 12px;
            }
        """)
        exhaust_layout.addWidget(self.exhaust_file_path)
        
        exhaust_browse_btn = QPushButton("Parcourir...")
        exhaust_browse_btn.setStyleSheet("""
            QPushButton {
                background-color: #1ecce8;
                color: #010001;
                border: none;
                border-radius: 4px;
                padding: 5px 10px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #1bb8d4;
            }
        """)
        exhaust_browse_btn.clicked.connect(self._browse_exhaust_file)
        exhaust_layout.addWidget(exhaust_browse_btn)
        
        files_layout.addLayout(exhaust_layout)
        
        layout.addWidget(files_group)
        
        # Section des critères de jointure
        join_group = QGroupBox("🔗 Critères de jointure")
        join_layout = QVBoxLayout(join_group)
        
        join_info = QLabel(
            "• Feuille source: 'Sld.CPI.calc.instru' (Récapitulatif statistique)\n"
            "• Feuille cible: 'Result Exhaustivité' (Exhaustivité)\n"
            "• Critère: DATE REGLEMENT ↔ DATE REGLEMENT CPI\n"
            "• Type: LEFT JOIN"
        )
        join_info.setStyleSheet("font-size: 13px; color: #e2e8f0; padding: 10px;")
        join_layout.addWidget(join_info)
        
        layout.addWidget(join_group)
        
        # Section de destination
        output_group = QGroupBox("💾 Fichier de sortie")
        output_layout = QVBoxLayout(output_group)
        
        output_info = QLabel("Le résultat sera créé dans un nouvel onglet 'Ventilation' dans le fichier Excel d'exhaustivité")
        output_info.setStyleSheet("font-size: 13px; color: #1ecce8; padding: 10px;")
        output_layout.addWidget(output_info)
        
        layout.addWidget(output_group)
        
        # Espace flexible
        layout.addStretch()
        
        return tab
    
    def _browse_recap_file(self):
        """Parcourir pour sélectionner le fichier Récapitulatif statistique"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Sélectionner le fichier Récapitulatif statistique",
            "",
            "Fichiers Excel (*.xlsx *.xls)"
        )
        if file_path:
            self.recap_file_path.setText(file_path)
    
    def _browse_exhaust_file(self):
        """Parcourir pour sélectionner le fichier Exhaustivité"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Sélectionner le fichier Exhaustivité",
            "",
            "Fichiers Excel (*.xlsx *.xls)"
        )
        if file_path:
            self.exhaust_file_path.setText(file_path)
    
    def _perform_exhaustivite_export(self, dialog, tab):
        """Effectue l'export des résultats de rapprochement"""
        try:
            if not hasattr(self, 'rapprochement_data') or self.rapprochement_data is None:
                QMessageBox.warning(self, "Attention", "Aucune donnée de rapprochement disponible!")
                return
            
            # Récupérer les paramètres
            format_export = tab.format_combo.currentText()
            filter_selection = tab.filter_combo.currentText()
            selected_columns = [tab.columns_list.item(i).text() for i in range(tab.columns_list.count()) 
                              if tab.columns_list.item(i).isSelected()]
            
            if not selected_columns:
                QMessageBox.warning(self, "Attention", "Veuillez sélectionner au moins une colonne!")
                return
            
            # Appliquer le filtre de rapprochement selon le STATUT
            df_filtered = self.rapprochement_data.clone()
            
            # Vérifier si la colonne STATUT existe
            if 'STATUT' in df_filtered.columns:
                if filter_selection == "Uniquement les écritures rapprochées":
                    df_filtered = df_filtered.filter(pl.col('STATUT') == 'RAPPROCHE')
                    self.log_message(f"Filtre: Écritures rapprochées uniquement ({df_filtered.height:,} lignes)", 'INFO')
                elif filter_selection == "Uniquement les écritures non rapprochées":
                    df_filtered = df_filtered.filter(
                        (pl.col('STATUT') == 'NON RAPPROCHE CPI') | 
                        (pl.col('STATUT') == 'NON RAPPROCHE BKHIS')
                    )
                    self.log_message(f"Filtre: Écritures non rapprochées uniquement ({df_filtered.height:,} lignes)", 'INFO')
                else:
                    self.log_message(f"Filtre: Toute la base ({df_filtered.height:,} lignes)", 'INFO')
            else:
                self.log_message(f"Filtre: Toute la base (colonne STATUT non disponible) ({df_filtered.height:,} lignes)", 'INFO')
            
            # 🔥 CRÉER LA COLONNE SOLDE BKHIS AVANT LA VÉRIFICATION DES COLONNES
            if 'SENS' in df_filtered.columns and 'MONT' in df_filtered.columns:
                df_filtered = df_filtered.with_columns([
                    pl.when(pl.col('SENS') == 'C')
                    .then(pl.col('MONT'))
                    .otherwise(-pl.col('MONT'))
                    .alias('SOLDE BKHIS')
                ])
                self.log_message("✅ Colonne 'SOLDE BKHIS' créée (condition sur SENS)", 'INFO')
            else:
                if 'SENS' not in df_filtered.columns:
                    self.log_message("⚠️ Colonne 'SENS' non trouvée, SOLDE BKHIS non créé", 'WARNING')
                if 'MONT' not in df_filtered.columns:
                    self.log_message("⚠️ Colonne 'MONT' non trouvée, SOLDE BKHIS non créé", 'WARNING')
            
            # S'assurer que les colonnes essentielles sont incluses
            colonnes_essentielles = ['N°']
            if 'STATUT' in df_filtered.columns:
                colonnes_essentielles.append('STATUT')
            
            colonnes_disponibles = df_filtered.columns
            
            # Ajouter les colonnes essentielles si elles ne sont pas déjà sélectionnées
            for col in colonnes_essentielles:
                if col in colonnes_disponibles and col not in selected_columns:
                    selected_columns.insert(0, col)
            
            # Réorganiser les colonnes dans l'ordre souhaité
            ordre_colonnes = [
                'N°', 'DATE REGLEMENT CPI', 'SENS CPI', 'SOLDE CPI', 
                'AGE', 'NCP', 'DCO', 'OPE', 'PIE', 'SENS', 'SOLDE BKHIS', 
                'LIBELLE_OPR', 'STATUT', 'ECART'
            ]
            
            # Créer la liste finale des colonnes dans l'ordre
            colonnes_finales = []
            for col in ordre_colonnes:
                if col in selected_columns and col in colonnes_disponibles:
                    colonnes_finales.append(col)
            
            # Ajouter les colonnes sélectionnées restantes (sauf MONT)
            for col in selected_columns:
                if col not in colonnes_finales and col in colonnes_disponibles and col != 'MONT':
                    colonnes_finales.append(col)
            
            # Vérifier qu'il y a des colonnes à exporter
            if not colonnes_finales:
                QMessageBox.warning(self, "Attention", "Aucune colonne valide à exporter!")
                return
            
            # Sélectionner les colonnes dans le bon ordre (sans MONT)
            df_export = df_filtered.select(colonnes_finales)
            
            # Demander le fichier de destination
            extensions = {
                "Excel": ".xlsx", "CSV": ".csv", "TXT": ".txt", 
                "DSV": ".dsv", "Parquet": ".parquet"
            }
            extension = extensions.get(format_export, ".xlsx")
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, f"Exporter les résultats du FULL OUTER JOIN",
                f"FULL_OUTER_JOIN_EXHAUSTIVITE_{extension}",
                f"{format_export} files (*{extension})"
            )
            
            if not file_path:
                return
            
            # Exporter selon le format
            self.log_message(f"Export FULL OUTER JOIN: {file_path}", 'INFO')
            self.log_message(f"Format: {format_export} | Colonnes: {len(colonnes_finales)} | Lignes: {df_export.height}", 'INFO')
            
            if format_export == "Excel":
                import pandas as pd
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                # Convertir Polars en pandas pour l'export Excel
                pandas_df = df_export.to_pandas()
                
                # 🔥 REMPLIR LES ÉCARTS VIDES AVEC LES SOLDES CPI POUR LES NON RAPPROCHÉS
                # Trouver l'index des colonnes
                colonnes = pandas_df.columns.tolist()
                statut_idx = None
                ecart_idx = None
                solde_cpi_idx = None
                
                for i, col in enumerate(colonnes):
                    if col == 'STATUT':
                        statut_idx = i
                    elif col == 'ECART':
                        ecart_idx = i
                    elif col == 'SOLDE CPI':
                        solde_cpi_idx = i
                
                # Remplir les écarts vides pour les non rapprochés
                if statut_idx is not None and ecart_idx is not None and solde_cpi_idx is not None:
                    for idx, row in pandas_df.iterrows():
                        statut = row.iloc[statut_idx]
                        ecart = row.iloc[ecart_idx]
                        solde_cpi = row.iloc[solde_cpi_idx]
                        
                        # Si c'est non rapproché et que l'écart est vide/None
                        if statut in ['NON RAPPROCHE CPI', 'NON RAPPROCHE BKHIS'] and (ecart is None or ecart == '' or pd.isna(ecart)):
                            pandas_df.iloc[idx, ecart_idx] = solde_cpi
                    
                    self.log_message("✅ Écarts vides remplis avec les soldes CPI pour les non rapprochés", 'INFO')
                else:
                    self.log_message("⚠️ Colonnes requises non trouvées pour remplir les écarts", 'WARNING')
                
                # Créer un classeur Excel avec plusieurs feuilles
                wb = Workbook()
                
                # Supprimer la feuille par défaut
                wb.remove(wb.active)
                
                # === FEUILLE 1: Rappro + Non Rappro (toutes les données) ===
                ws1 = wb.create_sheet("Rappro + Non Rappro")
                
                # Définir la couleur noire pour l'onglet
                ws1.sheet_properties.tabColor = "000000"  # Noir
                
                # Ajouter les données
                for r_idx, row in enumerate(dataframe_to_rows(pandas_df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws1.cell(row=r_idx, column=c_idx, value=value)
                
                # Formater la première ligne en gras
                for cell in ws1[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Ajuster la largeur des colonnes
                for column in ws1.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws1.column_dimensions[column_letter].width = adjusted_width
                
                self.log_message("✅ Feuille 'Rappro + Non Rappro' créée (couleur: noir)", 'INFO')
                
                # === FEUILLE 2: Result Exhaustivité ===
                ws2 = wb.create_sheet("Result Exhaustivité")
                
                # Définir la couleur blanche pour l'onglet
                ws2.sheet_properties.tabColor = "FFFFFF"  # Blanc
                
                # Filtrer les données CPI (STATUT contient 'CPI')
                df_cpi = pandas_df[pandas_df['STATUT'].isin(['RAPPROCHE', 'NON RAPPROCHE CPI'])].copy()
                
                # Créer le tableau structuré
                tableau_cpi = df_cpi[['N°', 'DATE REGLEMENT CPI', 'SOLDE CPI', 'DCO', 'SOLDE BKHIS', 'ECART']].copy()
                
                # Renommer les colonnes pour l'affichage
                tableau_cpi.columns = ['N°', 'DATE REGLEMENT CPI', 'SOLDE CPI', 'DCO', 'SOLDE BKHIS', 'ECARTS']
                
                # Ajouter les données
                for r_idx, row in enumerate(dataframe_to_rows(tableau_cpi, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws2.cell(row=r_idx, column=c_idx, value=value)
                
                # Formater la première ligne
                for cell in ws2[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Calculer les indicateurs
                total_cpi = len(df_cpi)
                rapproches_cpi = len(df_cpi[df_cpi['STATUT'] == 'RAPPROCHE'])
                non_rapproches_cpi = len(df_cpi[df_cpi['STATUT'] == 'NON RAPPROCHE CPI'])
                taux_exhaustivite = (rapproches_cpi / total_cpi * 100) if total_cpi > 0 else 0
                
                # 🔥 CALCULER LES TOTAUX
                total_solde_cpi = tableau_cpi['SOLDE CPI'].sum()
                total_solde_bkhis = tableau_cpi['SOLDE BKHIS'].sum()
                total_ecarts = tableau_cpi['ECARTS'].sum()
                
                # Ajouter les indicateurs en dessous du tableau
                start_row = len(tableau_cpi) + 3
                
                # 🔥 AJOUTER LES TOTAUX JUSTE EN DESSOUS DES DATES
                # Écrire "TOTAL" en majuscule, gras, noir
                ws2.cell(row=start_row, column=1, value="TOTAL").font = Font(bold=True, color="000000")
                
                # Écrire les totaux en noir
                ws2.cell(row=start_row, column=3, value=total_solde_cpi).font = Font(bold=True, color="000000")
                ws2.cell(row=start_row, column=5, value=total_solde_bkhis).font = Font(bold=True, color="000000")
                ws2.cell(row=start_row, column=6, value=total_ecarts).font = Font(bold=True, color="000000")
                
                # Décaler les indicateurs vers le bas
                start_row = start_row + 2
                
                # Taux d'exhaustivité
                ws2.cell(row=start_row, column=1, value="Taux d'exhaustivité:").font = Font(bold=True)
                ws2.cell(row=start_row, column=2, value=f"{taux_exhaustivite:.2f}%").font = Font(bold=True, color="000000")
                
                # Nombre total des journées CPI
                ws2.cell(row=start_row + 1, column=1, value="Nombre total des journées CPI:").font = Font(bold=True)
                ws2.cell(row=start_row + 1, column=2, value=str(total_cpi)).font = Font(bold=True, color="000000")
                
                # Nombre des journées CPI rapprochées
                ws2.cell(row=start_row + 2, column=1, value="Nombre des journées CPI rapprochées:").font = Font(bold=True)
                ws2.cell(row=start_row + 2, column=2, value=str(rapproches_cpi)).font = Font(bold=True, color="000000")
                
                # Nombre des journées CPI non rapprochées
                ws2.cell(row=start_row + 3, column=1, value="Nombre des journées CPI non rapprochées:").font = Font(bold=True)
                ws2.cell(row=start_row + 3, column=2, value=str(non_rapproches_cpi)).font = Font(bold=True, color="000000")
                
                # Ajuster la largeur des colonnes
                for column in ws2.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws2.column_dimensions[column_letter].width = adjusted_width
                
                self.log_message(f"✅ Feuille 'Result Exhaustivité' créée (couleur: blanc) - Taux: {taux_exhaustivite:.2f}%", 'INFO')
                self.log_message(f"    Totaux - SOLDE CPI: {total_solde_cpi:,.2f}, SOLDE BKHIS: {total_solde_bkhis:,.2f}, ECARTS: {total_ecarts:,.2f}", 'INFO')
                
                # === FEUILLE 3: Rapproché (uniquement les lignes rapprochées) ===
                ws3 = wb.create_sheet("Rapproché")
                
                # Définir la couleur verte pour l'onglet
                ws3.sheet_properties.tabColor = "00FF00"  # Vert
                
                # Filtrer uniquement les lignes rapprochées
                df_rapproche = pandas_df[pandas_df['STATUT'] == 'RAPPROCHE'].copy()
                
                # Refaire le numéro d'ordre
                df_rapproche['N°'] = range(1, len(df_rapproche) + 1)
                
                # Ajouter les données
                for r_idx, row in enumerate(dataframe_to_rows(df_rapproche, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws3.cell(row=r_idx, column=c_idx, value=value)
                
                # Formater la première ligne
                for cell in ws3[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Ajuster la largeur des colonnes
                for column in ws3.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws3.column_dimensions[column_letter].width = adjusted_width
                
                self.log_message(f"✅ Feuille 'Rapproché' créée (couleur: verte) - {len(df_rapproche)} lignes", 'INFO')
                
                # === FEUILLE 4: Non rapproché (CPI puis BKHIS) ===
                ws4 = wb.create_sheet("Non_rapproché")
                
                # Définir la couleur rouge pour l'onglet
                ws4.sheet_properties.tabColor = "FF0000"  # Rouge
                
                # Filtrer les lignes non rapprochées
                df_non_rapproche = pandas_df[pandas_df['STATUT'].isin(['NON RAPPROCHE CPI', 'NON RAPPROCHE BKHIS'])].copy()
                
                # Séparer CPI et BKHIS
                df_non_rapproche_cpi = df_non_rapproche[df_non_rapproche['STATUT'] == 'NON RAPPROCHE CPI'].copy()
                df_non_rapproche_bkhis = df_non_rapproche[df_non_rapproche['STATUT'] == 'NON RAPPROCHE BKHIS'].copy()
                
                # Refaire le numéro d'ordre pour CPI
                df_non_rapproche_cpi['N°'] = range(1, len(df_non_rapproche_cpi) + 1)
                
                # Refaire le numéro d'ordre pour BKHIS (recommence à 1)
                df_non_rapproche_bkhis['N°'] = range(1, len(df_non_rapproche_bkhis) + 1)
                
                # Combiner CPI puis BKHIS
                df_non_rapproche_final = pd.concat([df_non_rapproche_cpi, df_non_rapproche_bkhis], ignore_index=True)
                
                # Ajouter les données
                for r_idx, row in enumerate(dataframe_to_rows(df_non_rapproche_final, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws4.cell(row=r_idx, column=c_idx, value=value)
                
                # Formater la première ligne
                for cell in ws4[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Ajuster la largeur des colonnes
                for column in ws4.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws4.column_dimensions[column_letter].width = adjusted_width
                
                self.log_message(f"✅ Feuille 'Non_rapproché' créée (couleur: rouge) - CPI: {len(df_non_rapproche_cpi)}, BKHIS: {len(df_non_rapproche_bkhis)}", 'INFO')
                
                # Réorganiser les feuilles dans le bon ordre
                # Supprimer et recréer les feuilles dans l'ordre correct
                sheets_order = ["Rappro + Non Rappro", "Result Exhaustivité", "Rapproché", "Non_rapproché"]
                for sheet_name in sheets_order:
                    if sheet_name in wb.sheetnames:
                        wb.move_sheet(wb[sheet_name], -1)  # Déplacer à la fin
                
                # Réorganiser dans l'ordre inverse pour obtenir le bon ordre
                for sheet_name in reversed(sheets_order):
                    if sheet_name in wb.sheetnames:
                        wb.move_sheet(wb[sheet_name], 0)  # Déplacer au début
                
                # Sauvegarder le fichier Excel
                wb.save(file_path)
                
                self.log_message(f"✅ Fichier Excel multi-feuilles créé: {file_path}", 'SUCCESS')
                self.log_message(f"   📊 Feuille 1 'Rappro + Non Rappro': {len(pandas_df)} lignes (couleur: noir)", 'INFO')
                self.log_message(f"   📊 Feuille 2 'Result Exhaustivité': {len(df_cpi)} lignes (couleur: blanc, Taux: {taux_exhaustivite:.2f}%)", 'INFO')
                self.log_message(f"   📊 Feuille 3 'Rapproché': {len(df_rapproche)} lignes (couleur: verte)", 'INFO')
                self.log_message(f"   📊 Feuille 4 'Non_rapproché': {len(df_non_rapproche_final)} lignes (couleur: rouge)", 'INFO')
                
            elif format_export == "CSV":
                df_export.write_csv(file_path, separator=';')
            elif format_export == "TXT":
                df_export.write_csv(file_path, separator='\t')
            elif format_export == "DSV":
                df_export.write_csv(file_path, separator='|')
            elif format_export == "Parquet":
                df_export.write_parquet(file_path)
            
            self.log_message(f"Export FULL OUTER JOIN réussi: {file_path}", 'SUCCESS')
            self.log_message(f"Lignes exportées: {df_export.height:,}", 'SUCCESS')
            
            QMessageBox.information(self, "Succès", 
                f"Export des résultats du FULL OUTER JOIN réussi!\n\n"
                f"Fichier: {file_path}\n"
                f"Format: {format_export}\n"
                f"Filtre: {filter_selection}\n"
                f"Lignes exportées: {df_export.height:,}\n"
                f"Colonnes: {len(colonnes_finales)}\n\n"
                f"✅ Colonne 'SOLDE BKHIS' créée (condition sur SENS)\n"
                f"❌ Colonne 'MONT' supprimée de l'export\n\n"
                f"📊 Fichier Excel multi-feuilles personnalisé:\n"
                f"   📋 Feuille 1: 'Rappro + Non Rappro' (couleur: noir)\n"
                f"   📋 Feuille 2: 'Result Exhaustivité' (couleur: blanc)\n"
                f"   📋 Feuille 3: 'Rapproché' (couleur: verte)\n"
                f"   📋 Feuille 4: 'Non_rapproché' (couleur: rouge)\n\n"
                f"📈 Indicateurs ajoutés dans 'Result Exhaustivité':\n"
                f"   🧮 Taux d'exhaustivité\n"
                f"   📅 Nombre total des journées CPI\n"
                f"   ✅ Nombre des journées CPI rapprochées\n"
                f"   ❌ Nombre des journées CPI non rapprochées")
            
            dialog.accept()
            
        except Exception as e:
            self.log_message(f"Erreur lors de l'export du FULL OUTER JOIN: {str(e)}", 'ERROR')
            QMessageBox.critical(self, "Erreur", f"Échec de l'export du FULL OUTER JOIN:\n{str(e)}")
    
    def _get_file_extension(self, format_name):
        """Retourne l'extension de fichier selon le format"""
        extensions = {
            "Excel": "Fichiers Excel (*.xlsx)",
            "CSV": "Fichiers CSV (*.csv)",
            "TXT": "Fichiers texte (*.txt)",
            "DSV": "Fichiers DSV (*.dsv)",
            "Parquet": "Fichiers Parquet (*.parquet)"
        }
        return extensions.get(format_name, "Tous les fichiers (*)")
    
    def _select_all_columns(self, columns_list):
        """Sélectionne toutes les colonnes"""
        for i in range(columns_list.count()):
            columns_list.item(i).setSelected(True)
    
    def _deselect_all_columns(self, columns_list):
        """Désélectionne toutes les colonnes"""
        for i in range(columns_list.count()):
            columns_list.item(i).setSelected(False)
    
    def _perform_filtered_export(self, dialog, format_combo, columns_list, 
                                instruments_list, status_list, data_to_export, data_name):
        """Effectue l'exportation avec les filtres appliqués"""
        # Récupérer les colonnes sélectionnées
        selected_columns = [columns_list.item(i).text() for i in range(columns_list.count()) 
                           if columns_list.item(i).isSelected()]
        
        if not selected_columns:
            QMessageBox.warning(self, "Attention", "Veuillez sélectionner au moins une colonne!")
            return
        
        # Commencer avec toutes les données
        filtered_data = data_to_export
        
        # Filtrer par instruments de paiement (AVANT de sélectionner les colonnes)
        if instruments_list is not None:
            selected_instruments = [instruments_list.item(i).text() for i in range(instruments_list.count()) 
                                  if instruments_list.item(i).isSelected()]
            
            if selected_instruments:
                # Vérifier si la colonne INST PAIEMENT existe
                if "INST PAIEMENT" in filtered_data.columns:
                    filtered_data = filtered_data.filter(
                        pl.col("INST PAIEMENT").is_in(selected_instruments)
                    )
                    self.log_message(f"Filtre instruments: {selected_instruments}", 'DEBUG')
                    self.log_message(f"Lignes après filtre instruments: {filtered_data.height}", 'DEBUG')
                else:
                    self.log_message("Colonne INST PAIEMENT non trouvée dans les données", 'WARNING')
        
        # Filtrer par statut (AVANT de sélectionner les colonnes)
        if status_list is not None:
            selected_statuses = [status_list.item(i).text() for i in range(status_list.count()) 
                               if status_list.item(i).isSelected()]
            
            if selected_statuses:
                # Vérifier si la colonne Statut existe
                if "Statut" in filtered_data.columns:
                    filtered_data = filtered_data.filter(
                        pl.col("Statut").is_in(selected_statuses)
                    )
                    self.log_message(f"Filtre statuts: {selected_statuses}", 'DEBUG')
                    self.log_message(f"Lignes après filtre statuts: {filtered_data.height}", 'DEBUG')
                else:
                    self.log_message("Colonne Statut non trouvée dans les données", 'WARNING')
        
        # Maintenant sélectionner uniquement les colonnes demandées
        # S'assurer que les colonnes de filtrage sont incluses si elles existent
        final_columns = selected_columns.copy()
        if "INST PAIEMENT" in filtered_data.columns and "INST PAIEMENT" not in final_columns:
            final_columns.append("INST PAIEMENT")
        if "Statut" in filtered_data.columns and "Statut" not in final_columns:
            final_columns.append("Statut")
        
        # Ne garder que les colonnes qui existent réellement
        final_columns = [col for col in final_columns if col in filtered_data.columns]
        
        if final_columns:
            filtered_data = filtered_data.select(final_columns)
        
        # Récupérer le format
        format_choice = format_combo.currentText()
        extensions = {"Parquet": ".parquet", "CSV": ".csv", "TXT": ".txt", 
                     "DSV": ".dsv", "Excel": ".xlsx"}
        extension = extensions.get(format_choice, ".csv")
        
        # Boîte de dialogue pour choisir l'emplacement
        file_path, _ = QFileDialog.getSaveFileName(
            self, f"Exporter en {format_choice}",
            f"CPI_EXPORT_FILTRE_{data_name.upper()}{extension}",
            f"{format_choice} files (*{extension})"
        )
        
        if not file_path:
            return
        
        try:
            self.log_message(f"Export filtré en {format_choice}...", 'INFO')
            self.log_message(f"Colonnes: {len(selected_columns)}, Lignes: {filtered_data.height}", 'INFO')
            
            if format_choice == "Parquet":
                filtered_data.write_parquet(file_path)
            elif format_choice == "CSV":
                filtered_data.write_csv(file_path, separator='|')
            elif format_choice == "TXT":
                filtered_data.write_csv(file_path, separator='|')
            elif format_choice == "DSV":
                filtered_data.write_csv(file_path, separator='|')
            elif format_choice == "Excel":
                import pandas as pd
                # Convertir Polars en pandas en préservant les noms de colonnes
                pandas_df = filtered_data.to_pandas()
                pandas_df.to_excel(file_path, index=False, header=True)
            
            self.log_message(f"Export réussi: {file_path}", 'SUCCESS')
            QMessageBox.information(self, "Succès", 
                f"Données exportées avec succès!\n"
                f"Fichier: {file_path}\n"
                f"Lignes exportées: {filtered_data.height:,}\n"
                f"Colonnes: {len(selected_columns)}")
            
            dialog.accept()
            
        except Exception as e:
            self.log_message(f"Erreur export: {str(e)}", 'ERROR')
            QMessageBox.critical(self, "Erreur", f"Échec de l'export:\n{str(e)}")
    
    def verifier_exhaustivite(self):
        """Vérification de l'exhaustivité - Étape 1: Charger la base BKHIS"""
        self.log_message("Démarrage de la vérification d'exhaustivité...", 'INFO')
        
        # Ouvrir une boîte de dialogue pour sélectionner le fichier BKHIS
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            "Sélectionner la base BKHIS",
            "",
            "Fichiers texte (*.txt);;Fichiers séparés par des pipes (*.dsv);;Tous les fichiers (*)"
        )
        
        if not file_path:
            self.log_message("❌ Aucun fichier BKHIS sélectionné", 'WARNING')
            return
        
        try:
            self.log_message(f"📂 Chargement de la base BKHIS: {file_path}", 'INFO')
            
            # Définir les noms de colonnes corrects pour BKHIS
            noms_colonnes_bkhis = [
                'AGE', 'DEV', 'NCP', 'UNKNOWN_COL_4', 'DCO', 'OPE', 'MVT', 'UNKNOWN_COL_8',
                'DVA', 'DATE', 'MONT', 'SENS', 'LIBELLE_OPR', 'EXO', 'PIE', 'UNKNOWN_COL_16',
                'UNKNOWN_COL_17', 'UNKNOWN_COL_18', 'UNKNOWN_COL_19', 'UNKNOWN_COL_20',
                'UTIL', 'UNKNOWN_COL_22', 'UNKNOWN_COL_23', 'UNKNOWN_COL_24', 'UNKNOWN_COL_25',
                'DAG', 'NCP_OTHER', 'UNKNOWN_COL_28', 'UNKNOWN_COL_29', 'UNKNOWN_COL_30',
                'UNKNOWN_COL_31', 'UNKNOWN_COL_32', 'UNKNOWN_COL_33', 'UNKNOWN_COL_34',
                'UNKNOWN_COL_35', 'UNKNOWN_COL_36', 'UNKNOWN_COL_37', 'UNKNOWN_COL_38',
                'UNKNOWN_COL_39', 'UNKNOWN_COL_40'
            ]
            
            # Charger le fichier BKHIS avec séparateur pipe | et sans en-têtes
            # Essayer différents encodages et approches pour supporter les fichiers UTF-8 et malformés
            encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
            bkhis_loaded = False
            
            for encoding in encodings_to_try:
                try:
                    self.log_message(f"📂 Tentative de chargement avec encodage: {encoding}", 'DEBUG')
                    
                    # Essai 1: Approche standard avec ignore_errors
                    try:
                        self.df_bkhis = pl.read_csv(
                            file_path, 
                            separator='|',
                            has_header=False,
                            try_parse_dates=True,
                            infer_schema_length=10000,
                            new_columns=noms_colonnes_bkhis,
                            encoding=encoding,
                            ignore_errors=True
                        )
                        bkhis_loaded = True
                        self.log_message(f"✅ Fichier chargé avec succès (standard) en utilisant l'encodage: {encoding}", 'SUCCESS')
                        break
                    except Exception as e1:
                        self.log_message(f"🔄 Échec standard avec {encoding}: {str(e1)}", 'DEBUG')
                        
                        # Essai 2: Approche tout-string
                        try:
                            self.df_bkhis = pl.read_csv(
                                file_path, 
                                separator='|',
                                has_header=False,
                                try_parse_dates=False,
                                infer_schema_length=0,
                                dtypes={col: pl.Utf8 for col in noms_colonnes_bkhis},
                                new_columns=noms_colonnes_bkhis,
                                encoding=encoding,
                                ignore_errors=True
                            )
                            bkhis_loaded = True
                            self.log_message(f"✅ Fichier chargé avec succès (tout-string) en utilisant l'encodage: {encoding}", 'SUCCESS')
                            break
                        except Exception as e2:
                            self.log_message(f"🔄 Échec tout-string avec {encoding}: {str(e2)}", 'DEBUG')
                            continue
                            

                except Exception as e:
                    self.log_message(f"❌ Échec complet avec encodage {encoding}: {str(e)}", 'DEBUG')
                    continue
            
            if not bkhis_loaded:
                # Dernière tentative: pandas avec tolérance
                try:
                    self.log_message("🔄 Dernière tentative avec pandas...", 'INFO')
                    import pandas as pd
                    
                    df_temp = pd.read_csv(
                        file_path,
                        sep='|',
                        header=None,
                        names=noms_colonnes_bkhis,
                        encoding='latin-1',  # Plus tolérant
                        engine='python',
                        on_bad_lines='skip',  # Sauter lignes malformées
                        dtype=str
                    )
                    
                    self.df_bkhis = pl.from_pandas(df_temp)
                    bkhis_loaded = True
                    self.log_message("✅ Fichier chargé avec succès (pandas)", 'SUCCESS')
                except Exception as e:
                    self.log_message(f"❌ Échec final: {str(e)}", 'ERROR')
                    raise Exception(f"Impossible de charger le fichier BKHIS avec aucune méthode")
            
            # Afficher les informations de base sur la base BKHIS
            self.log_message(f"✅ Base BKHIS chargée avec succès!", 'SUCCESS')
            self.log_message(f"📊 Dimensions: {self.df_bkhis.height:,} lignes × {self.df_bkhis.width} colonnes", 'INFO')
            self.log_message(f"📋 Colonnes créées: {', '.join(self.df_bkhis.columns)}", 'INFO')
            
            # Vérifier les colonnes importantes
            colonnes_importantes = ['AGE', 'DEV', 'NCP', 'DCO', 'OPE', 'MVT', 'DVA', 'DATE', 'MONT', 'SENS', 'LIBELLE_OPR', 'EXO', 'PIE', 'UTIL', 'DAG', 'NCP_OTHER']
            
            # Étape 3: Créer les colonnes conditionnelles DEBIT et CREDIT
            self.log_message("Création des colonnes conditionnelles DEBIT et CREDIT...", 'INFO')
            
            # Vérifier que les colonnes nécessaires existent
            if 'SENS' in self.df_bkhis.columns and 'MONT' in self.df_bkhis.columns:
                # Nettoyer la colonne MONT d'abord
                self.df_bkhis = self.df_bkhis.with_columns(
                    pl.col('MONT').str.replace_all(',', '.').str.replace_all(r'\s+', '').cast(pl.Float64)
                )
                
                # Trouver l'index de la colonne SENS pour insérer après
                sens_index = self.df_bkhis.columns.index('SENS')
                
                # Créer les nouvelles colonnes avec gestion des valeurs nulles
                self.df_bkhis = self.df_bkhis.with_columns([
                    # Colonne DEBIT: prend MONT si SENS = "D", sinon null (pas de conversion pour éviter NaN)
                    pl.when(pl.col('SENS') == 'D')
                    .then(pl.col('MONT'))
                    .otherwise(pl.lit(0.0))  # Utiliser 0.0 au lieu de None
                    .alias('DEBIT'),
                    
                    # Colonne CREDIT: prend MONT si SENS = "C", sinon null  
                    pl.when(pl.col('SENS') == 'C')
                    .then(pl.col('MONT'))
                    .otherwise(pl.lit(0.0))  # Utiliser 0.0 au lieu de None
                    .alias('CREDIT')
                ])
                
                # Maintenant, remplacer les 0.0 par None là où ils ne devraient pas avoir de valeurs
                self.df_bkhis = self.df_bkhis.with_columns([
                    pl.when(pl.col('SENS') != 'D')
                    .then(pl.lit(None))
                    .otherwise(pl.col('DEBIT'))
                    .alias('DEBIT'),
                    
                    pl.when(pl.col('SENS') != 'C')
                    .then(pl.lit(None))
                    .otherwise(pl.col('CREDIT'))
                    .alias('CREDIT')
                ])
                
                # Réorganiser les colonnes pour placer DEBIT et CREDIT juste après SENS
                nouvelles_colonnes = []
                for i, col in enumerate(self.df_bkhis.columns):
                    nouvelles_colonnes.append(col)
                    # Si on vient de traiter SENS, ajouter DEBIT et CREDIT
                    if col == 'SENS':
                        nouvelles_colonnes.extend(['DEBIT', 'CREDIT'])
                
                # Supprimer les doublons DEBIT et CREDIT s'ils existent déjà à la fin
                nouvelles_colonnes = list(dict.fromkeys(nouvelles_colonnes))
                
                # Réorganiser le DataFrame
                self.df_bkhis = self.df_bkhis.select(nouvelles_colonnes)
                
                self.log_message("✅ Colonnes DEBIT et CREDIT créées avec succès!", 'SUCCESS')
                self.log_message(f"📊 Nouvelles dimensions: {self.df_bkhis.height:,} lignes × {self.df_bkhis.width} colonnes", 'INFO')
                
                # Statistiques sur les nouvelles colonnes
                total_debit = self.df_bkhis.select(pl.col('DEBIT').sum()).item()
                total_credit = self.df_bkhis.select(pl.col('CREDIT').sum()).item()
                nb_debit = self.df_bkhis.select(pl.col('DEBIT').count()).item()
                nb_credit = self.df_bkhis.select(pl.col('CREDIT').count()).item()
                
                # self.log_message(f" Total DEBIT: {total_debit:,.2f} ({nb_debit} opérations)", 'INFO')
                # self.log_message(f" Total CREDIT: {total_credit:,.2f} ({nb_credit} opérations)", 'INFO')
                
                # Étape 4: Lettrage automatique avec approche vectorielle polars
                # self.log_message("🔧 Étape 4: Lettrage automatique (approche vectorielle)...", 'INFO')
                
                # Vérifier que les colonnes nécessaires existent
                if 'NCP' in self.df_bkhis.columns and 'DEBIT' in self.df_bkhis.columns and 'CREDIT' in self.df_bkhis.columns:
                    
                    # 1️⃣ Créer une colonne MONT_UNIFIED pour le lettrage (DEBIT ou CREDIT)
                    self.df_bkhis = self.df_bkhis.with_columns(
                        pl.when(pl.col('DEBIT').is_not_null())
                        .then(pl.col('DEBIT'))
                        .when(pl.col('CREDIT').is_not_null())
                        .then(pl.col('CREDIT'))
                        .otherwise(None)
                        .alias('MONT_UNIFIED')
                    )
                    
                    # 2️⃣ Créer une clé de lettrage (NCP + MONT)
                    self.df_bkhis = self.df_bkhis.with_columns(
                        pl.concat_str(
                            [
                                pl.col("NCP").cast(pl.Utf8),
                                pl.col("MONT_UNIFIED").cast(pl.Utf8)
                            ],
                            separator="_"
                        ).alias("CLE_LETTRAGE")
                    )
                    
                    # 3️⃣ Compter les écritures par clé (pour trouver les paires)
                    self.df_bkhis = self.df_bkhis.with_columns(
                        pl.count().over("CLE_LETTRAGE").alias("NB_ECR_PAR_CLE")
                    )
                    
                    # 4️⃣ Marquer les écritures lettrables (au moins 2 écritures avec même NCP+MONT)
                    self.df_bkhis = self.df_bkhis.with_columns(
                        pl.when(
                            (pl.col("NB_ECR_PAR_CLE") >= 2) & 
                            (pl.col('MONT_UNIFIED').is_not_null())
                        )
                        .then(pl.lit('L'))
                        .otherwise(None)
                        .alias('LETTRAGE')
                    )
                    
                    # 5️⃣ Nettoyage des colonnes temporaires
                    self.df_bkhis = self.df_bkhis.drop(['MONT_UNIFIED', 'CLE_LETTRAGE', 'NB_ECR_PAR_CLE'])
                    
                    # Réorganiser les colonnes pour placer LETTRAGE juste après CREDIT
                    nouvelles_colonnes_avec_lettrage = []
                    for col in self.df_bkhis.columns:
                        nouvelles_colonnes_avec_lettrage.append(col)
                        if col == 'CREDIT':
                            nouvelles_colonnes_avec_lettrage.append('LETTRAGE')
                    
                    # Supprimer les doublons et réorganiser
                    nouvelles_colonnes_avec_lettrage = list(dict.fromkeys(nouvelles_colonnes_avec_lettrage))
                    self.df_bkhis = self.df_bkhis.select(nouvelles_colonnes_avec_lettrage)
                    
                    self.log_message("✅ Lettrage vectoriel effectué avec succès!", 'SUCCESS')
                    self.log_message(f"📊 Dimensions finales: {self.df_bkhis.height:,} lignes × {self.df_bkhis.width} colonnes", 'INFO')
                    
                    # Statistiques de lettrage
                    total_lettre = self.df_bkhis.select(pl.col('LETTRAGE').count()).item()
                    total_lignes = self.df_bkhis.height
                    taux_lettrage = (total_lettre / total_lignes * 100) if total_lignes > 0 else 0
                    
                    self.log_message(f"🔗 Lignes lettrées: {total_lettre:,} / {total_lignes:,} ({taux_lettrage:.1f}%)", 'INFO')
                    
                    # Détail par NCP lettré
                    ncp_lettres = self.df_bkhis.filter(
                        pl.col('LETTRAGE') == 'L'
                    ).select('NCP').unique().to_pandas()['NCP'].tolist()
                    
                    self.log_message(f"📋 NCP lettrés: {len(ncp_lettres)} comptes", 'INFO')
                    
                else:
                    self.log_message("❌ Colonnes NCP, DEBIT ou CREDIT non trouvées pour le lettrage", 'ERROR')
                    return
                
            else:
                self.log_message("❌ Colonnes SENS ou MONT non trouvées pour créer DEBIT/CREDIT", 'ERROR')
                return
            
            QMessageBox.information(self, "Base BKHIS chargée", 
                                  f"Base BKHIS chargée avec succès!\n\n"
                                  f"Fichier: {file_path}\n"
                                  f"Lignes: {self.df_bkhis.height:,}\n"
                                  f"Colonnes: {self.df_bkhis.width}\n\n"
                                  f"Prêt pour l'export et l'analyse.")
            
            # Étape 5: Rapprochement avec la feuille Sld.CPI.calc.date
            self.log_message("Rapprochement BKHIS avec Sld.CPI.calc.date...", 'INFO')
            self.rapprochement_bkhis_cpi()
            
        except Exception as e:
            self.log_message(f"❌ Erreur lors du chargement de BKHIS: {str(e)}", 'ERROR')
            QMessageBox.critical(self, "Erreur", f"Échec du chargement de la base BKHIS:\n{str(e)}")
    
    def rapprochement_bkhis_cpi(self):
        """Rapprochement entre BKHIS et la feuille Sld.CPI.calc.date"""
        try:
            # Vérifier que les données BKHIS existent
            if not hasattr(self, 'df_bkhis') or self.df_bkhis is None:
                self.log_message("❌ Base BKHIS non chargée", 'ERROR')
                return
            
            self.log_message("📊 Préparation des données pour le rapprochement...", 'INFO')
            
            # 1️⃣ Utiliser toutes les écritures BKHIS (plus de filtrage)
            bkhis_pour_rapprochement = self.df_bkhis
            
            self.log_message(f"📋 Total écritures BKHIS utilisées: {bkhis_pour_rapprochement.height:,} lignes", 'INFO')
            
            # 2️⃣ Charger les données CPI depuis le fichier original (comme dans l'export statistique)
            cpi_file_path = None
            
            # Vérifier si le fichier CPI est déjà spécifié
            if hasattr(self, 'file_path') and self.file_path:
                cpi_file_path = self.file_path
                self.log_message(f"📂 Utilisation du fichier CPI existant: {cpi_file_path}", 'INFO')
            else:
                # Demander à l'utilisateur de sélectionner le fichier CPI
                cpi_file_path, _ = QFileDialog.getOpenFileName(
                    self, 
                    "Sélectionner le fichier CPI pour le rapprochement",
                    "",
                    "Fichiers Excel (*.xlsx);;Fichiers Excel 97-2003 (*.xls);;Tous les fichiers (*)"
                )
                
                if not cpi_file_path:
                    self.log_message("❌ Aucun fichier CPI sélectionné", 'ERROR')
                    return
                
                self.log_message(f"📂 Fichier CPI sélectionné: {cpi_file_path}", 'INFO')
            
            try:
                # Charger les données
                raw_data = pl.read_excel(cpi_file_path)
                
                # Vérifier les colonnes disponibles
                self.log_message(f"📋 Colonnes trouvées dans le fichier CPI: {raw_data.columns}", 'INFO')
                
                # Vérifier que la colonne solde_total existe
                if 'solde_total' not in raw_data.columns:
                    self.log_message("❌ Colonne 'solde_total' non trouvée dans le fichier CPI", 'ERROR')
                    self.log_message(f"Colonnes disponibles: {raw_data.columns}", 'ERROR')
                    return
                
                # Vérifier le type de la colonne solde_total
                solde_dtype = raw_data['solde_total'].dtype
                self.log_message(f"📊 Type de colonne solde_total: {solde_dtype}", 'INFO')
                
                # Traiter la colonne solde_total selon son type
                if solde_dtype in [pl.Utf8, pl.String]:
                    # Si c'est du texte, nettoyer et convertir
                    cpi_data = raw_data.with_columns([
                        pl.col('solde_total').str.replace_all(',', '.').str.replace_all(r'\s+', '').cast(pl.Float64)
                    ])
                else:
                    # Si c'est déjà numérique, utiliser directement
                    cpi_data = raw_data
                
                # Dupliquer solde_total en Montant CPI (valeur absolue)
                cpi_data = cpi_data.with_columns([
                    pl.col('solde_total').abs().alias('Montant CPI')
                ])
                
                self.log_message(f"✅ Données CPI chargées: {cpi_data.height:,} lignes", 'SUCCESS')
                self.log_message("✅ Colonne 'Montant CPI' créée (copie de solde_total en valeur absolue)", 'SUCCESS')
                
            except Exception as e:
                self.log_message(f"❌ Erreur lors du chargement du fichier CPI: {str(e)}", 'ERROR')
                return
            
            # 4️⃣ Effectuer le rapprochement avec tolérance sur les montants et correspondance des SENS
            self.log_message("🔗 Rapprochement avec tolérance (Montant CPI ↔ MONT BKHIS) ±0.05 + correspondance SENS...", 'INFO')
            
            # Renommer la colonne MONT de BKHIS pour éviter les conflits
            bkhis_pour_join = bkhis_pour_rapprochement.with_columns(
                pl.col('MONT').alias('MONT_BKHIS')
            )
            
            # Créer une fonction de rapprochement FULL OUTER JOIN avec tolérance
            def rapprocher_full_outer_join(df_cpi, df_bkhis, tolerance=0.05):
                """Effectue un FULL OUTER JOIN avec tolérance sur les montants ET correspondance des SENS"""
                import pandas as pd  # Import local pour éviter les conflits
                
                try:
                    self.log_message("🔍 Début du FULL OUTER JOIN avec tolérance...", 'INFO')
                    
                    # Convertir en pandas pour faciliter le traitement
                    cpi_df = df_cpi.to_pandas()
                    bkhis_df = df_bkhis.to_pandas()
                    
                    self.log_message(f"📊 CPI: {len(cpi_df)} lignes, BKHIS: {len(bkhis_df)} lignes", 'INFO')
                    
                    # Diagnostic des colonnes disponibles (non affiché dans le journal)
                    # self.log_message(f"📋 Colonnes CPI: {list(cpi_df.columns)}", 'DEBUG')
                    # self.log_message(f"📋 Colonnes BKHIS: {list(bkhis_df.columns)}", 'DEBUG')
                    
                    # Vérifier les colonnes nécessaires
                    if 'Montant CPI' not in cpi_df.columns:
                        self.log_message("❌ Colonne 'Montant CPI' manquante dans CPI", 'ERROR')
                        return pl.DataFrame()
                    
                    if 'MONT' not in bkhis_df.columns:
                        self.log_message("❌ Colonne 'MONT' manquante dans BKHIS", 'ERROR')
                        return pl.DataFrame()
                    
                    # Préparer les données pour le FULL OUTER JOIN
                    results = []
                    matched_cpi_indices = set()
                    matched_bkhis_indices = set()
                    
                    # Étape 1: Trouver toutes les correspondances avec tolérance
                    for idx_cpi, row_cpi in cpi_df.iterrows():
                        montant_cpi = float(row_cpi['Montant CPI'])
                        best_match = None
                        best_diff = float('inf')
                        best_bkhis_idx = None
                        
                        for idx_bkhis, row_bkhis in bkhis_df.iterrows():
                            if idx_bkhis in matched_bkhis_indices:
                                continue  # Déjà matché
                            
                            montant_bkhis = float(row_bkhis['MONT'])
                            sens_bkhis = str(row_bkhis.get('SENS', '')).strip()
                            sens_cpi = str(row_cpi.get('Sens', '')).strip()
                            
                            # Vérifier la compatibilité des sens
                            # D (Débit) CPI correspond à D (Débit) BKHIS
                            # C (Crédit) CPI correspond à C (Crédit) BKHIS
                            if sens_cpi != sens_bkhis:
                                continue  # Sens différents, pas de rapprochement possible
                            
                            diff = abs(montant_cpi - montant_bkhis)
                            
                            if diff <= tolerance and diff < best_diff:
                                best_diff = diff
                                best_match = row_bkhis
                                best_bkhis_idx = idx_bkhis
                        
                        if best_match is not None:
                            # Créer la ligne fusionnée
                            merged_row = {
                                'N°': len(results) + 1,
                                'DATE REGLEMENT CPI': row_cpi.get('DATEREGLEMENT', ''),
                                'SENS CPI': row_cpi.get('Sens', ''),
                                'SOLDE CPI': row_cpi.get('solde_total', 0),
                                'AGE': best_match.get('AGE', ''),
                                'NCP': best_match.get('NCP', ''),
                                'DCO': best_match.get('DCO', ''),
                                'OPE': best_match.get('OPE', ''),
                                'PIE': best_match.get('PIE', ''),
                                'SENS': best_match.get('SENS', ''),
                                'MONT': best_match.get('MONT', 0),
                                'LETTRAGE': best_match.get('LETTRAGE', ''),
                                'LIBELLE_OPR': best_match.get('LIBELLE_OPR', ''),
                                'STATUT': 'RAPPROCHE',
                                'ECART': best_diff
                            }
                            results.append(merged_row)
                            matched_cpi_indices.add(idx_cpi)
                            matched_bkhis_indices.add(best_bkhis_idx)
                    
                    # Étape 2: Ajouter les lignes CPI non rapprochées
                    for idx_cpi, row_cpi in cpi_df.iterrows():
                        if idx_cpi not in matched_cpi_indices:
                            merged_row = {
                                'N°': len(results) + 1,
                                'DATE REGLEMENT CPI': row_cpi.get('DATEREGLEMENT', ''),
                                'SENS CPI': row_cpi.get('Sens', ''),
                                'SOLDE CPI': row_cpi.get('solde_total', 0),
                                'AGE': '',
                                'NCP': '',
                                'DCO': '',
                                'OPE': '',
                                'PIE': '',
                                'SENS': '',
                                'MONT': 0,
                                'LETTRAGE': None,  # Pas de données BKHIS = None
                                'LIBELLE_OPR': '',
                                'STATUT': 'NON RAPPROCHE CPI',
                                'ECART': None
                            }
                            results.append(merged_row)
                    
                    # Étape 3: Ajouter les lignes BKHIS non rapprochées
                    for idx_bkhis, row_bkhis in bkhis_df.iterrows():
                        if idx_bkhis not in matched_bkhis_indices:
                            merged_row = {
                                'N°': len(results) + 1,
                                'DATE REGLEMENT CPI': '',
                                'SENS CPI': '',
                                'SOLDE CPI': 0,
                                'AGE': row_bkhis.get('AGE', ''),
                                'NCP': row_bkhis.get('NCP', ''),
                                'DCO': row_bkhis.get('DCO', ''),
                                'OPE': row_bkhis.get('OPE', ''),
                                'PIE': row_bkhis.get('PIE', ''),
                                'SENS': row_bkhis.get('SENS', ''),
                                'MONT': row_bkhis.get('MONT', 0),
                                'LETTRAGE': row_bkhis.get('LETTRAGE', ''),
                                'LIBELLE_OPR': row_bkhis.get('LIBELLE_OPR', ''),
                                'STATUT': 'NON RAPPROCHE BKHIS',
                                'ECART': None
                            }
                            results.append(merged_row)
                    
                    # Statistiques
                    total_rapproches = len([r for r in results if r['STATUT'] == 'RAPPROCHE'])
                    total_cpi_seul = len([r for r in results if r['STATUT'] == 'NON RAPPROCHE CPI'])
                    total_bkhis_seul = len([r for r in results if r['STATUT'] == 'NON RAPPROCHE BKHIS'])
                    
                    self.log_message(f"✅ FULL OUTER JOIN terminé:", 'SUCCESS')
                    self.log_message(f"  Total lignes: {len(results)}", 'INFO')
                    self.log_message(f"  Rapprochées: {total_rapproches}", 'INFO')
                    self.log_message(f"  CPI seulement: {total_cpi_seul}", 'INFO')
                    self.log_message(f"  BKHIS seulement: {total_bkhis_seul}", 'INFO')
                    
                    # Convertir en DataFrame Polars avec gestion des types
                    if results:
                        try:
                            result_df_pandas = pd.DataFrame(results)
                             
                            # Nettoyer les types de données problématiques
                            for col in result_df_pandas.columns:
                                if result_df_pandas[col].dtype == 'object':
                                    # Convertir les chaînes vides en None/NaN
                                    result_df_pandas[col] = result_df_pandas[col].replace('', None)
                                    
                                    # Pour les colonnes numériques, convertir en float avec gestion des erreurs
                                    if col in ['SOLDE CPI', 'MONT', 'ECART']:
                                        result_df_pandas[col] = pd.to_numeric(result_df_pandas[col], errors='coerce')
                             
                            # self.log_message(f"Conversion pandas→polars: {len(result_df_pandas)} lignes, {len(result_df_pandas.columns)} colonnes", 'DEBUG')
                            return pl.from_pandas(result_df_pandas)
                             
                        except Exception as e:
                            self.log_message(f"Erreur conversion DataFrame: {str(e)}", 'ERROR')
                            # En cas d'erreur, créer un DataFrame Polars manuellement
                            try:
                                # Créer un DataFrame Polars vide avec les bonnes colonnes
                                schema = {
                                    'N°': pl.Int64,
                                    'DATE REGLEMENT CPI': pl.Utf8,
                                    'SENS CPI': pl.Utf8,
                                    'SOLDE CPI': pl.Float64,
                                    'AGE': pl.Utf8,
                                    'NCP': pl.Utf8,
                                    'DCO': pl.Utf8,
                                    'OPE': pl.Utf8,
                                    'PIE': pl.Utf8,
                                    'SENS': pl.Utf8,
                                    'MONT': pl.Float64,
                                    'LETTRAGE': pl.Utf8,
                                    'LIBELLE_OPR': pl.Utf8,
                                    'STATUT': pl.Utf8,
                                    'ECART': pl.Float64
                                }
                                 
                                # Créer le DataFrame avec les données nettoyées
                                clean_results = []
                                for row in results:
                                    clean_row = {}
                                    for key, value in row.items():
                                        if key in ['SOLDE CPI', 'MONT', 'ECART']:
                                            clean_row[key] = float(value) if value not in ['', None] else 0.0
                                        elif key == 'N°':
                                            clean_row[key] = int(value) if value not in ['', None] else 0
                                        else:
                                            clean_row[key] = str(value) if value is not None else ''
                                    clean_results.append(clean_row)
                                 
                                return pl.DataFrame(clean_results, schema=schema)
                                 
                            except Exception as e2:
                                self.log_message(f"Erreur création manuelle: {str(e2)}", 'ERROR')
                                return pl.DataFrame()
                    else:
                        self.log_message("Aucun résultat généré", 'WARNING')
                        return pl.DataFrame()
                    
                except Exception as e:
                    self.log_message(f"Erreur dans FULL OUTER JOIN: {str(e)}", 'ERROR')
                    return pl.DataFrame()
            
            # Effectuer le FULL OUTER JOIN avec tolérance
            rapprochement_result = rapprocher_full_outer_join(cpi_data, bkhis_pour_rapprochement, tolerance=0.05)
            
            self.log_message(f"Rapprochement terminé: {rapprochement_result.height:,} lignes", 'SUCCESS')
            self.log_message(f"Colonnes après rapprochement: {rapprochement_result.columns}", 'INFO')
            # self.log_message(f"✅ FULL OUTER JOIN effectué: {rapprochement_result.height:,} lignes", 'SUCCESS') # Doublon supprimé
            self.log_message(f"📋 Colonnes après rapprochement: {rapprochement_result.columns}", 'INFO')
            
            # Vérifier si le résultat est vide
            if rapprochement_result.height == 0:
                self.log_message("⚠️ Aucune donnée de rapprochement trouvée", 'WARNING')
                QMessageBox.warning(self, "Attention", 
                    "Aucune donnée de rapprochement n'a été trouvée.\n\n"
                    "Vérifiez que les fichiers CPI et BKHIS contiennent des données compatibles.")
                return
            
            # 5️⃣ Statistiques du FULL OUTER JOIN
            # Compter par statut uniquement si la colonne STATUT existe
            rapprochees = 0
            cpi_seul = 0
            bkhis_seul = 0
            
            if 'STATUT' in rapprochement_result.columns:
                stats_par_statut = rapprochement_result.group_by('STATUT').count()
                
                for stat in stats_par_statut.to_pandas().itertuples():
                    if stat.STATUT == 'RAPPROCHE':
                        rapprochees = stat.count
                    elif stat.STATUT == 'NON RAPPROCHE CPI':
                        cpi_seul = stat.count
                    elif stat.STATUT == 'NON RAPPROCHE BKHIS':
                        bkhis_seul = stat.count
            else:
                self.log_message("⚠️ Colonne STATUT non trouvée, calcul de statistiques impossible", 'WARNING')
            
            #  TOTAL LIGNES CPI SEULEMENT (CPI rapprochées + CPI non rapprochées)
            total_lignes_cpi = rapprochees + cpi_seul
            
            # Calcul du taux de rapprochement basé sur le total CPI seulement
            taux_rapprochement = (rapprochees / total_lignes_cpi * 100) if total_lignes_cpi > 0 else 0
            
            self.log_message(f"📊 Statistiques du rapprochement:", 'INFO')
            self.log_message(f"  Total lignes CPI: {total_lignes_cpi:,}", 'INFO')
            self.log_message(f"  Rapprochées: {rapprochees:,} ({taux_rapprochement:.1f}%)", 'INFO')
            self.log_message(f"  CPI reste: {cpi_seul:,} ({cpi_seul/total_lignes_cpi*100:.1f}%)", 'INFO')
            self.log_message(f"  BKHIS reste: {bkhis_seul:,}", 'INFO')
            
            #  METTRE À JOUR LA JAUGE DE PROGRESSION
            self.update_progress_jauge(taux_rapprochement)
            
            # 6️⃣ Stocker les résultats pour l'export
            self.rapprochement_data = rapprochement_result
            
            # 7️⃣ Créer un résumé des rapprochements (non affiché dans le journal)
            if rapprochees > 0:
                # Utiliser les bonnes colonnes pour le résumé
                detail_cols = ['Montant CPI', 'NCP', 'SENS', 'DATE']
                available_cols = [col for col in detail_cols if col in rapprochement_result.columns]
                
                if available_cols:
                    rapprochement_detail = rapprochement_result.filter(
                        pl.col('NCP').is_not_null()
                    ).select(available_cols).head(10)
                    
                    # Les exemples de rapprochements ne sont plus affichés dans le journal
                    # self.log_message("📝 Exemples de rapprochements trouvés:", 'INFO')
                    # for idx, row in rapprochement_detail.to_pandas().iterrows():
                    #     montant = row['Montant CPI'] if 'Montant CPI' in row else 'N/A'
                    #     ncp = row['NCP'] if 'NCP' in row else 'N/A'
                    #     sens = f" ({row['SENS']})" if 'SENS' in row and row['SENS'] else ""
                    #     date = f" - {row['DATE']}" if 'DATE' in row and row['DATE'] else ""
                    #     
                    #     # Formatter le montant uniquement si c'est numérique
                    #     try:
                    #         montant_formate = f"{float(montant):,.2f}" if montant != 'N/A' else 'N/A'
                    #     except (ValueError, TypeError):
                    #         montant_formate = str(montant)
                    #     
                    #     self.log_message(f"  Montant: {montant_formate} → NCP: {ncp}{sens}{date}", 'INFO')
            
            self.log_message("PHASE 1 - Vérification de l'exhaustivité des journées CPI terminé avec succès !", 'SUCCESS')
            
            # Message de confirmation
            QMessageBox.information(self, "FULL OUTER JOIN effectué", 
                                  f"FULL OUTER JOIN BKHIS-CPI terminé!\n\n"
                                  f"Total lignes CPI: {total_lignes_cpi:,}\n"
                                  f"Rapprochées: {rapprochees:,} ({taux_rapprochement:.1f}%)\n"
                                  f"CPI reste: {cpi_seul:,}\n"
                                  f"BKHIS reste: {bkhis_seul:,}\n\n"
                                  f"Prêt pour l'export des résultats.")
            
        except Exception as e:
            self.log_message(f"❌ Erreur lors du rapprochement: {str(e)}", 'ERROR')
            QMessageBox.critical(self, "Erreur", f"Échec du rapprochement:\n{str(e)}")
    
    def diagnostiquer_montants_non_rapproches(self, cpi_data, bkhis_pour_join, rapprochement_result):
        """Diagnostic détaillé des montants non rapprochés spécifiques"""
        try:
            self.log_message("🔍 DIAGNOSTIC: Analyse des montants non rapprochés...", 'INFO')
            
            # Montants spécifiques à analyser
            montants_a_analyser = [
                491628135.88, 73549447.11, 89522782.53, 719664118.49,
                601864653.19, 630538705.36, 71928646.03, 99574250.12,
                612940498.06, 54329866.31, 791306206.17, 31600528.98,
                642366415.32, 26410115.79
            ]
            
            # Convertir en pandas pour faciliter l'analyse
            cpi_df = cpi_data.to_pandas()
            bkhis_df = bkhis_pour_join.to_pandas()
            rapprochement_df = rapprochement_result.to_pandas()
            
            self.log_message(f"📊 Base CPI: {len(cpi_df)} lignes", 'INFO')
            self.log_message(f"📊 Base BKHIS: {len(bkhis_df)} lignes", 'INFO')
            self.log_message(f"📊 Rapprochement: {len(rapprochement_df)} lignes", 'INFO')
            
            # Analyser chaque montant
            for montant in montants_a_analyser:
                self.log_message(f"🔍 Analyse du montant: {montant:,.2f}", 'INFO')
                
                # 🔍 DIAGNOSTIC PRÉCIS: Vérifier les types et valeurs exactes
                self.log_message(f"🔬 Type du montant recherché: {type(montant)} - valeur: {repr(montant)}", 'DEBUG')
                
                # Chercher dans CPI avec diagnostic détaillé
                self.log_message("🔍 Recherche CPI avec diagnostic...", 'DEBUG')
                cpi_match_exact = cpi_df[cpi_df['Montant CPI'] == montant]
                self.log_message(f"🔍 Résultat recherche exacte: {len(cpi_match_exact)} matchs", 'DEBUG')
                
                # Diagnostic des valeurs dans CPI
                if len(cpi_df) > 0:
                    sample_cpi = cpi_df['Montant CPI'].iloc[0]
                    self.log_message(f"🔬 Type Montant CPI: {type(sample_cpi)} - exemple: {repr(sample_cpi)}", 'DEBUG')
                
                # Chercher avec une approche différente
                cpi_match_close = cpi_df[abs(cpi_df['Montant CPI'] - montant) < 0.01]
                self.log_message(f"🔍 Résultat recherche proche (<0.01): {len(cpi_match_close)} matchs", 'DEBUG')
                
                if len(cpi_match_close) > 0:
                    for idx, row in cpi_match_close.iterrows():
                        val = row['Montant CPI']
                        diff = abs(val - montant)
                        self.log_message(f"🔬 CPI trouvé: {repr(val)} (type: {type(val)}, diff: {diff})", 'DEBUG')
                
                if len(cpi_match_exact) > 0:
                    self.log_message(f"  ✅ Trouvé dans CPI: {len(cpi_match_exact)} fois", 'INFO')
                    for idx, row in cpi_match_exact.iterrows():
                        self.log_message(f"    - Ligne {idx}: solde_total={row.get('solde_total', 'N/A')}", 'DEBUG')
                else:
                    self.log_message(f"  ❌ Non trouvé dans CPI", 'WARNING')
                    # Chercher des montants proches
                    cpi_proches = cpi_df[abs(cpi_df['Montant CPI'] - montant) < 1]
                    if len(cpi_proches) > 0:
                        self.log_message(f"  🔍 Montants proches dans CPI:", 'DEBUG')
                        for idx, row in cpi_proches.head(3).iterrows():
                            diff = abs(row['Montant CPI'] - montant)
                            try:
                                montant_formate = f"{float(row['Montant CPI']):,.2f}"
                            except (ValueError, TypeError):
                                montant_formate = str(row['Montant CPI'])
                            self.log_message(f"    - {montant_formate} (diff: {diff:.2f})", 'DEBUG')
                
                # Chercher dans BKHIS
                bkhis_match = bkhis_df[bkhis_df['MONT_BKHIS'] == montant]
                if len(bkhis_match) > 0:
                    self.log_message(f"  ✅ Trouvé dans BKHIS: {len(bkhis_match)} fois", 'INFO')
                    for idx, row in bkhis_match.iterrows():
                        self.log_message(f"    - Ligne {idx}: NCP={row.get('NCP', 'N/A')}, SENS={row.get('SENS', 'N/A')}", 'DEBUG')
                else:
                    self.log_message(f"  ❌ Non trouvé dans BKHIS", 'WARNING')
                    # Chercher des montants proches
                    bkhis_proches = bkhis_df[abs(bkhis_df['MONT_BKHIS'] - montant) < 1]
                    if len(bkhis_proches) > 0:
                        self.log_message(f"  🔍 Montants proches dans BKHIS:", 'DEBUG')
                        for idx, row in bkhis_proches.head(3).iterrows():
                            diff = abs(row['MONT_BKHIS'] - montant)
                            try:
                                montant_formate = f"{float(row['MONT_BKHIS']):,.2f}"
                            except (ValueError, TypeError):
                                montant_formate = str(row['MONT_BKHIS'])
                            self.log_message(f"    - {montant_formate} (diff: {diff:.2f})", 'DEBUG')
                
                # Vérifier le rapprochement
                rapprochement_match = rapprochement_df[
                    (rapprochement_df['Montant CPI'] == montant) & 
                    (rapprochement_df['NCP'].notna())
                ]
                if len(rapprochement_match) > 0:
                    self.log_message(f"  ✅ Rapproché: {len(rapprochement_match)} fois", 'SUCCESS')
                else:
                    self.log_message(f"  ❌ Non rapproché", 'ERROR')
                    # Vérifier s'il existe dans le résultat sans rapprochement
                    non_rapproche_match = rapprochement_df[
                        rapprochement_df['Montant CPI'] == montant
                    ]
                    if len(non_rapproche_match) > 0:
                        self.log_message(f"  📋 Présent mais non rapproché: {len(non_rapproche_match)} fois", 'WARNING')
                
                self.log_message("", 'INFO')  # Ligne vide pour la lisibilité
            
            # Analyse générale des différences
            self.log_message("📊 Analyse générale des différences...", 'INFO')
            
            # Montants CPI non rapprochés
            non_rapproches_cpi = rapprochement_df[
                rapprochement_df['NCP'].isna()
            ]['Montant CPI'].tolist()
            
            self.log_message(f"📋 Montants CPI non rapprochés: {len(non_rapproches_cpi)}", 'INFO')
            self.log_message(f"📋 10 premiers montants non rapprochés:", 'INFO')
            for i, montant in enumerate(non_rapproches_cpi[:10]):
                try:
                    montant_formate = f"{float(montant):,.2f}"
                except (ValueError, TypeError):
                    montant_formate = str(montant)
                self.log_message(f"  {i+1}. {montant_formate}", 'INFO')
            
            # Vérifier si les montants spécifiques sont dans les non rapprochés
            montants_non_rapproches_specifiques = [
                m for m in montants_a_analyser if m in non_rapproches_cpi
            ]
            
            if montants_non_rapproches_specifiques:
                self.log_message(f"❌ Montants spécifiques non rapprochés: {len(montants_non_rapproches_specifiques)}/14", 'ERROR')
                for montant in montants_non_rapproches_specifiques:
                    self.log_message(f"  - {montant:,.2f}", 'ERROR')
            else:
                self.log_message("✅ Tous les montants spécifiques sont rapprochés!", 'SUCCESS')
            
            self.log_message(" Diagnostic terminé!", 'SUCCESS')
            
        except Exception as e:
            self.log_message(f"❌ Erreur lors du diagnostic: {str(e)}", 'ERROR')
    
    def verifier_ventilation(self):
        """Ouvre une boîte de dialogue indépendante pour la ventilation"""
        self.log_message("Ouverture de la boîte de dialogue de ventilation...", 'INFO')
        
        # Créer une boîte de dialogue personnalisée pour la ventilation
        dialog = QDialog(self)
        dialog.setWindowTitle("Ventilation")
        dialog.setFixedSize(600, 500)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #010001;
                color: #e2e8f0;
            }
            QLabel {
                color: #e2e8f0;
                font-size: 14px;
                padding: 5px;
            }
            QGroupBox {
                color: #e2e8f0;
                border: 2px solid #1ecce8;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                background-color: #010001;
            }
            QGroupBox::title {
                color: #60a5fa;
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
            QLineEdit {
                background-color: #010001;
                color: #e2e8f0;
                border: 1px solid #1ecce8;
                border-radius: 4px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton {
                background-color: #1ecce8;
                color: #010001;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #1bb8d4;
            }
            QPushButton:pressed {
                background-color: #1498a8;
            }
            QPushButton:disabled {
                background-color: #6b7280;
                color: #9ca3af;
            }
        """)
        
        layout = QVBoxLayout(dialog)
        
        # Titre
        title_label = QLabel("Ventilation")
        title_label.setStyleSheet("font-size: 20px; font-weight: bold; color: #60a5fa; padding: 15px;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # Section des fichiers à charger
        files_group = QGroupBox("📁 Fichiers Excel à charger")
        files_layout = QVBoxLayout(files_group)
        
        # Fichier Récapitulatif statistique
        recap_layout = QHBoxLayout()
        recap_label = QLabel("Fichier Récapitulatif statistique:")
        recap_label.setStyleSheet("font-size: 13px; color: #e2e8f0; padding: 5px; width: 200px;")
        recap_layout.addWidget(recap_label)
        
        recap_file_path = QLineEdit()
        recap_file_path.setPlaceholderText("Sélectionner le fichier Excel contenant 'Sld.CPI.calc.instru'...")
        recap_layout.addWidget(recap_file_path)
        
        recap_browse_btn = QPushButton("Parcourir...")
        recap_browse_btn.clicked.connect(lambda: self._browse_file_ventilation(recap_file_path, "Récapitulatif statistique"))
        recap_layout.addWidget(recap_browse_btn)
        
        files_layout.addLayout(recap_layout)
        
        # Fichier Exhaustivité
        exhaust_layout = QHBoxLayout()
        exhaust_label = QLabel("Fichier Exhaustivité:")
        exhaust_label.setStyleSheet("font-size: 13px; color: #e2e8f0; padding: 5px; width: 200px;")
        exhaust_layout.addWidget(exhaust_label)
        
        exhaust_file_path = QLineEdit()
        exhaust_file_path.setPlaceholderText("Sélectionner le fichier Excel contenant 'Result Exhaustivité'...")
        exhaust_layout.addWidget(exhaust_file_path)
        
        exhaust_browse_btn = QPushButton("Parcourir...")
        exhaust_browse_btn.clicked.connect(lambda: self._browse_file_ventilation(exhaust_file_path, "Exhaustivité"))
        exhaust_layout.addWidget(exhaust_browse_btn)
        
        files_layout.addLayout(exhaust_layout)
        
        # Fichier BKHIS
        bkhis_layout = QHBoxLayout()
        bkhis_label = QLabel("Fichier BKHIS:")
        bkhis_label.setStyleSheet("font-size: 13px; color: #e2e8f0; padding: 5px; width: 200px;")
        bkhis_layout.addWidget(bkhis_label)
        
        bkhis_file_path = QLineEdit()
        bkhis_file_path.setPlaceholderText("Sélectionner le fichier Excel BKHIS...")
        bkhis_layout.addWidget(bkhis_file_path)
        
        bkhis_browse_btn = QPushButton("Parcourir...")
        bkhis_browse_btn.clicked.connect(lambda: self._browse_file_ventilation(bkhis_file_path, "BKHIS"))
        bkhis_layout.addWidget(bkhis_browse_btn)
        
        files_layout.addLayout(bkhis_layout)
        
        layout.addWidget(files_group)
        
        # Section des critères de jointure
        join_group = QGroupBox("🔗 Critères de jointure")
        join_layout = QVBoxLayout(join_group)
        
        join_info = QLabel(
            "• Feuille source: 'Sld.CPI.calc.instru' (Récapitulatif statistique)\n"
            "• Feuille cible: 'Result Exhaustivité' (Exhaustivité)\n"
            "• Critère: DATE REGLEMENT ↔ DATE REGLEMENT CPI\n"
            "• Type: FULL OUTER JOIN\n"
            "• Tolérance BKHIS: ±0.02 sur les montants"
        )
        join_info.setStyleSheet("font-size: 13px; color: #e2e8f0; padding: 10px;")
        join_layout.addWidget(join_info)
        
        layout.addWidget(join_group)
        
        # Section de destination
        output_group = QGroupBox("💾 Fichier de sortie")
        output_layout = QVBoxLayout(output_group)
        
        output_info = QLabel("Le résultat sera créé dans un nouvel onglet 'Ventilation' dans le fichier Excel d'exhaustivité")
        output_info.setStyleSheet("font-size: 13px; color: #1ecce8; padding: 10px;")
        output_layout.addWidget(output_info)
        
        layout.addWidget(output_group)
        
        # Boutons d'action
        buttons_layout = QHBoxLayout()
        
        # Bouton principal "Ident.Vent.Exhaustive"
        vent_button = QPushButton("Ident.Vent.Exhaustive")
        vent_button.setStyleSheet("""
            QPushButton {
                background-color: #f59e0b;
                color: #010001;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #d97706;
            }
            QPushButton:pressed {
                background-color: #b45309;
            }
        """)
        vent_button.clicked.connect(lambda: self._perform_ventilation_identification(dialog, recap_file_path.text(), exhaust_file_path.text(), bkhis_file_path.text()))
        
        cancel_button = QPushButton("Annuler")
        cancel_button.clicked.connect(dialog.reject)
        
        buttons_layout.addStretch()
        buttons_layout.addWidget(vent_button)
        buttons_layout.addWidget(cancel_button)
        
        layout.addLayout(buttons_layout)
        
        # Afficher la boîte de dialogue
        dialog.exec()
    
    def _browse_file_ventilation(self, line_edit, file_type):
        """Parcourir pour sélectionner un fichier Excel pour la ventilation"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            f"Sélectionner le fichier {file_type}",
            "",
            "Fichiers Excel (*.xlsx *.xls)"
        )
        if file_path:
            line_edit.setText(file_path)
            self.log_message(f"Fichier {file_type} sélectionné: {file_path}", 'INFO')
    
    def _perform_ventilation_identification(self, dialog, recap_file, exhaust_file, bkhis_file=None):
        try:
            # Vérifier que les fichiers sont sélectionnés
            recap_file = recap_file.strip()
            exhaust_file = exhaust_file.strip()
            
            if not recap_file:
                QMessageBox.warning(self, "Attention", "Veuillez sélectionner le fichier Récapitulatif statistique!")
                return
            
            if not exhaust_file:
                QMessageBox.warning(self, "Attention", "Veuillez sélectionner le fichier Exhaustivité!")
                return
            
            self.log_message("Début de l'identification Ventilation...", 'INFO')
            self.log_message(f"Fichier Récapitulatif: {recap_file}", 'INFO')
            self.log_message(f"Fichier Exhaustivité: {exhaust_file}", 'INFO')
            
            # Charger les données avec pandas (nécessaire pour openpyxl)
            import pandas as pd
            
            # Charger la feuille Sld.CPI.calc.instru depuis le fichier récapitulatif
            try:
                df_recap = pd.read_excel(recap_file, sheet_name="Sld.CPI.calc.instru")
                self.log_message(f"✅ Feuille 'Sld.CPI.calc.instru' chargée: {len(df_recap)} lignes", 'SUCCESS')
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Impossible de charger la feuille 'Sld.CPI.calc.instru':\n{str(e)}")
                return
            
            # Charger la feuille Result Exhaustivité depuis le fichier exhaustivité
            try:
                df_exhaust = pd.read_excel(exhaust_file, sheet_name="Result Exhaustivité")
                self.log_message(f"✅ Feuille 'Result Exhaustivité' chargée: {len(df_exhaust)} lignes", 'SUCCESS')
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Impossible de charger la feuille 'Result Exhaustivité':\n{str(e)}")
                return
            
            # Vérifier les colonnes nécessaires
            if "DATE REGLEMENT" not in df_recap.columns:
                QMessageBox.critical(self, "Erreur", "La colonne 'DATE REGLEMENT' n'existe pas dans la feuille 'Sld.CPI.calc.instru'")
                return
            
            if "DATE REGLEMENT CPI" not in df_exhaust.columns:
                QMessageBox.critical(self, "Erreur", "La colonne 'DATE REGLEMENT CPI' n'existe pas dans la feuille 'Result Exhaustivité'")
                return
            
            # Convertir les dates en string pour éviter les problèmes de format
            df_recap["DATE REGLEMENT"] = df_recap["DATE REGLEMENT"].astype(str)
            df_exhaust["DATE REGLEMENT CPI"] = df_exhaust["DATE REGLEMENT CPI"].astype(str)
            
            # Effectuer le LEFT JOIN
            self.log_message("🔄 Effectuation du LEFT JOIN...", 'INFO')
            df_ventilation = df_recap.merge(
                df_exhaust,
                left_on="DATE REGLEMENT",
                right_on="DATE REGLEMENT CPI",
                how="left",
                suffixes=("", "_exhaust")
            )
            
            self.log_message(f"✅ LEFT JOIN terminé: {len(df_ventilation)} lignes résultantes", 'SUCCESS')
            
            # Ajouter la colonne Observation
            self.log_message("📝 Ajout de la colonne 'Observation'...", 'INFO')
            
            # Créer la colonne Observation en fonction de SOLDE BKHIS
            df_ventilation['Observation'] = df_ventilation.apply(
                lambda row: 'Non rapproché' if pd.notna(row.get('SOLDE BKHIS')) and float(row.get('SOLDE BKHIS', 0)) == 0 else 'Rapproché', 
                axis=1
            )
            
            # Compter les statuts
            nb_rapprochees = len(df_ventilation[df_ventilation['Observation'] == 'Rapproché'])
            nb_non_rapprochees = len(df_ventilation[df_ventilation['Observation'] == 'Non rapproché'])
            
            self.log_message(f"📊 Statistiques d'observation: {nb_rapprochees} rapprochées, {nb_non_rapprochees} non rapprochées", 'INFO')
            
            # Ajouter la feuille Ventilation au fichier Excel d'exhaustivité
            try:
                from openpyxl import load_workbook
                
                # Charger le fichier Excel existant
                wb = load_workbook(exhaust_file)
                
                # Supprimer la feuille Ventilation si elle existe déjà
                if "Ventilation" in wb.sheetnames:
                    del wb["Ventilation"]
                    self.log_message("🗑️ Ancienne feuille 'Ventilation' supprimée", 'INFO')
                
                # Créer la nouvelle feuille Ventilation
                ws_ventilation = wb.create_sheet("Ventilation")
                
                # Écrire les en-têtes en premier
                for c_idx, column_name in enumerate(df_ventilation.columns, 1):
                    ws_ventilation.cell(row=1, column=c_idx, value=column_name)
                
                # Appliquer le style aux en-têtes
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for c_idx in range(1, len(df_ventilation.columns) + 1):
                    cell = ws_ventilation.cell(row=1, column=c_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Écrire les données à partir de la ligne 2
                for r_idx, row in enumerate(df_ventilation.itertuples(index=False), 2):  # Commence à 2
                    for c_idx, value in enumerate(row, 1):
                        ws_ventilation.cell(row=r_idx, column=c_idx, value=value)
                
                # Mettre la couleur de l'onglet en orange
                ws_ventilation.sheet_properties.tabColor = "FF6600"
                
                # Créer une copie de la feuille Ventilation avec transformations
                self.log_message("🔄 Création de la copie 'Ventilation Rappro BKHIS' avec transformations...", 'INFO')
                
                # Supprimer la feuille Ventilation Rappro BKHIS si elle existe déjà
                if "Ventilation Rappro BKHIS" in wb.sheetnames:
                    del wb["Ventilation Rappro BKHIS"]
                    self.log_message("🗑️ Ancienne feuille 'Ventilation Rappro BKHIS' supprimée", 'INFO')
                
                # Créer la nouvelle feuille Ventilation Rappro BKHIS
                ws_ventilation_transform = wb.create_sheet("Ventilation Rappro BKHIS")
                
                # Préparer les données transformées
                df_transformed = df_ventilation.copy()
                
                # Filtrer uniquement les lignes "Rapproché" pour la feuille Ventilation Rappro BKHIS
                if "Observation" in df_transformed.columns:
                    nb_total_avant_filtre = len(df_transformed)
                    df_transformed = df_transformed[df_transformed["Observation"] == "Rapproché"].copy()
                    nb_total_apres_filtre = len(df_transformed)
                    self.log_message(f"🔍 Filtre des lignes rapprochées: {nb_total_apres_filtre:,} lignes conservées sur {nb_total_avant_filtre:,} au total", 'INFO')
                else:
                    self.log_message("⚠️ Colonne 'Observation' non trouvée, aucune ligne filtrée", 'WARNING')
                
                # 1- Créer NCP SANS CLE (supprimer les 3 derniers caractères)
                if "NCP" in df_transformed.columns:
                    df_transformed["NCP SANS CLE"] = df_transformed["NCP"].astype(str).str[:-3]
                    self.log_message("✅ Colonne 'NCP SANS CLE' créée (3 derniers caractères supprimés)", 'SUCCESS')
                else:
                    self.log_message("⚠️ Colonne 'NCP' non trouvée, création de 'NCP SANS CLE' ignorée", 'WARNING')
                
                # 2- Créer MONTANT (valeur absolue de SOLDES)
                if "SOLDES" in df_transformed.columns:
                    # Convertir en numérique puis prendre la valeur absolue
                    df_transformed["MONTANT"] = pd.to_numeric(df_transformed["SOLDES"], errors='coerce').abs()
                    self.log_message("✅ Colonne 'MONTANT' créée (valeur absolue de SOLDES)", 'SUCCESS')
                else:
                    self.log_message("⚠️ Colonne 'SOLDES' non trouvée, création de 'MONTANT' ignorée", 'WARNING')
                
                # 3- Supprimer les colonnes inutiles
                colonnes_a_supprimer = ["N°", "DATE REGLEMENT CPI", "SOLDE CPI", "DCO", "SOLDE BKHIS", "ECARTS", "Observation"]
                colonnes_supprimees = []
                for colonne in colonnes_a_supprimer:
                    if colonne in df_transformed.columns:
                        df_transformed = df_transformed.drop(columns=[colonne])
                        colonnes_supprimees.append(colonne)
                
                if colonnes_supprimees:
                    self.log_message(f"🗑️ Colonnes supprimées: {', '.join(colonnes_supprimees)}", 'SUCCESS')
                else:
                    self.log_message("⚠️ Aucune des colonnes à supprimer n'a été trouvée", 'WARNING')
                
                # Réorganiser les colonnes pour mettre les nouvelles colonnes à côté de leurs originales
                new_columns_order = []
                for col in df_transformed.columns:
                    new_columns_order.append(col)
                    if col == "NCP" and "NCP SANS CLE" in df_transformed.columns:
                        new_columns_order.append("NCP SANS CLE")
                    elif col == "SOLDES" and "MONTANT" in df_transformed.columns:
                        new_columns_order.append("MONTANT")
                
                # Éliminer les doublons tout en préservant l'ordre
                seen = set()
                final_columns = []
                for col in new_columns_order:
                    if col not in seen:
                        final_columns.append(col)
                        seen.add(col)
                
                df_transformed = df_transformed[final_columns]
                
                # Écrire les en-têtes de la feuille transformée
                for c_idx, column_name in enumerate(df_transformed.columns, 1):
                    ws_ventilation_transform.cell(row=1, column=c_idx, value=column_name)
                
                # Appliquer le style aux en-têtes de la feuille transformée
                header_font_transform = Font(bold=True, color="FFFFFF")
                header_fill_transform = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")  # Vert
                header_alignment_transform = Alignment(horizontal="center", vertical="center")
                
                for c_idx in range(1, len(df_transformed.columns) + 1):
                    cell = ws_ventilation_transform.cell(row=1, column=c_idx)
                    cell.font = header_font_transform
                    cell.fill = header_fill_transform
                    cell.alignment = header_alignment_transform
                
                # Écrire les données transformées à partir de la ligne 2
                for r_idx, row in enumerate(df_transformed.itertuples(index=False), 2):
                    for c_idx, value in enumerate(row, 1):
                        ws_ventilation_transform.cell(row=r_idx, column=c_idx, value=value)
                
                # Mettre la couleur de l'onglet en vert pour la feuille transformée
                ws_ventilation_transform.sheet_properties.tabColor = "10b981"
                
                # Ajuster la largeur des colonnes pour les deux feuilles
                for ws in [ws_ventilation, ws_ventilation_transform]:
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column].width = adjusted_width
                
                # Sauvegarder le fichier
                wb.save(exhaust_file)
                
                # Stocker le chemin du fichier pour la jauge de ventilation
                self.dernier_fichier_ventilation = exhaust_file
                
                self.log_message(f" Feuille 'Ventilation' créée avec succès dans: {exhaust_file}", 'SUCCESS')
                self.log_message(f"📊 Nombre de lignes dans la ventilation: {len(df_ventilation)}", 'INFO')
                self.log_message(f"🔄 Feuille 'Ventilation Rappro BKHIS' créée avec transformations", 'SUCCESS')
                if "Observation" in df_ventilation.columns:
                    self.log_message(f"🔍 Feuille filtrée: uniquement les lignes rapprochées ({len(df_transformed):,} lignes)", 'INFO')
                if colonnes_supprimees:
                    self.log_message(f"🗑️ Colonnes supprimées: {len(colonnes_supprimees)} colonnes inutiles", 'INFO')
                
                # Afficher les statistiques basées sur la colonne Observation
                self.log_message("📈 Statistiques de ventilation:", 'INFO')
                self.log_message(f"  ✅ Rapprochées: {nb_rapprochees}", 'INFO')
                self.log_message(f"  ❌ Non rapprochées: {nb_non_rapprochees}", 'INFO')
                
                # Stocker les statistiques pour le message de succès
                self.stats_ventilation = {
                    'nb_rapprochees': nb_rapprochees,
                    'nb_non_rapprochees': nb_non_rapprochees,
                    'len_df_ventilation': len(df_ventilation),
                    'len_df_transformed': len(df_transformed)
                }
                
                # 🌪️ METTRE À JOUR LA JAUGE DE VENTILATION
                self.update_ventilation_jauge()
                
                # Étape 2: Rapprochement BKHIS si le fichier est fourni
                if bkhis_file and bkhis_file.strip():
                    self._perform_bkhis_rapprochement(wb, df_transformed, bkhis_file.strip(), exhaust_file)
                
                # Préparer le message de succès avec les statistiques appropriées
                if hasattr(self, 'stats_bkhis_rapprochement'):
                    # Utiliser les statistiques BKHIS si disponibles
                    stats = self.stats_bkhis_rapprochement
                    message = (
                        f"Identification Ventilation Exhaustive effectuée avec succès!\n\n"
                        f"Fichier: {exhaust_file}\n\n"
                        f"📊 STATISTIQUES DE RAPPROCHEMENT BKHIS\n"
                        f"  • Rapprochées: {stats['nb_rapprochees']:,}\n"
                        f"  • Non rapprochées (CPI seule): {stats['nb_vent_seul']:,}\n"
                        f"  • Total CPI Ventilé: {stats['total_cpi_ventile']:,}\n"
                        f"  • Taux de ventilation correcte: {stats['taux_ventilation_correcte']:.2f}%\n"
                        f"  • Non rapprochées (BKHIS seule): {stats['nb_bkhis_seul']:,}\n\n"
                        f"📈 Feuille 'Statistique Rapprochement' créée avec résumé détaillé"
                    )
                else:
                    # Utiliser les statistiques de ventilation de base
                    stats = self.stats_ventilation
                    message = (
                        f"Identification Ventilation Exhaustive effectuée avec succès!\n\n"
                        f"Fichier: {exhaust_file}\n\n"
                        f"📊 STATISTIQUES DE VENTILATION\n"
                        f"  • Rapprochées: {stats['nb_rapprochees']:,}\n"
                        f"  • Non rapprochées: {stats['nb_non_rapprochees']:,}\n\n"
                        f"📈 Feuilles créées:\n"
                        f"  • Ventilation (orange) - {stats['len_df_ventilation']:,} lignes totales\n"
                        f"  • Ventilation Rappro BKHIS (vert) - {stats['len_df_transformed']:,} lignes rapprochées uniquement"
                    )
                
                QMessageBox.information(self, " Succès - Ventilation Terminée", message)
                
                dialog.accept()
                
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Impossible de créer la feuille 'Ventilation':\n{str(e)}")
                return
            
        except Exception as e:
            self.log_message(f"❌ Erreur lors de la ventilation: {str(e)}", 'ERROR')
            QMessageBox.critical(self, "Erreur", f"Échec de la ventilation:\n{str(e)}")

    def _perform_bkhis_rapprochement(self, wb, df_ventilation_rappro, bkhis_file, exhaust_file):
        """Effectue le rapprochement entre Ventilation Rappro BKHIS et le fichier BKHIS"""
        try:
            self.log_message("🔄 Début du rapprochement BKHIS...", 'INFO')
            self.log_message(f"Fichier BKHIS: {bkhis_file}", 'INFO')
            
            # Charger le fichier BKHIS
            import pandas as pd
            df_bkhis = pd.read_excel(bkhis_file)
            self.log_message(f"✅ Fichier BKHIS chargé: {len(df_bkhis)} lignes", 'SUCCESS')
            
            # Vérifier les colonnes requises
            colonnes_requises_bkhis = ['MONT', 'SENS BKHIS']
            colonnes_manquantes = [col for col in colonnes_requises_bkhis if col not in df_bkhis.columns]
            if colonnes_manquantes:
                QMessageBox.critical(self, "Erreur", f"Colonnes manquantes dans le fichier BKHIS: {', '.join(colonnes_manquantes)}")
                return
            
            colonnes_requises_vent = ['MONTANT', 'SENS']
            colonnes_manquantes = [col for col in colonnes_requises_vent if col not in df_ventilation_rappro.columns]
            if colonnes_manquantes:
                QMessageBox.critical(self, "Erreur", f"Colonnes manquantes dans Ventilation Rappro BKHIS: {', '.join(colonnes_manquantes)}")
                return
            
            # Détecter les doublons dans BKHIS (par montant ET sens)
            self.log_message("🔍 Détection des doublons dans BKHIS...", 'INFO')
            doublons_bkhis = df_bkhis[df_bkhis.duplicated(subset=['MONT', 'SENS BKHIS'], keep='first')]
            montants_sens_doublons = doublons_bkhis[['MONT', 'SENS BKHIS']].drop_duplicates()
            
            # Ajouter une colonne DOUBLON_BKHIS dans le dataframe BKHIS
            # Marquer tous les doublons sauf la première occurrence
            df_bkhis['DOUBLON_BKHIS'] = False
            for idx, row in montants_sens_doublons.iterrows():
                mask = (df_bkhis['MONT'] == row['MONT']) & (df_bkhis['SENS BKHIS'] == row['SENS BKHIS'])
                # Marquer comme doublon toutes les occurrences sauf la première
                doublon_indices = df_bkhis[mask].index.tolist()
                if len(doublon_indices) > 1:
                    df_bkhis.loc[doublon_indices[1:], 'DOUBLON_BKHIS'] = True
            
            nb_doublons = len(df_bkhis[df_bkhis['DOUBLON_BKHIS']])
            nb_paires_doublons = len(montants_sens_doublons)
            self.log_message(f"🔍 Doublons BKHIS détectés: {nb_doublons} lignes sur {len(df_bkhis)} ({nb_paires_doublons} paires montant+sens doublonnées)", 'INFO')
            
            # Préparer les données pour le rapprochement
            # S'assurer que les types sont compatibles
            df_ventilation_rappro['MONTANT'] = pd.to_numeric(df_ventilation_rappro['MONTANT'], errors='coerce')
            df_bkhis['MONT'] = pd.to_numeric(df_bkhis['MONT'], errors='coerce')
            
            # Séparer les doublons BKHIS pour les traiter séparément
            df_bkhis_unique = df_bkhis[~df_bkhis['DOUBLON_BKHIS']].copy()
            df_bkhis_doublons = df_bkhis[df_bkhis['DOUBLON_BKHIS']].copy()
            
            # Effectuer le rapprochement avec tolérance de ±0,02 sur les montants
            self.log_message("🔄 Effectuation du rapprochement avec tolérance ±0.02...", 'INFO')
            tolerance = 0.02
            
            def rapprocher_avec_tolerance(df_vent, df_bkhis, tol=0.02):
                """Effectue un rapprochement avec tolérance sur les montants et correspondance des sens"""
                results = []
                matched_vent_indices = set()
                matched_bkhis_indices = set()
                
                # Convertir en listes pour itération
                vent_rows = df_vent.to_dict('records')
                bkhis_rows = df_bkhis.to_dict('records')
                
                self.log_message(f"📊 Début rapprochement: {len(vent_rows)} ventilation, {len(bkhis_rows)} BKHIS", 'INFO')
                
                # Étape 1: Trouver les correspondances avec tolérance
                for i, vent_row in enumerate(vent_rows):
                    if i in matched_vent_indices:
                        continue
                    
                    montant_vent = float(vent_row.get('MONTANT', 0))
                    sens_vent = str(vent_row.get('SENS', '')).strip()
                    
                    best_match = None
                    best_diff = float('inf')
                    best_bkhis_idx = None
                    
                    for j, bkhis_row in enumerate(bkhis_rows):
                        if j in matched_bkhis_indices:
                            continue
                        
                        montant_bkhis = float(bkhis_row.get('MONT', 0))
                        sens_bkhis = str(bkhis_row.get('SENS BKHIS', '')).strip()
                        
                        # Vérifier la compatibilité des sens
                        if sens_vent != sens_bkhis:
                            continue
                        
                        diff = abs(montant_vent - montant_bkhis)
                        
                        if diff <= tol and diff < best_diff:
                            best_diff = diff
                            best_match = bkhis_row
                            best_bkhis_idx = j
                    
                    if best_match is not None:
                        # Créer la ligne fusionnée
                        merged_row = vent_row.copy()
                        # Ajouter les colonnes BKHIS avec suffixe _BKHIS
                        for key, value in best_match.items():
                            if key != 'SENS':  # Éviter de dupliquer SENS
                                merged_row[f"{key}_BKHIS"] = value
                            else:
                                merged_row[f"{key}_BKHIS"] = value  # Garder pour référence
                        
                        # Ajouter l'écart de tolérance
                        merged_row['ECART_TOLERANCE'] = best_diff
                        merged_row['_merge'] = 'both'
                        results.append(merged_row)
                        matched_vent_indices.add(i)
                        matched_bkhis_indices.add(best_bkhis_idx)
                
                # Étape 2: Ajouter les lignes ventilation non rapprochées
                for i, vent_row in enumerate(vent_rows):
                    if i not in matched_vent_indices:
                        vent_row['_merge'] = 'left_only'
                        vent_row['ECART_TOLERANCE'] = None
                        results.append(vent_row)
                
                # Étape 3: Ajouter les lignes BKHIS non rapprochées
                for j, bkhis_row in enumerate(bkhis_rows):
                    if j not in matched_bkhis_indices:
                        # Créer une ligne avec les colonnes ventilation vides
                        empty_vent_row = {col: None for col in df_vent.columns}
                        # Ajouter les colonnes BKHIS
                        for key, value in bkhis_row.items():
                            empty_vent_row[f"{key}_BKHIS"] = value
                        empty_vent_row['_merge'] = 'right_only'
                        empty_vent_row['ECART_TOLERANCE'] = None
                        results.append(empty_vent_row)
                
                # Statistiques
                nb_rapprochees = len([r for r in results if r.get('_merge') == 'both'])
                nb_vent_seul = len([r for r in results if r.get('_merge') == 'left_only'])
                nb_bkhis_seul = len([r for r in results if r.get('_merge') == 'right_only'])
                
                # Statistiques sur les écarts de tolérance
                ecarts = [r.get('ECART_TOLERANCE') for r in results if r.get('ECART_TOLERANCE') is not None]
                if ecarts:
                    ecart_moyen = sum(ecarts) / len(ecarts)
                    ecart_max = max(ecarts)
                    ecart_min = min(ecarts)
                    nb_avec_tolerance = len([e for e in ecarts if e > 0])
                    nb_exact = len([e for e in ecarts if e == 0])
                    
                    self.log_message(f"📊 Statistiques détaillées des écarts:", 'INFO')
                    self.log_message(f"  Écart moyen: {ecart_moyen:.4f}", 'INFO')
                    self.log_message(f"  Écart min: {ecart_min:.4f} | Écart max: {ecart_max:.4f}", 'INFO')
                    self.log_message(f"  Correspondances exactes: {nb_exact} | Avec tolérance: {nb_avec_tolerance}", 'INFO')
                
                self.log_message(f"✅ Rapprochement avec tolérance terminé:", 'SUCCESS')
                self.log_message(f"  Rapprochées: {nb_rapprochees}", 'INFO')
                self.log_message(f"  Ventilation seule: {nb_vent_seul}", 'INFO')
                self.log_message(f"  BKHIS seul: {nb_bkhis_seul}", 'INFO')
                
                return pd.DataFrame(results)
            
            # Utiliser la fonction de rapprochement avec tolérance
            df_rappro_final = rapprocher_avec_tolerance(df_ventilation_rappro, df_bkhis_unique, tolerance)
            
            # Créer la colonne OBS
            def create_observation(row):
                if row['_merge'] == 'both':
                    return 'Rapproché'
                elif row['_merge'] == 'left_only':
                    return 'Non rapproché (Ventilation seule)'
                else:
                    return 'Non rapproché (BKHIS seul)'
            
            df_rappro_final['OBS'] = df_rappro_final.apply(create_observation, axis=1)
            
            # Supprimer la colonne _merge
            df_rappro_final = df_rappro_final.drop(columns=['_merge'])
            
            # Ajouter les doublons BKHIS comme lignes non rapprochées
            if len(df_bkhis_doublons) > 0:
                self.log_message(f"🔄 Ajout de {len(df_bkhis_doublons)} doublons BKHIS comme non rapprochés...", 'INFO')
                
                # Créer un dataframe pour les doublons avec les mêmes colonnes que df_rappro_final
                df_doublons_final = pd.DataFrame()
                
                # Ajouter les colonnes de ventilation (vides)
                for col in df_ventilation_rappro.columns:
                    df_doublons_final[col] = None
                
                # Ajouter les colonnes BKHIS
                for col in df_bkhis.columns:
                    if col in df_rappro_final.columns:
                        df_doublons_final[col] = df_bkhis_doublons[col].values
                    else:
                        df_doublons_final[col] = df_bkhis_doublons[col].values
                
                # Ajouter les colonnes supplémentaires
                df_doublons_final['OBS'] = 'Non rapproché (Doublon BKHIS)'
                
                # Concaténer avec le résultat principal
                df_rappro_final = pd.concat([df_rappro_final, df_doublons_final], ignore_index=True)
            
            # Statistiques du rapprochement
            nb_rapprochees_bkhis = len(df_rappro_final[df_rappro_final['OBS'] == 'Rapproché'])
            nb_non_rapprochees_vent = len(df_rappro_final[df_rappro_final['OBS'] == 'Non rapproché (Ventilation seule)'])
            nb_non_rapprochees_bkhis = len(df_rappro_final[df_rappro_final['OBS'] == 'Non rapproché (BKHIS seul)'])
            
            # Calculer le taux de ventilation correcte pour la jauge
            total_cpi_ventile = nb_rapprochees_bkhis + nb_non_rapprochees_vent
            taux_ventilation_correcte = (nb_rapprochees_bkhis / total_cpi_ventile * 100) if total_cpi_ventile > 0 else 0
            
            # Stocker le taux pour la jauge et les statistiques pour le message
            self.taux_ventilation_actuel = taux_ventilation_correcte
            self.stats_bkhis_rapprochement = {
                'nb_rapprochees': nb_rapprochees_bkhis,
                'nb_vent_seul': nb_non_rapprochees_vent,
                'nb_bkhis_seul': nb_non_rapprochees_bkhis,
                'total_cpi_ventile': total_cpi_ventile,
                'taux_ventilation_correcte': taux_ventilation_correcte
            }
            
            self.log_message(f"📊 Statistiques du rapprochement BKHIS:", 'INFO')
            self.log_message(f"  ✅ Rapprochées: {nb_rapprochees_bkhis}", 'INFO')
            self.log_message(f"  ❌ Non rapprochées (Ventilation): {nb_non_rapprochees_vent}", 'INFO')
            self.log_message(f"  ❌ Non rapprochées (BKHIS): {nb_non_rapprochees_bkhis}", 'INFO')
            
            # Créer la nouvelle feuille de rapprochement
            if "Rapprochement Final" in wb.sheetnames:
                del wb["Rapprochement Final"]
                self.log_message("🗑️ Ancienne feuille 'Rapprochement Final' supprimée", 'INFO')
            
            ws_rappro_final = wb.create_sheet("Rapprochement Final")
            
            # Écrire les en-têtes
            for c_idx, column_name in enumerate(df_rappro_final.columns, 1):
                ws_rappro_final.cell(row=1, column=c_idx, value=column_name)
            
            # Appliquer le style aux en-têtes
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="8b5cf6", end_color="8b5cf6", fill_type="solid")  # Violet
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for c_idx in range(1, len(df_rappro_final.columns) + 1):
                cell = ws_rappro_final.cell(row=1, column=c_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Écrire les données
            for r_idx, row in enumerate(df_rappro_final.itertuples(index=False), 2):
                for c_idx, value in enumerate(row, 1):
                    ws_rappro_final.cell(row=r_idx, column=c_idx, value=value)
            
            # Mettre la couleur de l'onglet en violet
            ws_rappro_final.sheet_properties.tabColor = "8b5cf6"
            
            # Ajuster la largeur des colonnes
            for col in ws_rappro_final.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_rappro_final.column_dimensions[column].width = adjusted_width
            
            # Créer la feuille statistique
            self._create_feuille_statistique_rapprochement(wb, df_rappro_final, nb_rapprochees_bkhis, nb_non_rapprochees_vent, nb_non_rapprochees_bkhis)
            
            # Sauvegarder le fichier
            wb.save(exhaust_file)
            
            self.log_message(f" Feuille 'Rapprochement Final' créée avec succès: {len(df_rappro_final)} lignes", 'SUCCESS')
            self.log_message(f"📊 Répartition: {nb_rapprochees_bkhis} rapprochées, {nb_non_rapprochees_vent + nb_non_rapprochees_bkhis} non rapprochées", 'INFO')
            self.log_message("📈 Feuille 'Statistique Rapprochement' créée avec résumé détaillé", 'SUCCESS')
            
            # 🌪️ METTRE À JOUR LA JAUGE DE VENTILATION (après rapprochement final)
            self.update_ventilation_jauge()
            
        except Exception as e:
            self.log_message(f"❌ Erreur lors du rapprochement BKHIS: {str(e)}", 'ERROR')
            QMessageBox.critical(self, "Erreur", f"Échec du rapprochement BKHIS:\n{str(e)}")
    
    def _create_feuille_statistique_rapprochement(self, wb, df_rappro_final, nb_rapprochees, nb_vent_seul, nb_bkhis_seul):
        """Crée une feuille statistique résumant les résultats du rapprochement BKHIS"""
        try:
            self.log_message("📊 Création de la feuille statistique...", 'INFO')
            
            # Importer les classes nécessaires pour le style
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Supprimer la feuille si elle existe déjà
            if "Statistique Rapprochement" in wb.sheetnames:
                del wb["Statistique Rapprochement"]
                self.log_message("🗑️ Ancienne feuille 'Statistique Rapprochement' supprimée", 'INFO')
            
            # Créer la nouvelle feuille
            ws_stat = wb.create_sheet("Statistique Rapprochement")
            
            # Statistiques de base
            total_cpi_ventile = nb_rapprochees + nb_vent_seul
            taux_ventilation_correcte = (nb_rapprochees / total_cpi_ventile * 100) if total_cpi_ventile > 0 else 0
            
            # Créer le tableau de statistiques
            statistiques = [
                ["STATISTIQUES DE RAPPROCHEMENT BKHIS", ""],
                ["", ""],
                ["RÉSULTATS DU RAPPROCHEMENT", ""],
                ["Rapprochées:", nb_rapprochees],
                ["Non rapprochées (CPI seule):", nb_vent_seul],
                ["Total CPI Ventilé:", total_cpi_ventile],
                ["Taux de ventilation correcte:", f"{taux_ventilation_correcte:.2f}%"],
                ["Non rapprochées (BKHIS seule):", nb_bkhis_seul],
            ]
            
            # Écrire les données dans la feuille
            for row_idx, row_data in enumerate(statistiques, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_stat.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Style pour les titres
                    if row_idx == 1 or (row_idx > 1 and statistiques[row_idx-2][0].endswith(":")):
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="1ecce8", end_color="1ecce8", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    # Style pour les sous-titres
                    elif row_idx > 1 and statistiques[row_idx-2][0] == "" and statistiques[row_idx-2][1] == "":
                        cell.font = Font(bold=True, color="60a5fa")
                        cell.fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    # Style pour les données
                    else:
                        cell.font = Font(color="e2e8f0")
                        if col_idx == 1:  # Première colonne (libellés)
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        else:  # Deuxième colonne (valeurs)
                            cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # Fusionner les cellules pour les titres
            ws_stat.merge_cells('A1:B1')
            ws_stat.merge_cells('A3:B3')
            
            # Fusionner les sous-titres
            for row_idx in range(2, len(statistiques)+1):
                if row_idx <= len(statistiques) and statistiques[row_idx-2][0] == "" and statistiques[row_idx-2][1] == "":
                    ws_stat.merge_cells(f'A{row_idx}:B{row_idx}')
            
            # Ajuster la largeur des colonnes
            ws_stat.column_dimensions['A'].width = 40
            ws_stat.column_dimensions['B'].width = 20
            
            # Mettre la couleur de l'onglet
            ws_stat.sheet_properties.tabColor = "f59e0b"  # Orange
            
            self.log_message("✅ Feuille 'Statistique Rapprochement' créée avec succès", 'SUCCESS')
            self.log_message(f"📊 {nb_rapprochees} rapprochées, {nb_vent_seul} CPI seule, {nb_bkhis_seul} BKHIS seule", 'INFO')
            
        except Exception as e:
            self.log_message(f"❌ Erreur lors de la création de la feuille statistique: {str(e)}", 'ERROR')

def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    # Créer la fenêtre de login
    login_window = LoginWindow()
    
    # Variable pour stocker l'application principale
    main_window = None
    
    def create_and_show_main_window():
        nonlocal main_window
        # Fermer la fenêtre de login
        login_window.close()
        # Créer l'application principale uniquement après login réussi
        main_window = CPIAnalyzerModern()
        main_window.show()
    
    # Connecter le signal de succès du login à la création et affichage de l'application principale
    login_window.login_success.connect(create_and_show_main_window)
    
    # Afficher la fenêtre de login
    login_window.show()
    
    sys.exit(app.exec())

if __name__ == '__main__':
    main()
